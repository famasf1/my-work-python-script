#import

###############################
# WHAT DAY IS TODAY?
mydayis = 0 #0 = today or 1 = yesterday or 2 = yesterday of yesterday
###############################

import pandas as pd
from tkinter import *
from tkinter import filedialog
import datetime
import openpyxl as pyxl
import dagdshort

today = datetime.datetime.today().strftime("%d-%m-%y")
yesterday = datetime.datetime.today() - datetime.timedelta(days=1)
yesterday = yesterday.strftime("%d-%m-%y")
dayafteryesterday = datetime.datetime.today() - datetime.timedelta(days=2)
dayafteryesterday = dayafteryesterday.strftime("%d-%m-%y")

def read():

    '''
    Load Excel with Pandas, find and replace PHYIDINSURE | PHYID with empty whitespace. Then split out the data into array with '-'
    once that's done, combine frame and send all the data to 'whatday' function to process into Excel file. 
    '''

    #read
    global sheet
    
    root = Tk()
    root.excel = filedialog.askopenfilename(title='Open DHL file',filetypes=[('Excel Files', '*.xls'), ('All Files' , '*.*')]) #Load Excel
    sheet = pd.read_excel(root.excel)
    remove_word = sheet['CCN'].replace(['PHYIDINSURE','PHYID'],'', regex=True).str.split('-')
    rw_df = pd.DataFrame(remove_word)
    rw_df2 = rw_df[['ID','Branch','Box Num']] = pd.DataFrame(rw_df.CCN.to_list(), index=rw_df.index)
    rw_df = rw_df.drop(['CCN'], axis=1)

    #combine frame
    frames = [sheet, rw_df]
    sheet = pd.concat(frames, axis=1)
    insure_only()
    FCB_only()
    FCB_only_out()
    whatday(mydayis)



def insure_only():
    '''
    After read() function finished. Take data from loaded Excel sheet and flitered out only value that match 'PHYIDINSURE', '33-1' using regex.
    then replace all the 'PHYIDINSURE' into blank, split '-' and return as Pandas Dataframe before sent it to whatday() function to process.
    '''

    #filter only insure from 33 to customer 
    global sheet_insure_only
    CCN_Insure = sheet[sheet['CCN'].str.contains(r"(^PHYIDINSURE.*)-(33-1)$",regex=True)]
    remove_word = CCN_Insure["CCN"].replace(['PHYIDINSURE'],'', regex=True).str.split('-')

    df_remove = pd.DataFrame(remove_word)
    #try:
    #    df_remove_as_table = df_remove[['ID','Branch','Box Num']] = pd.DataFrame(df_remove.CCN.to_list(), index=df_remove.index)
    #except:
    #    pass
    frames = [CCN_Insure, df_remove]
    sheet_insure_only = pd.concat(frames, axis=1)
    

def FCB_only():

    '''
    Separate only FCB Branch
    '''
    #filter only FCB branch
    global sheet_fcb_only

    only_fcb = sheet[sheet["POD"].str.contains("^FCB", regex=True)]
    only_fcb_phyid = only_fcb["CCN"].replace(['PHYID'],'',regex=True).str.split('-')

    only_fcb_out = sheet[sheet["Branch"].str.contains("(^2)...", regex=True, na=False)]
    only_fcb_out_phyid = only_fcb_out["CCN"].replace(['PHYID'],'',regex=True).str.split('-')

    df_only_fcb = pd.DataFrame(only_fcb_phyid)
    df_only_fcb_out = pd.DataFrame(only_fcb_out_phyid)

    #try:
    #    df_only_fcb_as_table = df_only_fcb[['ID','Branch','Box Num']] = pd.DataFrame(df_only_fcb.CCN.to_list(), index=df_only_fcb.index)
    #except:
    #    pass

    frames = [only_fcb, df_only_fcb, df_only_fcb_out]
    sheet_fcb_only = pd.concat(frames, axis=1)

def FCB_only_out():

    '''
    Separate only FCB Branch
    '''
    #filter only FCB branch
    global sheet_fcb_out_only

    only_fcb_out = sheet[sheet["Branch"].str.contains("(^2)...", regex=True, na=False)]
    only_fcb_out_phyid = only_fcb_out["CCN"].replace(['PHYID'],'',regex=True).str.split('-')

    df_only_fcb_out = pd.DataFrame(only_fcb_out_phyid)

    #try:
    #    df_only_fcb_as_table = df_only_fcb[['ID','Branch','Box Num']] = pd.DataFrame(df_only_fcb.CCN.to_list(), index=df_only_fcb.index)
    #except:
    #    pass

    frames = [only_fcb_out, df_only_fcb_out]
    sheet_fcb_out_only = pd.concat(frames, axis=1)


def whatday(whatday):

    '''
    Take data from 'read' function. Check whatday variable setting before export Pandas dataframe into Excel.
    '''

    global name
    if whatday == 0: #today
        name = f'DHL {today}'
    elif whatday == 1: #yesterday
        name = f'DHL {yesterday}'
    elif whatday == 2:
        name = f'DHL {dayafteryesterday}'
    print(whatday)
    sheet.to_excel(fr'C:\Users\Comseven\Documents\DHL\Completed\{name}.xlsx',index=False)
    sheet_insure_only.to_excel(fr'C:\Users\Comseven\Documents\DHL\Completed\{name}_INSURE_ONLY.xlsx',index=False)
    sheet_fcb_only.to_excel(fr'C:\Users\Comseven\Documents\DHL\Completed\{name}_FCB_ONLY.xlsx',index=False)
    with pd.ExcelWriter(fr'C:\Users\Comseven\Documents\DHL\Completed\{name}_FCB_ONLY.xlsx', engine="openpyxl", mode="a", if_sheet_exists="overlay") as file:
        sheet_fcb_out_only.to_excel(file, index=False, header=None, startrow=file.sheets['Sheet1'].max_row)

    bitly_api_activate()

def bitly_api_activate():

    '''
    Get data from selected row where URL link is located. Then convert them into shortlinks.
    '''

    bitly_connect = bitlyshortener.Shortener(tokens=list_token, max_cache_size=256)

    load_insure_wb = pyxl.load_workbook(fr'C:\Users\Comseven\Documents\DHL\Completed\{name}_INSURE_ONLY.xlsx', data_only=True)
    load_fcb_only = pyxl.load_workbook(fr'C:\Users\Comseven\Documents\DHL\Completed\{name}_FCB_ONLY.xlsx', data_only=True)
    load_insure_wb_sh = load_insure_wb.sheetnames
    load_fcb_wb_sheet = load_fcb_only.sheetnames
    active_Sheet = load_insure_wb[load_insure_wb_sh[0]]
    active_fcb_sheet = load_fcb_only[load_fcb_wb_sheet[0]]

    #for customer insure sheet
    def customer_link():
        for row in range(2, active_Sheet.max_row+1):

            long_link_list = []
            long_link = active_Sheet.cell(row=row, column=11).value
            long_link_list.append(long_link)
        
            short_link = bitly_connect.shorten_urls(long_link_list)
            active_Sheet.cell(row=row, column=15).value = str(short_link[0])
            load_insure_wb.save(fr'C:\Users\Comseven\Documents\DHL\Completed\{name}_INSURE_ONLY_1.xlsx')

    #for fcb sheet
    def fcb_link():
        for row in range(2,active_fcb_sheet.max_row+1):
            
            long_link_list = []
            long_link = active_fcb_sheet.cell(row=row, column=11).value
            long_link_list.append(long_link)
        
            short_link = bitly_connect.shorten_urls(long_link_list)
            active_fcb_sheet.cell(row=row, column=15).value = str(short_link[0])
            load_fcb_only.save(fr'C:\Users\Comseven\Documents\DHL\Completed\{name}_FCB_ONLY_1.xlsx')

    customer_link()
    fcb_link()



##### LEGACY ####

def bitly_api_activate_once():

    '''
    Get data from selected row where URL link is located. Then convert them into shortlinks.
    ***THIS IS MEANT TO BE USE ONCE***
    '''

    dagd_connect = dagdshort.Shortener(user_agent_suffix='famasf1/dagd_shorten', max_cache_size=256)

    load_insure_wb = pyxl.load_workbook(fr"C:\Users\Comseven\Documents\DHL\Completed\11-65\All_Nov_FCB.xlsx", data_only=True)
    load_insure_wb_sh = load_insure_wb.sheetnames
    active_Sheet = load_insure_wb[load_insure_wb_sh[0]]
    for row in range(2, active_Sheet.max_row+1):

        long_link_list = []
        long_link = active_Sheet.cell(row=row, column=11).value
        long_link_list.append(long_link)
    
        short_link = dagd_connect.shorten_urls(long_urls=long_link_list)
        short_link_result = list(short_link.values())
        active_Sheet.cell(row=row, column=15).value = str(short_link_result[0])
        load_insure_wb.save(fr"C:\Users\Comseven\Documents\DHL\Completed\11-65\All_Nov_FCB_1.xlsx")

if __name__ in '__main__':
    #read()

    bitly_api_activate_once()