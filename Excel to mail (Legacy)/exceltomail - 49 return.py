from nbformat import reads
import win32com.client
from tkinter import filedialog
from tkinter import *
import openpyxl
import os
import datetime as dt

root = Tk()
root.filename = filedialog.askopenfilename(initialdir="/",title="เลือกไฟล์ Excel",filetypes=(("Excel","*.xlsx"),("All files","*.*")))

workbook = openpyxl.load_workbook(root.filename, data_only=True)
sheets = workbook.sheetnames
readsheet = workbook[sheets[2]]

def sentmail(): #dummy formula
    for i in range(2,readsheet.max_row+1): #not sure why while work but ok

        store_ID = readsheet.cell(row=i, column=1).value
        dhlTracking = readsheet.cell(row=i, column=2).value
        phyID = readsheet.cell(row=i, column=4).value
        statusimage = readsheet.cell(row=i, column=5).value
        stockout_bill = readsheet.cell(row=i, column=6).value
        mail_list = readsheet.cell(row=i, column=7).value

        outlook = win32com.client.Dispatch("Outlook.Application") #Call outlook
        From = None
        for myEmailAddress in outlook.Session.Accounts:
            if "jirayuth.p@comseven.com" in str(myEmailAddress):
                From = myEmailAddress
                break
        mail = outlook.CreateItem(0)
        mail.Subject = f'[49 โยกเข้าคลัง]สอบถามสถานะพัสดุ ID: 49 สินค้าโยกเข้าคลัง PHYID : {phyID}'
        mail.To = str(mail_list)
        mail.Cc = f'mailto:Pairin@COMSEVEN.COM'
        html = f'''

        <p style='margin-right:0cm;margin-left:0cm;font-size:15px;font-family:"Calibri","sans-serif";margin:0cm;margin-bottom:.0001pt;'><strong><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";color:red;'>เรียน</span></strong><strong><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";'>&nbsp;เจ้าหน้าที่ที่เกี่ยวข้องประจำร้าน&nbsp;</span></strong><strong><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";'>ID :&nbsp;{store_ID}</span></strong></p>

<p style='margin-right:0cm;margin-left:0cm;font-size:15px;font-family:"Calibri","sans-serif";margin:0cm;margin-bottom:.0001pt;'><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";'>ขออนุญาติรบกวนสอบถามสถานะการจัดส่ง&nbsp;</span><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";'>PHYID บิล&nbsp;{phyID}&nbsp; Tracking {dhlTracking} ว่าได้มีการจัดส่งมาหรือยังครับ</span></p>
<img src={stockout_bill}/>
<p style='margin-right:0cm;margin-left:0cm;font-size:15px;font-family:"Calibri","sans-serif";margin:0cm;margin-bottom:.0001pt;'><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";'>เนื่องจากการตรวจสอบเบื้องต้นพบว่าเลข&nbsp;</span><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";'>Tracking ที่ว่ายังอยู่ในสถานะรอทาง&nbsp;DHL เข้ารับพัสดุ&nbsp;</span></p>
<img src={statusimage}/>
<p style='margin-right:0cm;margin-left:0cm;font-size:15px;font-family:"Calibri","sans-serif";margin:0cm;margin-bottom:.0001pt;'><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";'>ในกรณีที่ยังไม่มีการจัดส่ง ตอนนี้เราได้</span><strong><span style='font-size:24px;font-family:"TH Sarabun New","sans-serif";'>ประสานงานติดต่อกับทาง&nbsp;</span></strong><strong><u><span style='font-size:24px;font-family:"TH Sarabun New","sans-serif";color:#843C0C;'>DHL</span></u></strong><span style='font-size:24px;font-family:"TH Sarabun New","sans-serif";'>&nbsp;</span><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";'>ให้ส่งรถเข้าไปรับพัสดุเรียบร้อยแล้ว และจะมีรถเข้าไปรับพัสดุ</span><strong><span style='font-size:24px;font-family:"TH Sarabun New","sans-serif";'>ภายใน&nbsp;</span></strong><strong><span style='font-size:24px;font-family:"TH Sarabun New","sans-serif";'>1-2 วัน</span></strong><span style='font-size:24px;font-family:"TH Sarabun New","sans-serif";'>&nbsp;</span><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";'>ให้เตรียมพัสดุรอการจัดส่งได้เลยครับ&nbsp;</span></p>
<p style='margin-right:0cm;margin-left:0cm;font-size:15px;font-family:"Calibri","sans-serif";margin:0cm;margin-bottom:.0001pt;'><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";'>ถ้าหากตรวจสอบภายใน</span><strong><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";color:red;'>วัน</span></strong><strong><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";color:red;'>อังคาร</span></strong><strong><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";color:red;'>ที่&nbsp;</span></strong><strong><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";color:red;'>10</span></strong><strong><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";color:red;'>&nbsp;พฤษภาคม&nbsp;2565</span></strong><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";color:red;'>&nbsp;</span><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";'>นี้แล้วยังไม่มีการจัดส่งเข้ามา อาจจะมีการต้องขอโอนยอดกลับ เนื่องจากทางคลังต้องนับจำนวนสรุปจำนวนสินค้าในคลังทั้งหมดให้ทันรอบก่อนจะทำการ</span><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";'>&nbsp;Return สินค้าคืน&nbsp;Supplier ครับ&nbsp;</span></p>
<p style='margin-right:0cm;margin-left:0cm;font-size:15px;font-family:"Calibri","sans-serif";margin:0cm;margin-bottom:.0001pt;'><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";'>หากมีข้อสงสัยเพิ่มเติม ให้ติดต่อที่อีเมลพี่ไพรินทร์&nbsp;</span><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";'><a href="mailto:Pairin@COMSEVEN.COM">Pairin@COMSEVEN.COM</a> ได้เลยครับ ถ้าหากมีการจัดส่งแล้วรบกวนขอหลักฐานการจัดส่ง (ถ้ามี) ด้วยครับ</span></p>
<p style='margin-right:0cm;margin-left:0cm;font-size:15px;font-family:"Calibri","sans-serif";margin:0cm;margin-bottom:.0001pt;'><span style='font-size:21px;font-family:"TH Sarabun New","sans-serif";'>ขอบคุณครับ</span></p>
        
        '''       
        mail.HTMLBody = html
        if From != None:
        # This line basically calls the "mail.SendUsingAccount = xyz@email.com" outlook VBA command
            mail._oleobj_.Invoke(*(64209, 0, 8, 0, From))
        
        mail.Send()
sentmail()
