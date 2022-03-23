import win32com.client
from cProfile import run
from fileinput import filename
from tkinter import filedialog
from tkinter import *
import openpyxl
import os
import datetime as dt

root1 = Tk()
root1.filename = filedialog.askopenfilename(initialdir="/",title="เลือกไฟล์ Excel",filetypes=(("Excel","*.xlsx"),("All files","*.*")))

workbook = openpyxl.load_workbook(root1.filename, data_only=True)
sheets = workbook.sheetnames
readsheet = workbook[sheets[0]]

def sentmail(): #dummy formula
    attlist = []
    row = 1 
    while True: #not sure why while work but ok
        #value1 = readsheet.cell(row=row,column=3).value #first, get value to use as your header
        email2 = readsheet.cell(row=row,column=2).value #second, get email row
        branch3 = readsheet.cell(row=row,column=1).value #Branch ID
        #amouth4 = readsheet.cell(row=row,column=5).value #amouth
        # amouth4 = readsheet.cell(row=row,column=9).value #amouth of bill
        outlook = win32com.client.Dispatch("Outlook.Application") #Call outlook
        From = None
        for myEmailAddress in outlook.Session.Accounts:
            if "jirayuth.p@comseven.com" in str(myEmailAddress):
                From = myEmailAddress
                break
        j = range(1,row) #thank you stackoverflow 
        mail = outlook.CreateItem(0)
        mail.Subject = '[Trade-in]แจ้งขอเปลี่ยนวิธีการโอน Trade In มายังคลัง 49'
        mail.To = str(email2)
        mail.Cc = f'Store_Support_Group <store_support_group@comseven.com>'
        html = f'''
        <html> \
        <body> \
            <p style='margin:0cm;margin-bottom:.0001pt;font-size:15px;font-family:"Calibri","sans-serif";'><span style='font-size:27px;font-family:"TH Sarabun New","sans-serif";'>เรียน หน้าร้าน&nbsp;</span><span style='font-size:27px;font-family:"TH Sarabun New","sans-serif";'>ID : {branch3}&nbsp;</span></p>
<p style='margin:0cm;margin-bottom:.0001pt;font-size:15px;font-family:"Calibri","sans-serif";'><span style='font-size:27px;font-family:"TH Sarabun New","sans-serif";'>&nbsp; &nbsp; &nbsp; &nbsp; เนื่องจากการตรวจสอบเบื้องต้น มีการพบว่าทางสาขามีการส่งสินค้า Trade In <strong><span style="color:red;">โดยไม่ทำตาม<u>ขั้นตอนที่ถูกต้อง</u></span></strong>เป็นจำนวนมาก <span style="color:#1F497D;">จึงรบกวนขอความร่วมมือให้ทำตาม</span><strong>วิธีการทำส่งสินค้า&nbsp;</strong><strong>Trade In</strong> <strong>แบบใหม่</strong> โดยมีการเปลี่ยนแปลงหลักๆ เพิ่มเติมดังนี้</span></p>
<div style='margin:0cm;margin-bottom:.0001pt;font-size:15px;font-family:"Calibri","sans-serif";'>
    <p><span style='font-family:"TH Sarabun New","sans-serif";font-size:27px;'>- ในเมนูรับของส่งซ่อมจากลูกค้า (</span><span style="font-size:20.0pt;">Reverse Product) ตรงผู้ซื้อให้ใส่เป็นชื่อซัพพลายเออร์ลงไป ใช้รหัสตามภาพด้านล่างนี้</span></p>
</div>
<p style='margin:0cm;margin-bottom:.0001pt;font-size:15px;font-family:"Calibri","sans-serif";margin-left:18.0pt;'><br></p>
<p style='margin:0cm;margin-bottom:.0001pt;font-size:15px;font-family:"Calibri","sans-serif";margin-left:18.0pt;'><span style='font-size:27px;font-family:"TH Sarabun New","sans-serif";'><img src="https://i.imgur.com/SbiUyl5.png"></span></p>
<p><span style='font-family:"TH Sarabun New","sans-serif";font-size:27px;'>- หลังจากทำการใส่รายละเอียดเสร็จเรียบร้อยแล้ว เมื่อกด&nbsp;</span><span style='font-family:"TH Sarabun New","sans-serif";font-size:27px;'>OK แล้ว<strong>ให้&nbsp;</strong><strong>Print ใบส่งซ่อมแนบกับเครื่อง</strong>มาด้วย <strong><span style="background:silver;">(</span></strong><strong><span style="background:silver;">Print ใส่อะไรมาก็ได้ขอให้&nbsp;Print มา</span></strong><span style="background:;;">)</span></span></p>
<p style='margin:0cm;margin-bottom:.0001pt;font-size:15px;font-family:"Calibri","sans-serif";margin-left:18.0pt;'><span style='font-size:27px;font-family:"TH Sarabun New","sans-serif";'><img src="https://i.imgur.com/00MAV7d.png"></span></p>
<div style='margin:0cm;margin-bottom:.0001pt;font-size:15px;font-family:"Calibri","sans-serif";'>
    <p><span style='font-family:"TH Sarabun New","sans-serif";font-size:27px;'>- หน้าตาบิลรับซ่อมที่ต้องแนบมา (กรณีนี้คือ&nbsp;</span><span style='font-family:"TH Sarabun New","sans-serif";font-size:27px;'>Print จาก&nbsp;A4)</span></p>
    <p style='margin:0cm;margin-bottom:.0001pt;font-size:15px;font-family:"Calibri","sans-serif";margin-left:18.0pt;'><span style='font-size:27px;font-family:"TH Sarabun New","sans-serif";'><img src="https://i.imgur.com/8ZmNciW.png"></span></p>
</div>
<div style='margin:0cm;margin-bottom:.0001pt;font-size:15px;font-family:"Calibri","sans-serif";'>
    <p><span style='font-family:"TH Sarabun New","sans-serif";font-size:27px;'>- ตอนทำโอนสินค้า&nbsp;</span><span style="font-size:20.0pt;">Stock Out ออกมา ให้ทำโอนออกมาทีเดียวแบบนี้ <strong>(1 บิลต้องไม่เกิน 5 ชิ้น)&nbsp;</strong>ไม่จำเป็นต้อง&nbsp;Comment อะไรลงไป<span style="color:#1F497D;">&nbsp;ปริ้นใบ&nbsp;</span><span style="color:#1F497D;">Stock Out แบบในภาพใส่มาในกล่องด้วยครับ</span></span></p>
</div>
<p style='margin:0cm;margin-bottom:.0001pt;font-size:15px;font-family:"Calibri","sans-serif";'><br></p>
<p style='margin:0cm;margin-bottom:.0001pt;font-size:15px;font-family:"Calibri","sans-serif";margin-left:18.0pt;'><span style='font-size:27px;font-family:"TH Sarabun New","sans-serif";'><img src="https://i.imgur.com/lc78hFj.jpg"></span></p>
<p style='margin:0cm;margin-bottom:.0001pt;font-size:15px;font-family:"Calibri","sans-serif";'><span style='font-size:27px;font-family:"TH Sarabun New","sans-serif";color:#1F497D;'>ขอบคุณที่ให้ความร่วมมือครับ</span></p>
<p style='margin:0cm;margin-bottom:.0001pt;font-size:15px;font-family:"Calibri","sans-serif";'><span style='font-size:27px;font-family:"TH Sarabun New","sans-serif";'>&nbsp;</span></p>

        </body> \
        </html>      '''       
        mail.HTMLBody = html
        if From != None:
        # This line basically calls the "mail.SendUsingAccount = xyz@email.com" outlook VBA command
            mail._oleobj_.Invoke(*(64209, 0, 8, 0, From))
        
        mail.Send()
        row += 1
        if email2 is 'None':
            break
        else:
            continue
sentmail()