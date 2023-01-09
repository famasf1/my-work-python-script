import collections
import ctypes
import openpyxl
from tkinter import filedialog, messagebox
from tkinter import *
import os
import sys
import datetime as dt
import win32com.client
from barcode import Code128
from barcode.writer import ImageWriter
import treepoem as tp
from openpyxl.styles import Alignment

# MAIN


def main():
    global all_data, test_sheet
    root = Tk()
    root.excel = filedialog.askopenfilename(
        title='เลือกไฟล์ Excel', filetypes=(("Excel", "*.xlsx"), ('All Files', '*.*')))
    worksheet = openpyxl.load_workbook(root.excel, data_only=True)
    sheet = worksheet.sheetnames
    sticker_front = worksheet[sheet[0]]
    sticker_side = worksheet[sheet[1]]
    all_data = worksheet[sheet[2]]
    test_sheet = worksheet[sheet[3]]
    col1_tolist = []
    col2_tolist = []
    col3_tolist = []

    max_rows = int(all_data.max_row / 8) + 1

    # for this loop. if this loop number have more than max row of all_data sheet.
    # after divided it by 8 (barcode have only 8 slots)
    # stop the operation.
    # But if the row is less than 8, do a match case.
    rownumber = all_data.max_row
    match all_data.max_row:
        case 1:
            rownumber = all_data.max_row + 7
        case 2:
            rownumber = all_data.max_row + 6
        case 3:
            rownumber = all_data.max_row + 5
        case 4:
            rownumber = all_data.max_row + 4
        case 5:
            rownumber = all_data.max_row + 3
        case 6:
            rownumber = all_data.max_row + 2
        case other:
            rownumber = all_data.max_row + 1

    print(rownumber)
    for val in range(1, rownumber + 1):
        
        col1 = all_data.cell(row=val, column=1).value
        col1 = f'No.{col1}'
        col2 = all_data.cell(row=val, column=2).value
        col3 = all_data.cell(row=val, column=3).value

        col1_tolist.append(col1)
        col2_tolist.append(col2)
        col3_tolist.append(col3)


        def template_front1():  # top left
            # top left
            if len(col1_tolist) % 8 == 1:
                try:
                    sticker_front['B2'].value = col1_tolist[0]
                    sticker_front['B3'].value = col2_tolist[0]
                    sticker_front['B5'].value = col3_tolist[0]
                    with open(rf"D:\Workstuff\my-work-python-script\Print Barcode (dev)\result\{col3_tolist[0]}.png".replace("/00", "-00"), "wb+") as files:
                        img = tp.generate_barcode(
                            barcode_type="code128", data=f"{col3}",)
                        img.save(files)
                    img1 = openpyxl.drawing.image.Image(
                        rf"D:\Workstuff\my-work-python-script\Print Barcode (dev)\result\{col3}.png".replace("/00", "-00"))
                    img1.height = 48
                    img1.width = 173.8
                    img1.anchor = 'B4'
                    sticker_front.add_image(img1)
                except:
                    sticker_front['B2'].value = 'None'
                    sticker_front['B3'].value = 'None'
                    worksheet.save(root.excel)

        def template_front2():  # top right
            if len(col1_tolist) % 8 == 2:
                try:
                    sticker_front['E2'].value = col1_tolist[1]
                    sticker_front['E3'].value = col2_tolist[1]
                    sticker_front['E5'].value = col3_tolist[1]
                    with open(rf"D:\Workstuff\my-work-python-script\Print Barcode (dev)\result\{col3_tolist[1]}.png".replace("/00", "-00"), "wb+") as files:
                        img = tp.generate_barcode(barcode_type="code128", data=f"{col3}",)
                        img.save(files)
                    img1 = openpyxl.drawing.image.Image(rf"D:\Workstuff\my-work-python-script\Print Barcode (dev)\result\{col3}.png".replace("/00","-00"))
                    img1.height = 48
                    img1.width = 173.8
                    img1.anchor = 'E4'
                    sticker_front.add_image(img1)
                except:
                    sticker_front['E2'].value = 'None'
                    sticker_front['E3'].value = 'None'
                    worksheet.save(root.excel)

        def template_front3():  # 2nd left
            if len(col1_tolist) % 8 == 3:
                try:
                    sticker_front['B6'].value = col1_tolist[2]
                    sticker_front['B7'].value = col2_tolist[2]
                    sticker_front['B9'].value = col3_tolist[2]
                    with open(rf"D:\Workstuff\my-work-python-script\Print Barcode (dev)\result\{col3_tolist[2]}.png".replace("/00", "-00"), "wb+") as files:
                        img = tp.generate_barcode(barcode_type="code128", data=f"{col3}",)
                        img.save(files)
                    img1 = openpyxl.drawing.image.Image(rf"D:\Workstuff\my-work-python-script\Print Barcode (dev)\result\{col3}.png".replace("/00","-00"))
                    img1.height = 48
                    img1.width = 173.8
                    img1.anchor = 'B8'
                    sticker_front.add_image(img1)
                except:
                    sticker_front['B6'].value = 'None'
                    sticker_front['B7'].value = 'None'
                    worksheet.save(root.excel)

        def template_front4():  # 2nd left
            if len(col1_tolist) % 8 == 4:
                try:
                    sticker_front['E6'].value = col1_tolist[3]
                    sticker_front['E7'].value = col2_tolist[3]
                    sticker_front['E9'].value = col3_tolist[3]
                    with open(rf"D:\Workstuff\my-work-python-script\Print Barcode (dev)\result\{col3_tolist[3]}.png".replace("/00", "-00"), "wb+") as files:
                        img = tp.generate_barcode(barcode_type="code128", data=f"{col3}",)
                        img.save(files)
                    img1 = openpyxl.drawing.image.Image(rf"D:\Workstuff\my-work-python-script\Print Barcode (dev)\result\{col3}.png".replace("/00","-00"))
                    img1.height = 48
                    img1.width = 173.8
                    img1.anchor = 'E8'
                    sticker_front.add_image(img1)
                except:
                    sticker_front['E6'].value = 'None'
                    sticker_front['E7'].value = 'None'                    
                    worksheet.save(root.excel)

        def template_front5():  # 3rd left
            if len(col1_tolist) % 8 == 5:
                try:
                    sticker_front['B10'].value = col1_tolist[4]
                    sticker_front['B11'].value = col2_tolist[4]
                    sticker_front['B13'].value = col3_tolist[4]
                    with open(rf"D:\Workstuff\my-work-python-script\Print Barcode (dev)\result\{col3_tolist[4]}.png".replace("/00", "-00"), "wb+") as files:
                        img = tp.generate_barcode(barcode_type="code128", data=f"{col3}",)
                        img.save(files)
                    img1 = openpyxl.drawing.image.Image(rf"D:\Workstuff\my-work-python-script\Print Barcode (dev)\result\{col3}.png".replace("/00","-00"))
                    img1.height = 48
                    img1.width = 173.8
                    img1.anchor = 'B12'
                    sticker_front.add_image(img1)
                except:
                    sticker_front['B10'].value = 'None'
                    sticker_front['B11'].value = 'None'
                    worksheet.save(root.excel)

        def template_front6():  # 3rd left
            if len(col1_tolist) % 8 == 6:
                try:
                    sticker_front['E10'].value = col1_tolist[5]
                    sticker_front['E11'].value = col2_tolist[5]
                    sticker_front['E13'].value = col3_tolist[5]
                    with open(rf"D:\Workstuff\my-work-python-script\Print Barcode (dev)\result\{col3_tolist[5]}.png".replace("/00", "-00"), "wb+") as files:
                        img = tp.generate_barcode(barcode_type="code128", data=f"{col3}",)
                        img.save(files)
                    img1 = openpyxl.drawing.image.Image(rf"D:\Workstuff\my-work-python-script\Print Barcode (dev)\result\{col3}.png".replace("/00","-00"))
                    img1.height = 48
                    img1.width = 173.8
                    img1.anchor = 'E12'
                    sticker_front.add_image(img1)
                except:
                    sticker_front['E10'].value = 'None'
                    sticker_front['E11'].value = 'None'
                    worksheet.save(root.excel)

        def template_front7():  # bottom left
            if len(col1_tolist) % 8 == 7:
                try:
                    sticker_front['B14'].value = col1_tolist[6]
                    sticker_front['B15'].value = col2_tolist[6]
                    sticker_front['B17'].value = col3_tolist[6]
                    with open(rf"D:\Workstuff\my-work-python-script\Print Barcode (dev)\result\{col3_tolist[6]}.png".replace("/00", "-00"), "wb+") as files:
                        img = tp.generate_barcode(barcode_type="code128", data=f"{col3}",)
                        img.save(files)
                    img1 = openpyxl.drawing.image.Image(rf"D:\Workstuff\my-work-python-script\Print Barcode (dev)\result\{col3}.png".replace("/00","-00"))
                    img1.height = 48
                    img1.width = 173.8
                    img1.anchor = 'B16'
                    sticker_front.add_image(img1)
                except:
                    sticker_front['B14'].value = 'None'
                    sticker_front['B15'].value = 'None'
                    worksheet.save(root.excel)

        def template_front_final():  # bottom right
            if len(col1_tolist) % 8 == 0:
                
                try:
                    sticker_front['E14'].value = col1_tolist[7]
                    sticker_front['E15'].value = col2_tolist[7]
                    sticker_front['E17'].value = col3_tolist[7]
                    with open(rf"D:\Workstuff\my-work-python-script\Print Barcode (dev)\result\{col3_tolist[7]}.png".replace("/00", "-00"), "wb+") as files:
                        img = tp.generate_barcode(barcode_type="code128", data=f"{col3}",)
                        img.save(files)
                    img1 = openpyxl.drawing.image.Image(rf"D:\Workstuff\my-work-python-script\Print Barcode (dev)\result\{col3}.png".replace("/00","-00"))
                    img1.height = 48
                    img1.width = 173.8
                    img1.anchor = 'E16'
                    sticker_front.add_image(img1)
                    worksheet.save(root.excel)
                    col1_tolist.clear()
                    col2_tolist.clear()
                    col3_tolist.clear()
                    dispatcher = win32com.client.Dispatch('Excel.Application')
                    dispatcher.visible = False
                    wb = dispatcher.Workbooks.Open(str(root.excel))
                    getsheet = wb.Worksheets([1])
                    #getsheet.PageSetup.FitToPagesTall = 1
                    #getsheet.PageSetup.FitToPagesWide = 1
                    getsheet.PrintOut()
                    wb.Close(True)
                except:
                    sticker_front['E14'].value = 'None'
                    sticker_front['E15'].value = 'None'
                    worksheet.save(root.excel)
                    col1_tolist.clear()
                    col2_tolist.clear()
                    col3_tolist.clear()
                    dispatcher = win32com.client.Dispatch('Excel.Application')
                    dispatcher.visible = False
                    wb = dispatcher.Workbooks.Open(str(root.excel))
                    getsheet = wb.Worksheets([1])
                    #getsheet.PageSetup.FitToPagesTall = 1
                    #getsheet.PageSetup.FitToPagesWide = 1
                    getsheet.PrintOut()
                    wb.Close(True)
    
        template_front1()
        template_front2()
        template_front3()
        template_front4()
        template_front5()
        template_front6()
        template_front7()
        template_front_final()





### TEST ROOM
def test_room():
    root = Tk()
    root.excel = filedialog.askopenfilename(title='เลือกไฟล์ Excel', filetypes=(("Excel", "*.xlsx"),('All Files','*.*')))
    worksheet = openpyxl.load_workbook(root.excel, data_only=True)
    sheet = worksheet.sheetnames
    sticker_front = worksheet[sheet[0]]
    sticker_side = worksheet[sheet[1]]
    all_data = worksheet[sheet[2]]  
    col1_tolist = []
    col2_tolist = []
    col3_tolist = []

    for val in range(1, all_data.max_row+1):
        col1 = all_data.cell(row=val, column=1).value
        col1 = f'No.{col1}'
        col2 = all_data.cell(row=val, column=2).value
        col3 = all_data.cell(row=val, column=3).value
        col1_tolist.append(col1)
        col2_tolist.append(col2)
        col3_tolist.append(col3)
        

        def template_front1(): #top left
            if len(col1_tolist) % 4 == 1:
                sticker_front['A1'].value = col1_tolist[0]
                sticker_front['A2'].value = col2_tolist[0]
                sticker_front['A4'].value = col3_tolist[0]
                with open(rf"D:\Workstuff\my-work-python-script\Print Barcode (dev)\result\{col3}.png".replace("/00","-00"), "wb+") as files:
                    img = tp.generate_barcode(barcode_type="code128", data=f"{col3}",)
                    img.save(files)
                img1 = openpyxl.drawing.image.Image(rf"D:\Workstuff\my-work-python-script\Print Barcode (dev)\result\{col3}.png".replace("/00","-00"))
                img1.height = 48
                img1.width = 173.8
                img1.alignment = Alignment(horizontal='center', vertical='center')
                sticker_front.add_image(img1, 'A3')
                worksheet.save(root.excel)
        template_front1()
        break

### Extra
def MBox(title, text, style):
    return ctypes.windll.user32.MessageBoxW(0,text,title,style)

def barcode_generator():
    for i in range(1, all_data.max_row+1):
        values = all_data.cell(row=i, column=3).value
        with open(rf"D:\Workstuff\my-work-python-script\Print Barcode\result\{values}.png".replace("/00","-00"), "wb+") as files:
            Code128(values, writer=ImageWriter()).write(files)
        break

if __name__ in '__main__':
    main()
    #test_room()

