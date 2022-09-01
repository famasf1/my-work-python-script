import collections
import ctypes
import openpyxl
from tkinter import filedialog, messagebox
from tkinter import *
import os
import sys
import datetime as dt
import win32com.client
from barcode import Code128
from barcode.writer import ImageWriter
from PIL import Image

### MAIN
def main():
    global all_data, test_sheet
    root = Tk()
    root.excel = filedialog.askopenfilename(title='เลือกไฟล์ Excel', filetypes=(("Excel", "*.xlsx"),('All Files','*.*')))
    worksheet = openpyxl.load_workbook(root.excel, data_only=True)
    sheet = worksheet.sheetnames
    sticker_front = worksheet[sheet[0]]
    sticker_side = worksheet[sheet[1]]
    all_data = worksheet[sheet[2]]
    test_sheet = worksheet[sheet[3]]
    col1_tolist = []
    col2_tolist = []
    col3_tolist = []

    for i in range(1,sticker_side.max_row+1):
        pass


### TEST ROOM
def test_room():
    pass

### Extra
def MBox(title, text, style):
    return ctypes.windll.user32.MessageBoxW(0,text,title,style)

def barcode_generator():
    for i in range(1, all_data.max_row+1):
        values = all_data.cell(row=i, column=3).value
        with open(rf"D:\Workstuff\my-work-python-script\Print Barcode\result\{values}.png".replace("/00","-00"), "rb+") as files:
            Code128(values, writer=ImageWriter()).write(files)
        break

if __name__ in '__main__':
    main()
    #test_room()