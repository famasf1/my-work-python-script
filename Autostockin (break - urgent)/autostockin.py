from netrc import NetrcParseError
from tkinter import filedialog, messagebox
from tkinter import *
import pyautogui as pyg
import openpyxl
import pyperclip
import random
############### CLASS ################

### List all employee ###
class Employeelist:
    #เขียนครบ ตามด้วยไอดีพนักงาน
    def __init__(self,index, name, staffid):
        self.name = name
        self.staffid = staffid
        self.index = index
    #ขานชื่อ - เรียกชื่อจริง
    #ขานเลขที่ - เรียกเลขที่
    def ขานชื่อ(self):
        return self.name

    def ขานเลขพนักงาน(self):
        return self.staffid

    def get_index(self):
        return self.index

########################################

##### SETTING

def start_script():
    try:
        global workbook, worksheet,readpicerrorfound, ok, employeelist_index, employeelist_name, employeelist_id
        global จิรายุทธ, วรัญญู, มรกต, วุฒิภัทร, numb
        ###################################### LIST EMPLOYEE #####################################
        จิรายุทธ = Employeelist(0,'ครบ / โบ้', 22608) 
        วรัญญู = Employeelist(1,'ครบ / ตั้ม', 24179)
        มรกต = Employeelist(2,'ครบ / ปาน', 23947)
        วุฒิภัทร = Employeelist(3,'ครบ / มาร์ค',23800)
        numb = [0,1,2,3]
        employeelist_index = [จิรายุทธ.index, วรัญญู.index, มรกต.index, วุฒิภัทร.index]
        ##########################################################################################

        root = Tk()
        root.excel = filedialog.askopenfilename(initialdir='/Desktop',title='เลือกไฟล์ Excel สำหรับ Stock-Out33', filetypes=(('Excel','*.xlsx'),('All Files','*.*')))
        workbook = openpyxl.load_workbook(root.excel, data_only=True)
        root.withdraw()
        sheet = workbook.sheetnames
        worksheet = workbook[sheet[0]] 
        readpicerrorfound = pyg.locateCenterOnScreen('asset/checkerror.png')
        ok = pyg.locateOnScreen('asset/ok.png')

    except Exception as e:
        messagebox.showerror('Python Error', f'{e}')
        exit()

############################
############################
############################
###### MAIN FUNCTION #######

def main(): #short for default behavior
    ### First time starting
    start_script()

    def firstStart(start):
        if start == 1:
            pyg.hotkey('alt','k')
            pyg.press('i')
        else:
            pass
    ### Nect
    def nextStart(next):
        if next == 1:
            pass
        else:
            pyg.hotkey('alt','k')
            pyg.press('i')

    for i in range(1, worksheet.max_row+1):
        stockoutid = worksheet.cell(row=i, column=1).value
        if stockoutid:
            nextStart(i)
            firstStart(i)
            print('Start Stock In')

            #pick staff id

            r = random.choice(employeelist_index)
            match r:
                case จิรายุทธ.index:
                    pyg.typewrite(จิรายุทธ.staffid)
                case วรัญญู.index:
                    pyg.typewrite(วรัญญู.staffid)
                case มรกต.index:
                    pyg.typewrite(มรกต.staffid)
                case วุฒิภัทร.index:
                    pyg.typewrite(วุฒิภัทร.staffid)

            pressenter(2)
            pyg.typewrite(str(stockoutid)) #stockout
            pressenter(2)

            #name

            match r:
                case จิรายุทธ.index:
                    pyg.typewrite(จิรายุทธ.name)
                case วรัญญู.index:
                    pyg.typewrite(วรัญญู.name)
                case มรกต.index:
                    pyg.typewrite(มรกต.name)
                case วุฒิภัทร.index:
                    pyg.typewrite(วุฒิภัทร.name)
            pyg.hotkey('ctrl', 'v')
            pyg.sleep(15)
            ##ready

            pyg.hotkey('alt','f')
            pyg.press('o')
            try:
                if readpicerrorfound:
                    pressenter(1)
                    print('Found Error')
                    continue
                else:
                    print('Error is not Found')
                    pressenter(1)
                    pyg.press('left')
                    pressenter(1)
                    pyg.sleep(3)
                    pressenter(4)
                    continue
            except Exception as e:
                messagebox.showerror('Python Error', f'{e}')
        else: break

####################################################################
######### Extra Function that doesn't involve with any of the above

def custom_comment(comment):
    pyperclip.copy(comment)
    pyg.hotkey('ctrl', 'v')

def pressenter(numberoftimes):
    for i in range(0, numberoftimes):
        pyg.press('enter')
        if i == numberoftimes:
            break

def clickleft(numberoftimes):
    for i in range(0, numberoftimes):
        pyg.leftClick()
        pyg.sleep(1.5)
        if i == numberoftimes:
            break

def test():
    จิรายุทธ = Employeelist(0,'ครบ / โบ้', 22608) 
    วรัญญู = Employeelist(1,'ครบ / ตั้ม', 24179)
    มรกต = Employeelist(2,'ครบ / ปาน', 23947)
    วุฒิภัทร = Employeelist(3,'ครบ / มาร์ค',23800)
    print(จิรายุทธ.name)

try:
  #  test()
    main()
except Exception as e:
    messagebox.showerror('Python Error', f'{e}')
    exit()








































