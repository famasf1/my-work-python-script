#BLOCK 117

import openpyxl
import os, glob, shutil
from tkinter import Tk, filedialog
import itertools
import natsort as ns
import time
import pandas as pd

asset_dir = os.path.join(r"D:\Workstuff\my-work-python-script\project_bebephone", r"Asset")

root = Tk()
#get asset
asset_dir_117 = os.path.join(asset_dir, r"block117template.xlsm")
block117_wb = openpyxl.load_workbook(asset_dir_117, data_only=True, keep_vba=True)
block117_ws = block117_wb.sheetnames
block117 = block117_wb[block117_ws[0]]

#get data
get_wb = filedialog.askopenfilename(title="เลือกไฟล์ Excel", filetypes=([("*Excel Files", "*.xlsx"), ("All Files", "*.*")]))
data_wb = openpyxl.load_workbook(get_wb, data_only=True)
data_sheet = data_wb.sheetnames
sheet = data_wb[data_sheet[0]]
data_list = []
new_dir = os.path.join(os.getcwd(), "excel")


def movefiletofolder(filenameid):
    def grouper(S, n): #https://stackoverflow.com/questions/12559055/for-every-x-number-of-files-create-new-directory-and-move-files-using-python
        iterator = iter(S)
        while True:
            item = list(itertools.islice(iterator, n))
            print(item)
            if len(item) == 0:
                break
            yield item

    os.chdir(rf"D:\Workstuff\my-work-python-script\project_bebephone\excel\{filenameid}")
    fnames = ns.natsorted(glob.glob('*.xlsm'))
    for i, fnames in enumerate(grouper(fnames,10)):
        print(i)
        print(fnames)
        dirname = f"{i + 1}"
        try:
            os.mkdir(dirname)
        except FileExistsError:
            pass
        for fname in fnames:
            shutil.move(fname, dirname)


def main():
    bebe_startrow = 2
    com7barcode_startrow = 3
    com7barcodetext_startrow = 4
    filename = sheet.cell(row=1, column=3).value
    roundset = 1
    os.chdir(rf"D:\Workstuff\my-work-python-script\project_bebephone\excel")
    try:
        os.mkdir(str(filename))
    except FileExistsError:
        pass
    os.chdir(rf"D:\Workstuff\my-work-python-script\project_bebephone\excel\{filename}")
    row = 2
    while row != sheet.max_row+1:
        bebecode = sheet.cell(row=row, column=1).value
        productcode = sheet.cell(row=row, column=2).value
        try:
            needed_set = int(sheet.cell(row=row, column=3).value / 9)
        except TypeError:
            break
        data_list.append(needed_set)
        if needed_set >= 14:
            with open("โปรด manual.txt", mode="a") as file:
                file.writelines(productcode)
                row += 1
                continue
        elif sum(data_list) > 14:
            bebe_startrow = 2
            com7barcode_startrow = 3
            com7barcodetext_startrow = 4
            data_list.clear()
            block117_wb.save(f"{filename}_{row-1}.xlsm")
            continue
        elif sum(data_list) == 14:
            bebe_startrow = 2
            com7barcode_startrow = 3
            com7barcodetext_startrow = 4
            data_list.clear()
            block117_wb.save(f"{filename}_{row-1}.xlsm")
            continue
        else:
            row += 1
            if row == sheet.max_row+1:
                for round in range(1,needed_set+1):
                    for col in range(2,11):
                        block117.cell(row=bebe_startrow, column=col).value = bebecode 
                        block117.cell(row=com7barcode_startrow, column=col).value = f'''=code128("{productcode}")'''
                        block117.cell(row=com7barcodetext_startrow, column=col).value = productcode
                    if needed_set >= 1:
                        bebe_startrow += 4
                        com7barcode_startrow += 4
                        com7barcodetext_startrow += 4
                    else:
                        break
                row += 1
                block117_wb.save(f"{filename}_{row-1}.xlsm")
        #set
        
        for round in range(1,needed_set+1):
            for col in range(2,11):
                block117.cell(row=bebe_startrow, column=col).value = bebecode 
                block117.cell(row=com7barcode_startrow, column=col).value = f'''=code128("{productcode}")'''
                block117.cell(row=com7barcodetext_startrow, column=col).value = productcode
                #end of round one
            if needed_set >= 1:
                bebe_startrow += 4
                com7barcode_startrow += 4
                com7barcodetext_startrow += 4
                if sum(data_list) >= 14:
                    break
                else:
                    continue
            
    movefiletofolder(filename)
def combine():
    os.chdir("D:\Workstuff\my-work-python-script\project_bebephone\excel\Mixed")
    print(os.listdir(r"D:\Workstuff\my-work-python-script\project_bebephone\excel\Mixed"))
    for filename in os.listdir(fr"D:\Workstuff\my-work-python-script\project_bebephone\excel\Mixed"):
        data = pd.read_excel(filename, dtype=object)
        with pd.ExcelWriter(r"D:\Workstuff\my-work-python-script\project_bebephone\excel\รวม.xlsx", mode="a", engine="openpyxl", if_sheet_exists="overlay") as file:
            file.number_format = openpyxl.styles.numbers.FORMAT_TEXT
            data.to_excel(file,sheet_name="รวม", header=False, index=False, startrow=file.sheets['รวม'].max_row)
if __name__ in "__main__":
    main()
    #combine()
    