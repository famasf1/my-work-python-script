import os, glob, shutil
import openpyxl
from tkinter import Tk, filedialog
import itertools
import natsort as ns

asset_directory = os.path.join(r"D:\Workstuff\my-work-python-script\project_bebephone", r"Asset")

root = Tk()

#get asset
PRINT_dir = r"D:\Workstuff\my-work-python-script\project_bebephone\PRINT"
asset_dir_90 = os.path.join(asset_directory, r"block90template.xlsm")
block90_wb = openpyxl.load_workbook(asset_dir_90, data_only=True, keep_vba=True)
block90_ws = block90_wb.sheetnames
block90 = block90_wb[block90_ws[0]]
#get data
get_wb = filedialog.askopenfilename(title="เลือกไฟล์ข้อมูล Excel", filetypes=([("*Excel Files", "*.xlsx"), ("All Files", "*.*")]))
data_wb = openpyxl.load_workbook(get_wb, data_only=True)
data_sheet = data_wb.sheetnames
get_sheet = data_wb[data_sheet[0]]

def main():

        #bebe row
    bebe_start_row_number = 2

    #barcode row
    com7barcode_startrow = 3

    #barcode text
    com7barcodetext_startrow = 4

    #data
    data_startrow = 3

    ran = False
    col = 2
    folder_count = 1

    for total_print_row in range(3,get_sheet.max_row+1):
        
        bebecode = get_sheet.cell(row=total_print_row,column = 2).value #0166
        com7productcode = get_sheet.cell(row=total_print_row, column = 1).value #51691
        total_sticker = get_sheet.cell(row=total_print_row, column=3).value
        total_print_data = get_sheet.cell(row=total_print_row, column = 4).value #8 #int

        if total_print_data > 1:
            os.chdir(r"D:\Workstuff\my-work-python-script\project_bebephone\ปริ้น 2 แผ่น")
            #b
            block90['B2'] = bebecode
            block90['B3'] = f'''=code128("{com7productcode}")'''
            block90['B4'] = com7productcode
            #e
            block90['E2'] = bebecode
            block90['E3'] = f'''=code128("{com7productcode}")'''
            block90['E4'] = com7productcode
            #h
            block90['H2'] = bebecode
            block90['H3'] = f'''=code128("{com7productcode}")'''
            block90['H4'] = com7productcode
            #k
            block90['K2'] = bebecode
            block90['K3'] = f'''=code128("{com7productcode}")'''
            block90['K4'] = com7productcode
            #n
            block90['N2'] = bebecode
            block90['N3'] = f'''=code128("{com7productcode}")'''
            block90['N4'] = com7productcode
            #b
            block90['B6'] = bebecode
            block90['B7'] = f'''=code128("{com7productcode}")'''
            block90['B8'] = com7productcode
            #e
            block90['E6'] = bebecode
            block90['E7'] = f'''=code128("{com7productcode}")'''
            block90['E8'] = com7productcode
            #h
            block90['H6'] = bebecode
            block90['H7'] = f'''=code128("{com7productcode}")'''
            block90['H8'] = com7productcode
            #k
            block90['K6'] = bebecode
            block90['K7'] = f'''=code128("{com7productcode}")'''
            block90['K8'] = com7productcode
            #n
            block90['N6'] = bebecode
            block90['N7'] = f'''=code128("{com7productcode}")'''
            block90['N8'] = com7productcode            
            #b
            block90['B10'] = bebecode
            block90['B11'] = f'''=code128("{com7productcode}")'''
            block90['B12'] = com7productcode
            #e
            block90['E10'] = bebecode
            block90['E11'] = f'''=code128("{com7productcode}")'''
            block90['E12'] = com7productcode
            #h
            block90['H10'] = bebecode
            block90['H11'] = f'''=code128("{com7productcode}")'''
            block90['H12'] = com7productcode
            #k
            block90['K10'] = bebecode
            block90['K11'] = f'''=code128("{com7productcode}")'''
            block90['K12'] = com7productcode
            #n
            block90['N10'] = bebecode
            block90['N11'] = f'''=code128("{com7productcode}")'''
            block90['N12'] = com7productcode            
            #b
            block90['B14'] = bebecode
            block90['B15'] = f'''=code128("{com7productcode}")'''
            block90['B16'] = com7productcode
            #e
            block90['E14'] = bebecode
            block90['E15'] = f'''=code128("{com7productcode}")'''
            block90['E16'] = com7productcode
            #h
            block90['H14'] = bebecode
            block90['H15'] = f'''=code128("{com7productcode}")'''
            block90['H16'] = com7productcode
            #k
            block90['K14'] = bebecode
            block90['K15'] = f'''=code128("{com7productcode}")'''
            block90['K16'] = com7productcode
            #n
            block90['N14'] = bebecode
            block90['N15'] = f'''=code128("{com7productcode}")'''
            block90['N16'] = com7productcode            
            #b
            block90['B18'] = bebecode
            block90['B19'] = f'''=code128("{com7productcode}")'''
            block90['B20'] = com7productcode
            #e
            block90['E18'] = bebecode
            block90['E19'] = f'''=code128("{com7productcode}")'''
            block90['E20'] = com7productcode
            #h
            block90['H18'] = bebecode
            block90['H19'] = f'''=code128("{com7productcode}")'''
            block90['H20'] = com7productcode
            #k
            block90['K18'] = bebecode
            block90['K19'] = f'''=code128("{com7productcode}")'''
            block90['K20'] = com7productcode
            #n
            block90['N18'] = bebecode
            block90['N19'] = f'''=code128("{com7productcode}")'''
            block90['N20'] = com7productcode            
            #b
            block90['B22'] = bebecode
            block90['B23'] = f'''=code128("{com7productcode}")'''
            block90['B24'] = com7productcode
            #e
            block90['E22'] = bebecode
            block90['E23'] = f'''=code128("{com7productcode}")'''
            block90['E24'] = com7productcode
            #h
            block90['H22'] = bebecode
            block90['H23'] = f'''=code128("{com7productcode}")'''
            block90['H24'] = com7productcode
            #k
            block90['K22'] = bebecode
            block90['K23'] = f'''=code128("{com7productcode}")'''
            block90['K24'] = com7productcode
            #n
            block90['N22'] = bebecode
            block90['N23'] = f'''=code128("{com7productcode}")'''
            block90['N24'] = com7productcode            
            #b
            block90['B26'] = bebecode
            block90['B27'] = f'''=code128("{com7productcode}")'''
            block90['B28'] = com7productcode
            #e
            block90['E26'] = bebecode
            block90['E27'] = f'''=code128("{com7productcode}")'''
            block90['E28'] = com7productcode
            #h
            block90['H26'] = bebecode
            block90['H27'] = f'''=code128("{com7productcode}")'''
            block90['H28'] = com7productcode
            #k
            block90['K26'] = bebecode
            block90['K27'] = f'''=code128("{com7productcode}")'''
            block90['K28'] = com7productcode
            #n
            block90['N26'] = bebecode
            block90['N27'] = f'''=code128("{com7productcode}")'''
            block90['N28'] = com7productcode            
            #b
            block90['B30'] = bebecode
            block90['B31'] = f'''=code128("{com7productcode}")'''
            block90['B32'] = com7productcode
            #e
            block90['E30'] = bebecode
            block90['E31'] = f'''=code128("{com7productcode}")'''
            block90['E32'] = com7productcode
            #h
            block90['H30'] = bebecode
            block90['H31'] = f'''=code128("{com7productcode}")'''
            block90['H32'] = com7productcode
            #k
            block90['K30'] = bebecode
            block90['K31'] = f'''=code128("{com7productcode}")'''
            block90['K32'] = com7productcode
            #n
            block90['N30'] = bebecode
            block90['N31'] = f'''=code128("{com7productcode}")'''
            block90['N32'] = com7productcode            
            #b
            block90['B34'] = bebecode
            block90['B35'] = f'''=code128("{com7productcode}")'''
            block90['B36'] = com7productcode
            #e
            block90['E34'] = bebecode
            block90['E35'] = f'''=code128("{com7productcode}")'''
            block90['E36'] = com7productcode
            #h
            block90['H34'] = bebecode
            block90['H35'] = f'''=code128("{com7productcode}")'''
            block90['H36'] = com7productcode
            #k
            block90['K34'] = bebecode
            block90['K35'] = f'''=code128("{com7productcode}")'''
            block90['K36'] = com7productcode
            #n
            block90['N34'] = bebecode
            block90['N35'] = f'''=code128("{com7productcode}")'''
            block90['N36'] = com7productcode            
            #b
            block90['B38'] = bebecode
            block90['B39'] = f'''=code128("{com7productcode}")'''
            block90['B40'] = com7productcode
            #e
            block90['E38'] = bebecode
            block90['E39'] = f'''=code128("{com7productcode}")'''
            block90['E40'] = com7productcode
            #h
            block90['H38'] = bebecode
            block90['H39'] = f'''=code128("{com7productcode}")'''
            block90['H40'] = com7productcode
            #k
            block90['K38'] = bebecode
            block90['K39'] = f'''=code128("{com7productcode}")'''
            block90['K40'] = com7productcode
            #n
            block90['N38'] = bebecode
            block90['N39'] = f'''=code128("{com7productcode}")'''
            block90['N40'] = com7productcode            
            #b
            block90['B42'] = bebecode
            block90['B43'] = f'''=code128("{com7productcode}")'''
            block90['B44'] = com7productcode
            #e
            block90['E42'] = bebecode
            block90['E43'] = f'''=code128("{com7productcode}")'''
            block90['E44'] = com7productcode
            #h4
            block90['H42'] = bebecode
            block90['H43'] = f'''=code128("{com7productcode}")'''
            block90['H44'] = com7productcode
            #k4
            block90['K42'] = bebecode
            block90['K43'] = f'''=code128("{com7productcode}")'''
            block90['K44'] = com7productcode
            #n4
            block90['N42'] = bebecode
            block90['N43'] = f'''=code128("{com7productcode}")'''
            block90['N44'] = com7productcode            
            #b
            block90['B46'] = bebecode
            block90['B47'] = f'''=code128("{com7productcode}")'''
            block90['B48'] = com7productcode
            #e
            block90['E46'] = bebecode
            block90['E47'] = f'''=code128("{com7productcode}")'''
            block90['E48'] = com7productcode
            #h
            block90['H46'] = bebecode
            block90['H47'] = f'''=code128("{com7productcode}")'''
            block90['H48'] = com7productcode
            #k
            block90['K46'] = bebecode
            block90['K47'] = f'''=code128("{com7productcode}")'''
            block90['K48'] = com7productcode
            #n
            block90['N46'] = bebecode
            block90['N47'] = f'''=code128("{com7productcode}")'''
            block90['N48'] = com7productcode            
            #b
            block90['B50'] = bebecode
            block90['B51'] = f'''=code128("{com7productcode}")'''
            block90['B52'] = com7productcode
            #e
            block90['E50'] = bebecode
            block90['E51'] = f'''=code128("{com7productcode}")'''
            block90['E52'] = com7productcode
            #h
            block90['H50'] = bebecode
            block90['H51'] = f'''=code128("{com7productcode}")'''
            block90['H52'] = com7productcode
            #k
            block90['K50'] = bebecode
            block90['K51'] = f'''=code128("{com7productcode}")'''
            block90['K52'] = com7productcode
            #n
            block90['N50'] = bebecode
            block90['N51'] = f'''=code128("{com7productcode}")'''
            block90['N52'] = com7productcode            
            #b
            block90['B54'] = bebecode
            block90['B55'] = f'''=code128("{com7productcode}")'''
            block90['B56'] = com7productcode
            #e
            block90['E54'] = bebecode
            block90['E55'] = f'''=code128("{com7productcode}")'''
            block90['E56'] = com7productcode
            #h
            block90['H54'] = bebecode
            block90['H55'] = f'''=code128("{com7productcode}")'''
            block90['H56'] = com7productcode
            #k
            block90['K54'] = bebecode
            block90['K55'] = f'''=code128("{com7productcode}")'''
            block90['K56'] = com7productcode
            #n
            block90['N54'] = bebecode
            block90['N55'] = f'''=code128("{com7productcode}")'''
            block90['N56'] = com7productcode            
            #b
            block90['B58'] = bebecode
            block90['B59'] = f'''=code128("{com7productcode}")'''
            block90['B60'] = com7productcode
            #e
            block90['E58'] = bebecode
            block90['E59'] = f'''=code128("{com7productcode}")'''
            block90['E60'] = com7productcode
            #h
            block90['H58'] = bebecode
            block90['H59'] = f'''=code128("{com7productcode}")'''
            block90['H60'] = com7productcode
            #k
            block90['K58'] = bebecode
            block90['K59'] = f'''=code128("{com7productcode}")'''
            block90['K60'] = com7productcode
            #n
            block90['N58'] = bebecode
            block90['N59'] = f'''=code128("{com7productcode}")'''
            block90['N60'] = com7productcode            
            #b
            block90['B62'] = bebecode
            block90['B63'] = f'''=code128("{com7productcode}")'''
            block90['B64'] = com7productcode
            #e6
            block90['E62'] = bebecode
            block90['E63'] = f'''=code128("{com7productcode}")'''
            block90['E64'] = com7productcode
            #h6
            block90['H62'] = bebecode
            block90['H63'] = f'''=code128("{com7productcode}")'''
            block90['H64'] = com7productcode
            #k6
            block90['K62'] = bebecode
            block90['K63'] = f'''=code128("{com7productcode}")'''
            block90['K64'] = com7productcode
            #n6
            block90['N62'] = bebecode
            block90['N63'] = f'''=code128("{com7productcode}")'''
            block90['N64'] = com7productcode            
            #b
            block90['B66'] = bebecode
            block90['B67'] = f'''=code128("{com7productcode}")'''
            block90['B68'] = com7productcode
            #e
            block90['E66'] = bebecode
            block90['E67'] = f'''=code128("{com7productcode}")'''
            block90['E68'] = com7productcode
            #h
            block90['H66'] = bebecode
            block90['H67'] = f'''=code128("{com7productcode}")'''
            block90['H68'] = com7productcode
            #k
            block90['K66'] = bebecode
            block90['K67'] = f'''=code128("{com7productcode}")'''
            block90['K68'] = com7productcode
            #n
            block90['N66'] = bebecode
            block90['N67'] = f'''=code128("{com7productcode}")'''
            block90['N68'] = com7productcode            
            #b
            block90['B70'] = bebecode
            block90['B71'] = f'''=code128("{com7productcode}")'''
            block90['B72'] = com7productcode
            #e
            block90['E70'] = bebecode
            block90['E71'] = f'''=code128("{com7productcode}")'''
            block90['E72'] = com7productcode
            #h
            block90['H70'] = bebecode
            block90['H71'] = f'''=code128("{com7productcode}")'''
            block90['H72'] = com7productcode
            #k
            block90['K70'] = bebecode
            block90['K71'] = f'''=code128("{com7productcode}")'''
            block90['K72'] = com7productcode
            #n
            block90['N70'] = bebecode
            block90['N71'] = f'''=code128("{com7productcode}")'''
            block90['N72'] = com7productcode
            #for row_cell in range(1,91):
            #    if row_cell == 91:
            #        pass
            #    else:
            #        for bebe_row in range(bebe_start_row_number,74,4):
            #            for com7barcoderow in range(com7barcode_startrow, 75, 4):
            #                for com7barcodetext_row in range(com7barcodetext_startrow, 76, 4):
            #                    if col == 14:
            #                        block90.cell(row=bebe_row, column=col).value = bebecode
            #                        block90.cell(row=com7barcoderow, column=col).value = com7productcode
            #                        block90.cell(row=com7barcodetext_row, column=col).value = com7productcode
            #                        col = 2
            #                        bebe_start_row_number = bebe_row + 4
            #                        com7barcode_startrow = com7barcoderow + 4
            #                        com7barcodetext_startrow = com7barcodetext_row + 4
            #                        print(bebe_start_row_number)
            #                        print(com7barcode_startrow)
            #                        print(com7barcodetext_startrow)
            #                        time.sleep(2)
            #                        break
            #                    else:                            
            #                        block90.cell(row=bebe_row, column=col).value = bebecode
            #                        block90.cell(row=com7barcoderow, column=col).value = com7productcode
            #                        block90.cell(row=com7barcodetext_row, column=col).value = com7productcode
            #                        col += 3
            #                        break
            #                    break
            #                break
            #            break
            block90_wb.save(f"{bebecode}.xlsm")

                
           
            
                
                
      
        elif total_print_data == 1:
            os.chdir(r"D:\Workstuff\my-work-python-script\project_bebephone\ปริ้น 1 แผ่น")
            #b
            block90['B2'] = bebecode
            block90['B3'] = f'''=code128("{com7productcode}")'''
            block90['B4'] = com7productcode
            #e
            block90['E2'] = bebecode
            block90['E3'] = f'''=code128("{com7productcode}")'''
            block90['E4'] = com7productcode
            #h
            block90['H2'] = bebecode
            block90['H3'] = f'''=code128("{com7productcode}")'''
            block90['H4'] = com7productcode
            #k
            block90['K2'] = bebecode
            block90['K3'] = f'''=code128("{com7productcode}")'''
            block90['K4'] = com7productcode
            #n
            block90['N2'] = bebecode
            block90['N3'] = f'''=code128("{com7productcode}")'''
            block90['N4'] = com7productcode
            #b
            block90['B6'] = bebecode
            block90['B7'] = f'''=code128("{com7productcode}")'''
            block90['B8'] = com7productcode
            #e
            block90['E6'] = bebecode
            block90['E7'] = f'''=code128("{com7productcode}")'''
            block90['E8'] = com7productcode
            #h
            block90['H6'] = bebecode
            block90['H7'] = f'''=code128("{com7productcode}")'''
            block90['H8'] = com7productcode
            #k
            block90['K6'] = bebecode
            block90['K7'] = f'''=code128("{com7productcode}")'''
            block90['K8'] = com7productcode
            #n
            block90['N6'] = bebecode
            block90['N7'] = f'''=code128("{com7productcode}")'''
            block90['N8'] = com7productcode            
            #b
            block90['B10'] = bebecode
            block90['B11'] = f'''=code128("{com7productcode}")'''
            block90['B12'] = com7productcode
            #e
            block90['E10'] = bebecode
            block90['E11'] = f'''=code128("{com7productcode}")'''
            block90['E12'] = com7productcode
            #h
            block90['H10'] = bebecode
            block90['H11'] = f'''=code128("{com7productcode}")'''
            block90['H12'] = com7productcode
            #k
            block90['K10'] = bebecode
            block90['K11'] = f'''=code128("{com7productcode}")'''
            block90['K12'] = com7productcode
            #n
            block90['N10'] = bebecode
            block90['N11'] = f'''=code128("{com7productcode}")'''
            block90['N12'] = com7productcode            
            #b
            block90['B14'] = bebecode
            block90['B15'] = f'''=code128("{com7productcode}")'''
            block90['B16'] = com7productcode
            #e
            block90['E14'] = bebecode
            block90['E15'] = f'''=code128("{com7productcode}")'''
            block90['E16'] = com7productcode
            #h
            block90['H14'] = bebecode
            block90['H15'] = f'''=code128("{com7productcode}")'''
            block90['H16'] = com7productcode
            #k
            block90['K14'] = bebecode
            block90['K15'] = f'''=code128("{com7productcode}")'''
            block90['K16'] = com7productcode
            #n
            block90['N14'] = bebecode
            block90['N15'] = f'''=code128("{com7productcode}")'''
            block90['N16'] = com7productcode            
            #b
            block90['B18'] = bebecode
            block90['B19'] = f'''=code128("{com7productcode}")'''
            block90['B20'] = com7productcode
            #e
            block90['E18'] = bebecode
            block90['E19'] = f'''=code128("{com7productcode}")'''
            block90['E20'] = com7productcode
            #h
            block90['H18'] = bebecode
            block90['H19'] = f'''=code128("{com7productcode}")'''
            block90['H20'] = com7productcode
            #k
            block90['K18'] = bebecode
            block90['K19'] = f'''=code128("{com7productcode}")'''
            block90['K20'] = com7productcode
            #n
            block90['N18'] = bebecode
            block90['N19'] = f'''=code128("{com7productcode}")'''
            block90['N20'] = com7productcode            
            #b
            block90['B22'] = bebecode
            block90['B23'] = f'''=code128("{com7productcode}")'''
            block90['B24'] = com7productcode
            #e
            block90['E22'] = bebecode
            block90['E23'] = f'''=code128("{com7productcode}")'''
            block90['E24'] = com7productcode
            #h
            block90['H22'] = bebecode
            block90['H23'] = f'''=code128("{com7productcode}")'''
            block90['H24'] = com7productcode
            #k
            block90['K22'] = bebecode
            block90['K23'] = f'''=code128("{com7productcode}")'''
            block90['K24'] = com7productcode
            #n
            block90['N22'] = bebecode
            block90['N23'] = f'''=code128("{com7productcode}")'''
            block90['N24'] = com7productcode            
            #b
            block90['B26'] = bebecode
            block90['B27'] = f'''=code128("{com7productcode}")'''
            block90['B28'] = com7productcode
            #e
            block90['E26'] = bebecode
            block90['E27'] = f'''=code128("{com7productcode}")'''
            block90['E28'] = com7productcode
            #h
            block90['H26'] = bebecode
            block90['H27'] = f'''=code128("{com7productcode}")'''
            block90['H28'] = com7productcode
            #k
            block90['K26'] = bebecode
            block90['K27'] = f'''=code128("{com7productcode}")'''
            block90['K28'] = com7productcode
            #n
            block90['N26'] = bebecode
            block90['N27'] = f'''=code128("{com7productcode}")'''
            block90['N28'] = com7productcode            
            #b
            block90['B30'] = bebecode
            block90['B31'] = f'''=code128("{com7productcode}")'''
            block90['B32'] = com7productcode
            #e
            block90['E30'] = bebecode
            block90['E31'] = f'''=code128("{com7productcode}")'''
            block90['E32'] = com7productcode
            #h
            block90['H30'] = bebecode
            block90['H31'] = f'''=code128("{com7productcode}")'''
            block90['H32'] = com7productcode
            #k
            block90['K30'] = bebecode
            block90['K31'] = f'''=code128("{com7productcode}")'''
            block90['K32'] = com7productcode
            #n
            block90['N30'] = bebecode
            block90['N31'] = f'''=code128("{com7productcode}")'''
            block90['N32'] = com7productcode            
            #b
            block90['B34'] = bebecode
            block90['B35'] = f'''=code128("{com7productcode}")'''
            block90['B36'] = com7productcode
            #e
            block90['E34'] = bebecode
            block90['E35'] = f'''=code128("{com7productcode}")'''
            block90['E36'] = com7productcode
            #h
            block90['H34'] = bebecode
            block90['H35'] = f'''=code128("{com7productcode}")'''
            block90['H36'] = com7productcode
            #k
            block90['K34'] = bebecode
            block90['K35'] = f'''=code128("{com7productcode}")'''
            block90['K36'] = com7productcode
            #n
            block90['N34'] = bebecode
            block90['N35'] = f'''=code128("{com7productcode}")'''
            block90['N36'] = com7productcode            
            #b
            block90['B38'] = bebecode
            block90['B39'] = f'''=code128("{com7productcode}")'''
            block90['B40'] = com7productcode
            #e
            block90['E38'] = bebecode
            block90['E39'] = f'''=code128("{com7productcode}")'''
            block90['E40'] = com7productcode
            #h
            block90['H38'] = bebecode
            block90['H39'] = f'''=code128("{com7productcode}")'''
            block90['H40'] = com7productcode
            #k
            block90['K38'] = bebecode
            block90['K39'] = f'''=code128("{com7productcode}")'''
            block90['K40'] = com7productcode
            #n
            block90['N38'] = bebecode
            block90['N39'] = f'''=code128("{com7productcode}")'''
            block90['N40'] = com7productcode            
            #b
            block90['B42'] = bebecode
            block90['B43'] = f'''=code128("{com7productcode}")'''
            block90['B44'] = com7productcode
            #e
            block90['E42'] = bebecode
            block90['E43'] = f'''=code128("{com7productcode}")'''
            block90['E44'] = com7productcode
            #h4
            block90['H42'] = bebecode
            block90['H43'] = f'''=code128("{com7productcode}")'''
            block90['H44'] = com7productcode
            #k4
            block90['K42'] = bebecode
            block90['K43'] = f'''=code128("{com7productcode}")'''
            block90['K44'] = com7productcode
            #n4
            block90['N42'] = bebecode
            block90['N43'] = f'''=code128("{com7productcode}")'''
            block90['N44'] = com7productcode            
            #b
            block90['B46'] = bebecode
            block90['B47'] = f'''=code128("{com7productcode}")'''
            block90['B48'] = com7productcode
            #e
            block90['E46'] = bebecode
            block90['E47'] = f'''=code128("{com7productcode}")'''
            block90['E48'] = com7productcode
            #h
            block90['H46'] = bebecode
            block90['H47'] = f'''=code128("{com7productcode}")'''
            block90['H48'] = com7productcode
            #k
            block90['K46'] = bebecode
            block90['K47'] = f'''=code128("{com7productcode}")'''
            block90['K48'] = com7productcode
            #n
            block90['N46'] = bebecode
            block90['N47'] = f'''=code128("{com7productcode}")'''
            block90['N48'] = com7productcode            
            #b
            block90['B50'] = bebecode
            block90['B51'] = f'''=code128("{com7productcode}")'''
            block90['B52'] = com7productcode
            #e
            block90['E50'] = bebecode
            block90['E51'] = f'''=code128("{com7productcode}")'''
            block90['E52'] = com7productcode
            #h
            block90['H50'] = bebecode
            block90['H51'] = f'''=code128("{com7productcode}")'''
            block90['H52'] = com7productcode
            #k
            block90['K50'] = bebecode
            block90['K51'] = f'''=code128("{com7productcode}")'''
            block90['K52'] = com7productcode
            #n
            block90['N50'] = bebecode
            block90['N51'] = f'''=code128("{com7productcode}")'''
            block90['N52'] = com7productcode            
            #b
            block90['B54'] = bebecode
            block90['B55'] = f'''=code128("{com7productcode}")'''
            block90['B56'] = com7productcode
            #e
            block90['E54'] = bebecode
            block90['E55'] = f'''=code128("{com7productcode}")'''
            block90['E56'] = com7productcode
            #h
            block90['H54'] = bebecode
            block90['H55'] = f'''=code128("{com7productcode}")'''
            block90['H56'] = com7productcode
            #k
            block90['K54'] = bebecode
            block90['K55'] = f'''=code128("{com7productcode}")'''
            block90['K56'] = com7productcode
            #n
            block90['N54'] = bebecode
            block90['N55'] = f'''=code128("{com7productcode}")'''
            block90['N56'] = com7productcode            
            #b
            block90['B58'] = bebecode
            block90['B59'] = f'''=code128("{com7productcode}")'''
            block90['B60'] = com7productcode
            #e
            block90['E58'] = bebecode
            block90['E59'] = f'''=code128("{com7productcode}")'''
            block90['E60'] = com7productcode
            #h
            block90['H58'] = bebecode
            block90['H59'] = f'''=code128("{com7productcode}")'''
            block90['H60'] = com7productcode
            #k
            block90['K58'] = bebecode
            block90['K59'] = f'''=code128("{com7productcode}")'''
            block90['K60'] = com7productcode
            #n
            block90['N58'] = bebecode
            block90['N59'] = f'''=code128("{com7productcode}")'''
            block90['N60'] = com7productcode            
            #b
            block90['B62'] = bebecode
            block90['B63'] = f'''=code128("{com7productcode}")'''
            block90['B64'] = com7productcode
            #e6
            block90['E62'] = bebecode
            block90['E63'] = f'''=code128("{com7productcode}")'''
            block90['E64'] = com7productcode
            #h6
            block90['H62'] = bebecode
            block90['H63'] = f'''=code128("{com7productcode}")'''
            block90['H64'] = com7productcode
            #k6
            block90['K62'] = bebecode
            block90['K63'] = f'''=code128("{com7productcode}")'''
            block90['K64'] = com7productcode
            #n6
            block90['N62'] = bebecode
            block90['N63'] = f'''=code128("{com7productcode}")'''
            block90['N64'] = com7productcode            
            #b
            block90['B66'] = bebecode
            block90['B67'] = f'''=code128("{com7productcode}")'''
            block90['B68'] = com7productcode
            #e
            block90['E66'] = bebecode
            block90['E67'] = f'''=code128("{com7productcode}")'''
            block90['E68'] = com7productcode
            #h
            block90['H66'] = bebecode
            block90['H67'] = f'''=code128("{com7productcode}")'''
            block90['H68'] = com7productcode
            #k
            block90['K66'] = bebecode
            block90['K67'] = f'''=code128("{com7productcode}")'''
            block90['K68'] = com7productcode
            #n
            block90['N66'] = bebecode
            block90['N67'] = f'''=code128("{com7productcode}")'''
            block90['N68'] = com7productcode            
            #b
            block90['B70'] = bebecode
            block90['B71'] = f'''=code128("{com7productcode}")'''
            block90['B72'] = com7productcode
            #e
            block90['E70'] = bebecode
            block90['E71'] = f'''=code128("{com7productcode}")'''
            block90['E72'] = com7productcode
            #h
            block90['H70'] = bebecode
            block90['H71'] = f'''=code128("{com7productcode}")'''
            block90['H72'] = com7productcode
            #k
            block90['K70'] = bebecode
            block90['K71'] = f'''=code128("{com7productcode}")'''
            block90['K72'] = com7productcode
            #n
            block90['N70'] = bebecode
            block90['N71'] = f'''=code128("{com7productcode}")'''
            block90['N72'] = com7productcode
            block90_wb.save(f"{bebecode}.xlsm")


def movefiletofolders():
    def grouper(S, n): #https://stackoverflow.com/questions/12559055/for-every-x-number-of-files-create-new-directory-and-move-files-using-python
        iterator = iter(S)
        while True:
            item = list(itertools.islice(iterator, n))
            print(item)
            if len(item) == 0:
                break
            yield item
    os.chdir(r"D:\Workstuff\my-work-python-script\project_bebephone\ปริ้น 1 แผ่น")
    fnames = ns.natsorted(glob.glob('*.xlsm'))
    for i, fnames in enumerate(grouper(fnames, 53)):
        dirname = f"{i + 1}"
        try:
            os.mkdir(dirname)
        except FileExistsError:
            pass
        for fname in fnames:
            shutil.move(fname, dirname)
            with open(f"filelist {i + 1}.txt", mode="a") as textfile:
                textfile.writelines(f"{fname}\n")
        shutil.move(f"filelist {i + 1}.txt", dirname)
        

if __name__ in "__main__"  :
    main()
    movefiletofolders()
     
