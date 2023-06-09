import os, glob, shutil
import openpyxl
from tkinter import Tk, filedialog
import itertools
import natsort as ns

asset_directory = os.path.join(r"D:\Workstuff\my-work-python-script\project_bebephone", r"Asset")

root = Tk()

#get asset
PRINT_dir = r"D:\Workstuff\my-work-python-script\project_bebephone\PRINT"
asset_dir_117 = os.path.join(asset_directory, r"block117template.xlsm")
block117_wb = openpyxl.load_workbook(asset_dir_117, data_only=True, keep_vba=True)
block117_ws = block117_wb.sheetnames
block117 = block117_wb[block117_ws[0]]
#get data
get_wb = filedialog.askopenfilename(title="เลือกไฟล์ข้อมูล Excel", filetypes=([("*Excel Files", "*.xlsx"), ("All Files", "*.*")]))
data_wb = openpyxl.load_workbook(get_wb, data_only=True)
data_sheet = data_wb.sheetnames
get_sheet = data_wb[data_sheet[0]]
def main():

        #bebe row
    bebe_start_row_number = 2

    #barcode row
    com7barcode_startrow = 3

    #barcode text
    com7barcodetext_startrow = 4

    #data
    data_startrow = 3

    ran = False
    col = 2
    folder_count = 1

    for total_print_row in range(3,get_sheet.max_row+1):
        bebecode = get_sheet.cell(row=total_print_row,column = 1).value #0166
        com7productcode = get_sheet.cell(row=total_print_row, column = 2).value #51691
        total_sticker = get_sheet.cell(row=total_print_row, column=3).value
        total_print_data = get_sheet.cell(row=total_print_row, column = 4).value #8 #int
        try:
            os.chdir(r"D:\Workstuff\my-work-python-script\project_bebephone")
            os.mkdir("Block 117_Manual")
            os.chdir(r"D:\Workstuff\my-work-python-script\project_bebephone\Block 117_Manual")
        except FileExistsError:
            pass
        os.chdir(r"D:\Workstuff\my-work-python-script\project_bebephone\Block 117_Manual")
        #1
        #b
        block117['B2'] = bebecode
        block117['B3'] = f'''=code128("{com7productcode}")'''
        block117['B4'] = com7productcode
                #c
        block117['C2'] = bebecode
        block117['C3'] = f'''=code128("{com7productcode}")'''
        block117['C4'] = com7productcode
        #d
        block117['D2'] = bebecode
        block117['D3'] = f'''=code128("{com7productcode}")'''
        block117['D4'] = com7productcode
                #b
        block117['E2'] = bebecode
        block117['E3'] = f'''=code128("{com7productcode}")'''
        block117['E4'] = com7productcode
        #f
        block117['F2'] = bebecode
        block117['F3'] = f'''=code128("{com7productcode}")'''
        block117['F4'] = com7productcode 
    
        #g
        block117['G2'] = bebecode
        block117['G3'] = f'''=code128("{com7productcode}")'''
        block117['G4'] = com7productcode
        #h
        block117['H2'] = bebecode
        block117['H3'] = f'''=code128("{com7productcode}")'''
        block117['H4'] = com7productcode 
        #i
        block117['I2'] = bebecode
        block117['I3'] = f'''=code128("{com7productcode}")'''
        block117['I4'] = com7productcode 
        #j
        block117['J2'] = bebecode
        block117['J3'] = f'''=code128("{com7productcode}")'''
        block117['J4'] = com7productcode 

        ######
                #1
        #b
        block117['B6'] = bebecode
        block117['B7'] = f'''=code128("{com7productcode}")'''
        block117['B8'] = com7productcode
                #c
        block117['C6'] = bebecode
        block117['C7'] = f'''=code128("{com7productcode}")'''
        block117['C8'] = com7productcode
        #d
        block117['D6'] = bebecode
        block117['D7'] = f'''=code128("{com7productcode}")'''
        block117['D8'] = com7productcode
                #b
        block117['E6'] = bebecode
        block117['E7'] = f'''=code128("{com7productcode}")'''
        block117['E8'] = com7productcode
        #f
        block117['F6'] = bebecode
        block117['F7'] = f'''=code128("{com7productcode}")'''
        block117['F8'] = com7productcode 
    
        #g
        block117['G6'] = bebecode
        block117['G7'] = f'''=code128("{com7productcode}")'''
        block117['G8'] = com7productcode
        #h
        block117['H6'] = bebecode
        block117['H7'] = f'''=code128("{com7productcode}")'''
        block117['H8'] = com7productcode 
        #i
        block117['I6'] = bebecode
        block117['I7'] = f'''=code128("{com7productcode}")'''
        block117['I8'] = com7productcode 
        #j
        block117['J6'] = bebecode
        block117['J7'] = f'''=code128("{com7productcode}")'''
        block117['J8'] = com7productcode 


                #1
        #b
        block117['B10'] = bebecode
        block117['B11'] = f'''=code128("{com7productcode}")'''
        block117['B12'] = com7productcode
                #c
        block117['C10'] = bebecode
        block117['C11'] = f'''=code128("{com7productcode}")'''
        block117['C12'] = com7productcode
        #d
        block117['D10'] = bebecode
        block117['D11'] = f'''=code128("{com7productcode}")'''
        block117['D12'] = com7productcode
                #b
        block117['E10'] = bebecode
        block117['E11'] = f'''=code128("{com7productcode}")'''
        block117['E12'] = com7productcode
        #f
        block117['F10'] = bebecode
        block117['F11'] = f'''=code128("{com7productcode}")'''
        block117['F12'] = com7productcode 
        
        #g
        block117['G10'] = bebecode
        block117['G11'] = f'''=code128("{com7productcode}")'''
        block117['G12'] = com7productcode
        #h
        block117['H10'] = bebecode
        block117['H11'] = f'''=code128("{com7productcode}")'''
        block117['H12'] = com7productcode 
        #i
        block117['I10'] = bebecode
        block117['I11'] = f'''=code128("{com7productcode}")'''
        block117['I12'] = com7productcode 
        #j
        block117['J10'] = bebecode
        block117['J11'] = f'''=code128("{com7productcode}")'''
        block117['J12'] = com7productcode 


                #1
        #b
        block117['B14'] = bebecode
        block117['B15'] = f'''=code128("{com7productcode}")'''
        block117['B16'] = com7productcode
                #c
        block117['C14'] = bebecode
        block117['C15'] = f'''=code128("{com7productcode}")'''
        block117['C16'] = com7productcode
        #d
        block117['D14'] = bebecode
        block117['D15'] = f'''=code128("{com7productcode}")'''
        block117['D16'] = com7productcode
                #b
        block117['E14'] = bebecode
        block117['E15'] = f'''=code128("{com7productcode}")'''
        block117['E16'] = com7productcode
        #
        block117['F14'] = bebecode
        block117['F15'] = f'''=code128("{com7productcode}")'''
        block117['F16'] = com7productcode 
        #g
        block117['G14'] = bebecode
        block117['G15'] = f'''=code128("{com7productcode}")'''
        block117['G16'] = com7productcode
        #h
        block117['H14'] = bebecode
        block117['H15'] = f'''=code128("{com7productcode}")'''
        block117['H16'] = com7productcode 
        #i
        block117['I14'] = bebecode
        block117['I15'] = f'''=code128("{com7productcode}")'''
        block117['I16'] = com7productcode 
        #j
        block117['J14'] = bebecode
        block117['J15'] = f'''=code128("{com7productcode}")'''
        block117['J16'] = com7productcode 


                #1
        #b
        block117['B18'] = bebecode
        block117['B19'] = f'''=code128("{com7productcode}")'''
        block117['B20'] = com7productcode
                #c
        block117['C18'] = bebecode
        block117['C19'] = f'''=code128("{com7productcode}")'''
        block117['C20'] = com7productcode
        #d
        block117['D18'] = bebecode
        block117['D19'] = f'''=code128("{com7productcode}")'''
        block117['D20'] = com7productcode
                #b
        block117['E18'] = bebecode
        block117['E19'] = f'''=code128("{com7productcode}")'''
        block117['E20'] = com7productcode
        #f
        block117['F18'] = bebecode
        block117['F19'] = f'''=code128("{com7productcode}")'''
        block117['F20'] = com7productcode 
        
        #g
        block117['G18'] = bebecode
        block117['G19'] = f'''=code128("{com7productcode}")'''
        block117['G20'] = com7productcode
        #h
        block117['H18'] = bebecode
        block117['H19'] = f'''=code128("{com7productcode}")'''
        block117['H20'] = com7productcode 
        #i
        block117['I18'] = bebecode
        block117['I19'] = f'''=code128("{com7productcode}")'''
        block117['I20'] = com7productcode 
        #j
        block117['J18'] = bebecode
        block117['J19'] = f'''=code128("{com7productcode}")'''
        block117['J20'] = com7productcode 


                #1
        #b
        block117['B22'] = bebecode
        block117['B23'] = f'''=code128("{com7productcode}")'''
        block117['B24'] = com7productcode
                #c
        block117['C22'] = bebecode
        block117['C23'] = f'''=code128("{com7productcode}")'''
        block117['C24'] = com7productcode
        #d
        block117['D22'] = bebecode
        block117['D23'] = f'''=code128("{com7productcode}")'''
        block117['D24'] = com7productcode
                #b
        block117['E22'] = bebecode
        block117['E23'] = f'''=code128("{com7productcode}")'''
        block117['E24'] = com7productcode
        #f
        block117['F22'] = bebecode
        block117['F23'] = f'''=code128("{com7productcode}")'''
        block117['F24'] = com7productcode 
        
        #g
        block117['G22'] = bebecode
        block117['G23'] = f'''=code128("{com7productcode}")'''
        block117['G24'] = com7productcode
        #h
        block117['H22'] = bebecode
        block117['H23'] = f'''=code128("{com7productcode}")'''
        block117['H24'] = com7productcode 
        #i
        block117['I22'] = bebecode
        block117['I23'] = f'''=code128("{com7productcode}")'''
        block117['I24'] = com7productcode 
        #j
        block117['J22'] = bebecode
        block117['J23'] = f'''=code128("{com7productcode}")'''
        block117['J24'] = com7productcode 


                #1
        #b
        block117['B26'] = bebecode
        block117['B27'] = f'''=code128("{com7productcode}")'''
        block117['B28'] = com7productcode
                #c
        block117['C26'] = bebecode
        block117['C27'] = f'''=code128("{com7productcode}")'''
        block117['C28'] = com7productcode
        #d
        block117['D26'] = bebecode
        block117['D27'] = f'''=code128("{com7productcode}")'''
        block117['D28'] = com7productcode
                #b
        block117['E26'] = bebecode
        block117['E27'] = f'''=code128("{com7productcode}")'''
        block117['E28'] = com7productcode
        #f
        block117['F26'] = bebecode
        block117['F27'] = f'''=code128("{com7productcode}")'''
        block117['F28'] = com7productcode 
 
        #g
        block117['G26'] = bebecode
        block117['G27'] = f'''=code128("{com7productcode}")'''
        block117['G28'] = com7productcode
        #h
        block117['H26'] = bebecode
        block117['H27'] = f'''=code128("{com7productcode}")'''
        block117['H28'] = com7productcode 
        #i
        block117['I26'] = bebecode
        block117['I27'] = f'''=code128("{com7productcode}")'''
        block117['I28'] = com7productcode 
        #j
        block117['J26'] = bebecode
        block117['J27'] = f'''=code128("{com7productcode}")'''
        block117['J28'] = com7productcode 


                #1
        #b
        block117['B30'] = bebecode
        block117['B31'] = f'''=code128("{com7productcode}")'''
        block117['B32'] = com7productcode
                #c
        block117['C30'] = bebecode
        block117['C31'] = f'''=code128("{com7productcode}")'''
        block117['C32'] = com7productcode
        #d
        block117['D30'] = bebecode
        block117['D31'] = f'''=code128("{com7productcode}")'''
        block117['D32'] = com7productcode
                #b
        block117['E30'] = bebecode
        block117['E31'] = f'''=code128("{com7productcode}")'''
        block117['E32'] = com7productcode
        #f
        block117['F30'] = bebecode
        block117['F31'] = f'''=code128("{com7productcode}")'''
        block117['F32'] = com7productcode 
        
        #g
        block117['G30'] = bebecode
        block117['G31'] = f'''=code128("{com7productcode}")'''
        block117['G32'] = com7productcode
        #h
        block117['H30'] = bebecode
        block117['H31'] = f'''=code128("{com7productcode}")'''
        block117['H32'] = com7productcode 
        #i
        block117['I30'] = bebecode
        block117['I31'] = f'''=code128("{com7productcode}")'''
        block117['I32'] = com7productcode 
        #j
        block117['J30'] = bebecode
        block117['J31'] = f'''=code128("{com7productcode}")'''
        block117['J32'] = com7productcode 


                #1
        #b
        block117['B34'] = bebecode
        block117['B35'] = f'''=code128("{com7productcode}")'''
        block117['B36'] = com7productcode
                #c
        block117['C34'] = bebecode
        block117['C35'] = f'''=code128("{com7productcode}")'''
        block117['C36'] = com7productcode
        #d
        block117['D34'] = bebecode
        block117['D35'] = f'''=code128("{com7productcode}")'''
        block117['D36'] = com7productcode
                #b
        block117['E34'] = bebecode
        block117['E35'] = f'''=code128("{com7productcode}")'''
        block117['E36'] = com7productcode
        #f
        block117['F34'] = bebecode
        block117['F35'] = f'''=code128("{com7productcode}")'''
        block117['F36'] = com7productcode 
        
        #g
        block117['G34'] = bebecode
        block117['G35'] = f'''=code128("{com7productcode}")'''
        block117['G36'] = com7productcode
        #h
        block117['H34'] = bebecode
        block117['H35'] = f'''=code128("{com7productcode}")'''
        block117['H36'] = com7productcode 
        #i
        block117['I34'] = bebecode
        block117['I35'] = f'''=code128("{com7productcode}")'''
        block117['I36'] = com7productcode 
        #j
        block117['J34'] = bebecode
        block117['J35'] = f'''=code128("{com7productcode}")'''
        block117['J36'] = com7productcode 


                #1
        #b
        block117['B38'] = bebecode
        block117['B39'] = f'''=code128("{com7productcode}")'''
        block117['B40'] = com7productcode
                #c
        block117['C38'] = bebecode
        block117['C39'] = f'''=code128("{com7productcode}")'''
        block117['C40'] = com7productcode
        #d
        block117['D38'] = bebecode
        block117['D39'] = f'''=code128("{com7productcode}")'''
        block117['D40'] = com7productcode
                #b
        block117['E38'] = bebecode
        block117['E39'] = f'''=code128("{com7productcode}")'''
        block117['E40'] = com7productcode
        #f
        block117['F38'] = bebecode
        block117['F39'] = f'''=code128("{com7productcode}")'''
        block117['F40'] = com7productcode 
        
        #g
        block117['G38'] = bebecode
        block117['G39'] = f'''=code128("{com7productcode}")'''
        block117['G40'] = com7productcode
        #h
        block117['H38'] = bebecode
        block117['H39'] = f'''=code128("{com7productcode}")'''
        block117['H40'] = com7productcode 
        #i
        block117['I38'] = bebecode
        block117['I39'] = f'''=code128("{com7productcode}")'''
        block117['I40'] = com7productcode 
        #j
        block117['J38'] = bebecode
        block117['J39'] = f'''=code128("{com7productcode}")'''
        block117['J40'] = com7productcode 


                #1
        #b
        block117['B42'] = bebecode
        block117['B43'] = f'''=code128("{com7productcode}")'''
        block117['B44'] = com7productcode
                #c
        block117['C42'] = bebecode
        block117['C43'] = f'''=code128("{com7productcode}")'''
        block117['C44'] = com7productcode
        #d
        block117['D42'] = bebecode
        block117['D43'] = f'''=code128("{com7productcode}")'''
        block117['D44'] = com7productcode
                #b
        block117['E42'] = bebecode
        block117['E43'] = f'''=code128("{com7productcode}")'''
        block117['E44'] = com7productcode
        #f
        block117['F42'] = bebecode
        block117['F43'] = f'''=code128("{com7productcode}")'''
        block117['F44'] = com7productcode 
        
        #g
        block117['G42'] = bebecode
        block117['G43'] = f'''=code128("{com7productcode}")'''
        block117['G44'] = com7productcode
        #h
        block117['H42'] = bebecode
        block117['H43'] = f'''=code128("{com7productcode}")'''
        block117['H44'] = com7productcode 
        #i
        block117['I42'] = bebecode
        block117['I43'] = f'''=code128("{com7productcode}")'''
        block117['I44'] = com7productcode 
        #j
        block117['J42'] = bebecode
        block117['J43'] = f'''=code128("{com7productcode}")'''
        block117['J44'] = com7productcode 


                #1
        #b
        block117['B46'] = bebecode
        block117['B47'] = f'''=code128("{com7productcode}")'''
        block117['B48'] = com7productcode
                #c
        block117['C46'] = bebecode
        block117['C47'] = f'''=code128("{com7productcode}")'''
        block117['C48'] = com7productcode
        #d
        block117['D46'] = bebecode
        block117['D47'] = f'''=code128("{com7productcode}")'''
        block117['D48'] = com7productcode
                #b
        block117['E46'] = bebecode
        block117['E47'] = f'''=code128("{com7productcode}")'''
        block117['E48'] = com7productcode
        #f
        block117['F46'] = bebecode
        block117['F47'] = f'''=code128("{com7productcode}")'''
        block117['F48'] = com7productcode 
        
        #g
        block117['G46'] = bebecode
        block117['G47'] = f'''=code128("{com7productcode}")'''
        block117['G48'] = com7productcode
        #h
        block117['H46'] = bebecode
        block117['H47'] = f'''=code128("{com7productcode}")'''
        block117['H48'] = com7productcode 
        #i
        block117['I46'] = bebecode
        block117['I47'] = f'''=code128("{com7productcode}")'''
        block117['I48'] = com7productcode 
        #j
        block117['J46'] = bebecode
        block117['J47'] = f'''=code128("{com7productcode}")'''
        block117['J48'] = com7productcode 


                #1
        #b
        block117['B50'] = bebecode
        block117['B51'] = f'''=code128("{com7productcode}")'''
        block117['B52'] = com7productcode
                #c
        block117['C50'] = bebecode
        block117['C51'] = f'''=code128("{com7productcode}")'''
        block117['C52'] = com7productcode
        #d
        block117['D50'] = bebecode
        block117['D51'] = f'''=code128("{com7productcode}")'''
        block117['D52'] = com7productcode
                #b
        block117['E50'] = bebecode
        block117['E51'] = f'''=code128("{com7productcode}")'''
        block117['E52'] = com7productcode
        #f
        block117['F50'] = bebecode
        block117['F51'] = f'''=code128("{com7productcode}")'''
        block117['F52'] = com7productcode 
        
        #g
        block117['G50'] = bebecode
        block117['G51'] = f'''=code128("{com7productcode}")'''
        block117['G52'] = com7productcode
        #h
        block117['H50'] = bebecode
        block117['H51'] = f'''=code128("{com7productcode}")'''
        block117['H52'] = com7productcode 
        #i
        block117['I50'] = bebecode
        block117['I51'] = f'''=code128("{com7productcode}")'''
        block117['I52'] = com7productcode 
        #j
        block117['J50'] = bebecode
        block117['J51'] = f'''=code128("{com7productcode}")'''
        block117['J52'] = com7productcode 
        
        block117_wb.save(f"{bebecode} - {total_print_data} แผ่น_{total_sticker}.xlsm")


def movefiletofolders():
    def grouper(S, n): #https://stackoverflow.com/questions/12559055/for-every-x-number-of-files-create-new-directory-and-move-files-using-python
        iterator = iter(S)
        while True:
            item = list(itertools.islice(iterator, n))
            print(item)
            if len(item) == 0:
                break
            yield item
    os.chdir(r"D:\Workstuff\my-work-python-script\project_bebephone\Block 117")
    fnames = ns.natsorted(glob.glob('*.xlsm'))
    for i, fnames in enumerate(grouper(fnames, 53)):
        dirname = f"{i + 1}"
        try:
            os.mkdir(dirname)
        except FileExistsError:
            pass
        for fname in fnames:
            shutil.move(fname, dirname)
            with open(f"filelist {i + 1}.txt", mode="a") as textfile:
                textfile.writelines(f"{fname}\n")
        shutil.move(f"filelist {i + 1}.txt", dirname)
        

if __name__ in "__main__"  :
    main()
    movefiletofolders()