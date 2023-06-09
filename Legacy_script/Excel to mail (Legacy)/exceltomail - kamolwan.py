import win32com.client
from cProfile import run
from fileinput import filename
from tkinter import filedialog
from tkinter import *
import openpyxl
import os
import datetime as dt

root1 = Tk()
root1.filedir = filedialog.askdirectory(title="เลือก Folder ที่เก็บไฟล์ PDF")
root1.filename = filedialog.askopenfilename(initialdir="/",title="เลือกไฟล์ Excel",filetypes=(("Excel","*.xlsx"),("All files","*.*")))

workbook = openpyxl.load_workbook(root1.filename, data_only=True)
sheets = workbook.sheetnames
readsheet = workbook[sheets[1]]

def sentmail(): #dummy formula
    os.chdir(root1.filedir)
    attlist = []
    row = 2 #skip first row because it's header
    while True: #not sure why while work but ok
        value1 = readsheet.cell(row=row,column=3).value #first, get value to use as your header
        email2 = readsheet.cell(row=row,column=4).value #second, get email row
        branch3 = readsheet.cell(row=row,column=2).value #Branch ID
        amouth4 = readsheet.cell(row=row,column=5).value #amouth
        # amouth4 = readsheet.cell(row=row,column=9).value #amouth of bill
        outlook = win32com.client.Dispatch("Outlook.Application") #Call outlook
        From = None
        for myEmailAddress in outlook.Session.Accounts:
            if "Kamolwan@comseven.com" in str(myEmailAddress):
                From = myEmailAddress
                break
        j = range(1,row) #thank you stackoverflow 
        mail = outlook.CreateItem(0)
        mail.Subject = 'Print Label DHL สินค้าส่งซ่อม Code 166 Ship To แผนก Service / รอบวันที่ ' + str(dt.datetime.today().strftime("%d/%m/%Y")) + ' / PHYID' + str(value1)
        mail.To = str(email2)
        mail.Cc = 'Saknarin.W <saknarin.w@comseven.com>; Store_Support_Group <store_support_group@comseven.com>'
        html = '''
        <html> \
        <body> \
            <p><strong>หลังบิลนี้ไป รบกวนขอความร่วมมือ ให้เริ่มใช้วิธีด้านล่าง ตั้งแต่บิลต่อไปนะคะ</strong></p>
<p><strong><span style='font-size:21px;font-family:"Leelawadee","sans-serif";color:blue;'>แจ้งเปลี่ยน ขั้นตอนการส่งซ่อมสินค้า</span></strong><strong><span style='font-size:21px;font-family:"Leelawadee","sans-serif";color:blue;'>&nbsp; ในระบบ &nbsp;ITECINSURANCE &nbsp;หรือ ITEC SERVICE ตามเมล์ด้านล่างนี้</span></strong></p>
<p style='margin-right:0cm;margin-left:0cm;font-size:19px;font-family:"Angsana New","serif";margin:0cm;margin-bottom:.0001pt;'><strong><u><span style='font-family:"Leelawadee","sans-serif";color:blue;background:yellow;'>** ขั้นตอนการเปิดบิลรับซ่อม</span></u></strong><strong><span style='font-family:"Leelawadee","sans-serif";color:blue;'>&nbsp;</span></strong><strong><span style='font-family:"Leelawadee","sans-serif";color:black;'><br>&nbsp;<br>&nbsp;</span></strong><strong><span style='font-family:"Leelawadee","sans-serif";color:blue;'>ในการเปิดบิลรับซ่อม จากเดิม ผู้ซื้อ เป็น&nbsp;</span></strong><strong><span style='font-family:"Leelawadee","sans-serif";color:blue;'>CASH ให้เปลี่ยนเป็น &nbsp;<u><span style="background:yellow;">RMA0004</span></u>&nbsp; (ลูกค้าส่งซ่อมที่สาขา</span></strong><span style='font-family:"Leelawadee","sans-serif";color:blue;'>)</span><span style='font-family:"Leelawadee","sans-serif";color:black;'>&nbsp;</span></p>
<p style='margin-right:0cm;margin-left:0cm;font-size:19px;font-family:"Angsana New","serif";margin:0cm;margin-bottom:.0001pt;'><strong><span style='font-family:"Leelawadee","sans-serif";color:blue;'>กรณีสินค้าหมดประกัน ลูกค้าต้องการส่งเคลม&nbsp;</span></strong><strong><span style='font-family:"Leelawadee","sans-serif";color:blue;font-weight:normal;'>ผู้ซื้อ&nbsp;</span></strong><strong><span style='font-family:"Leelawadee","sans-serif";color:blue;'>ให้ใช้&nbsp;</span></strong><strong><u><span style='font-family:"Leelawadee","sans-serif";color:blue;background:yellow;'>RMA0001</span></u></strong><strong><span style='font-family:"Leelawadee","sans-serif";color:blue;'>&nbsp;สินค้าฝากซ่อม มีค่าบริการ</span></strong></p>
<p style='margin-right:0cm;margin-left:0cm;font-size:19px;font-family:"Angsana New","serif";margin:0cm;margin-bottom:.0001pt;'><strong><span style='font-family:"Leelawadee","sans-serif";color:blue;'><img src="https://i.imgur.com/5IIsEbo.jpg"></span></strong><br></p>
<p style='margin-right:0cm;margin-left:0cm;font-size:19px;font-family:"Angsana New","serif";margin:0cm;margin-bottom:.0001pt;'><span style='font-family:"Leelawadee","sans-serif";color:black;'>&nbsp;&nbsp;</span><strong><u><span style='font-family:"Leelawadee","sans-serif";color:blue;background:yellow;'>** ขั้นตอนในการส่งซ่อม</span></u></strong><span style='font-family:"Leelawadee","sans-serif";color:black;'><br>&nbsp;</span><span style='font-family:"Leelawadee","sans-serif";color:black;'><br>&nbsp;</span><strong><span style='font-family:"Leelawadee","sans-serif";color:blue;'>สินค้าที่ทำส่งซ่อมใน&nbsp;</span></strong><strong><span style='font-family:"Leelawadee","sans-serif";color:blue;'>CODE : 166 ให้เปลี่ยนวิธีเป็นการ <u><span style="background:yellow;">Stock Out สินค้าเสีย ไปที่ 33</span></u> *ทำใน ITEC INSURANCE /FC - ITEC SERVICE (ทำเหมือน สินค้า Trade in</span></strong><span style='font-family:"Leelawadee","sans-serif";color:black;'>) <strong><span style="background:yellow;">ไม่ต้องขอ&nbsp;</span></strong></span><strong><span style='font-family:"Leelawadee","sans-serif";color:black;background:yellow;'>Get CODE</span></strong><strong><span style='font-size:21px;font-family:"Leelawadee","sans-serif";'><br></span></strong></p>
<p style='margin-right:0cm;margin-left:0cm;font-size:19px;font-family:"Angsana New","serif";margin:0cm;margin-bottom:.0001pt;'><strong><span style='font-family:"Leelawadee","sans-serif";color:blue;'><img src="https://i.imgur.com/lGmI5wo.jpg"></span></strong><br></p>
<p style='margin-right:0cm;margin-left:0cm;font-size:19px;font-family:"Angsana New","serif";margin:0cm;margin-bottom:.0001pt;'><strong><span style='font-family:"Leelawadee","sans-serif";color:blue;'><img src="https://i.imgur.com/8HsLQgO.jpg"></span></strong><br></p>
<p style='margin-right:0cm;margin-left:0cm;font-size:19px;font-family:"Angsana New","serif";margin:0cm;margin-bottom:.0001pt;'><strong><span style='font-size:21px;font-family:"Leelawadee","sans-serif";color:blue;'>&nbsp;&nbsp;</span></strong><strong><span style='font-size:21px;font-family:"Leelawadee","sans-serif";color:blue;'>**สอบถามเพิ่มเติม ขั้นตอนการเปลี่ยนแปลง การส่งซ่อมสินค้า**</span></strong></p>
<p style='margin-right:0cm;margin-left:0cm;font-size:19px;font-family:"Angsana New","serif";margin:0cm;margin-bottom:.0001pt;'><strong><span style='font-size:21px;font-family:"Leelawadee","sans-serif";color:blue;'>Line Q&amp;A Service : &nbsp;</span></strong><strong><span style='font-size:21px;font-family:"Leelawadee","sans-serif";color:blue;background:yellow;'><a href="https://page.line.me/533aagih"><strong><span style='font-family:"Leelawadee","sans-serif";color:blue;'>https://page.line.me/533aagih</span></strong></a></span></strong></p>
<p style='font-size: 19px; font-family: "Angsana New", "serif"; margin: 0cm 0cm 0.0001pt; text-align: center;'>+++++++++++++++++++++++++++++++++++++++</p>

<p style='margin:0cm;margin-bottom:.0001pt;font-size:15px;font-family:"Tahoma","sans-serif";'><img width="228" src="https://www.bb-talkin.eu/bbe/wp-content/uploads/2018/12/DHL-parcel-logo-250x200.jpg"></p>

        <p style='margin:0cm;margin-bottom:.0001pt;font-size:19px;font-family:"Tahoma","serif";'><span style="font-size:16px;color:#212121;">หน้าร้าน ID : ''' + str(branch3) + '''</span></p>
        <p style='margin-right:0cm;margin-left:0cm;font-size:19px;font-family:"Angsana New","serif";margin:0cm;margin-bottom:.0001pt;'><br></p>
        <p style='margin-right:0cm;margin-left:0cm;font-size:19px;font-family:"Angsana New","serif";margin:0cm;margin-bottom:.0001pt;'><span style='font-size:15px;font-family:"Tahoma","sans-serif";'>&nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp;&nbsp;</span><span style='font-size:19px;font-family:"Cordia New","sans-serif";'>แจ้งให้</span><span style='font-size:19px;font-family:"Cordia New","sans-serif";'>&nbsp;</span><span style='font-size:15px;font-family:"Tahoma","sans-serif";'>Print Label&nbsp;</span><span style='font-size:29px;font-family:"Cordia New","sans-serif";'>สินค้า</span><span style='font-size:24px;font-family:"Cordia New","sans-serif";color:#1F497D;'>ส่งซ่อม&nbsp;</span><span style='font-size:19px;font-family:"Cordia New","sans-serif";'>&nbsp;<span style="color:#1F497D;">(&nbsp;</span></span><u><span style='font-size:24px;font-family:"Cordia New","sans-serif";color:#1F497D;'>ปลายทาง แผนก&nbsp;</span></u><u><span style='font-size:24px;font-family:"Cordia New","sans-serif";color:#1F497D;'>Service (โปรแกรม&nbsp;ITEC Service -</span></u><u><span style='font-size:35px;font-family:"Cordia New","sans-serif";color:#1F497D;'>&nbsp;Insure</span></u><u><span style='font-size:24px;font-family:"Cordia New","sans-serif";color:#1F497D;'>&nbsp;)</span></u><span style='font-size:19px;font-family:"Cordia New","sans-serif";'>&nbsp;<span style="background:yellow;">เลขที่ บิล</span><span style="background:yellow;">&nbsp;</span></span><span style='font-size:15px;font-family:"Tahoma","sans-serif";background:yellow;'>PHY ID: ''' + str(value1) + ''' &nbsp;:</span><span style='font-size:15px;font-family:"Tahoma","sans-serif";'>&nbsp; &nbsp; &nbsp;&nbsp;</span><strong><u><span style='font-size:24px;font-family:"Cordia New","sans-serif";color:red;'>จำนวน&nbsp;</span></u></strong><strong><u><span style='font-size:24px;font-family:"Cordia New","sans-serif";color:red;'>1 &nbsp;กล่อง &nbsp;&nbsp;</span></u></strong><strong><u><span style='font-size:24px;font-family:"Tahoma","sans-serif";color:red;'>/ '''  + str(amouth4) + '''&nbsp;&nbsp;</span></u></strong><strong><u><span style='font-size:24px;font-family:"Cordia New","sans-serif";color:red;'>บิล</span></u></strong><span style='font-size:19px;font-family:"Cordia New","sans-serif";color:red;'>&nbsp;&nbsp;</span><span style='font-size:15px;font-family:"Tahoma","sans-serif";'>*** <u>Print Label&nbsp;</u></span><u><span style='font-size:19px;font-family:"Cordia New","sans-serif";'>ตาม</span></u><u><span style='font-size:19px;font-family:"Cordia New","sans-serif";'>&nbsp;</span></u><u><span style='font-size:15px;font-family:"Tahoma","sans-serif";'>file&nbsp;</span></u><u><span style='font-size:19px;font-family:"Cordia New","sans-serif";'>ที่แนบมาให้เท่านั้น</span></u><span style='font-size:19px;font-family:"Cordia New","sans-serif";'>&nbsp;</span><span style='font-size:15px;font-family:"Tahoma","sans-serif";'>***</span></p>
        <p style='margin:0cm;margin-bottom:.0001pt;font-size:19px;font-family:"Tahoma","serif";'><span style='font-size:15px;font-family:"Verdana","sans-serif";color:#212121;'>&nbsp;</span><span style='font-size:15px;font-family:"Cordia New","sans-serif";color:#212121;'>***</span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:#212121;'>&nbsp;&nbsp;</span><span style='font-size:15px;font-family:"Cordia New","sans-serif";color:#1F497D;'>ใช้ ขนส่ง</span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:#1F497D;'>&nbsp;</span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:#1F497D;'>DHL &nbsp;</span><span style="font-size:15px;color:#1F497D;">รบกวน เตรียมแพค สินค้าลงกล่อง รอ ขนส่งเข้ารับได้เลยค่ะ</span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:#1F497D;'>&nbsp;</span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:#1F497D;'>&nbsp;</span><span style='font-size:15px;font-family:"Cordia New","sans-serif";color:#1F497D;'>หากขนส่งไม่เข้ารับ</span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:#1F497D;'>&nbsp;&nbsp;</span><span style='font-size:15px;font-family:"Cordia New","sans-serif";color:#1F497D;'>ภายใน 2 วัน ให้ แจ้งกลับมา</span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:#212121;'>&nbsp;/&nbsp;</span><span style='font-size:15px;font-family:"Browallia New","sans-serif";color:#212121;'>ช่องทางการติดต่อ</span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:#212121;'>&nbsp;</span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:#212121;'>**&nbsp;</span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:#1F497D;'>&nbsp;</span><span style='font-size:15px;font-family:"Calibri","sans-serif";color:#1F497D;'><a href="https://page.line.me/533aagih" id="LPlnk282976"><span style='font-family:"Verdana","sans-serif";'>https://page.line.me/533aagih</span></a></span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:#1F497D;'>&nbsp;</span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:#1F497D;'>&nbsp; (พี่อ้อ)</span></p>
        <p style='margin:0cm;margin-bottom:.0001pt;font-size:19px;font-family:"Tahoma","serif";text-indent:36.0pt;'><span style='font-size:15px;font-family:"Cordia New","sans-serif";color:#1F497D;'>&nbsp; &nbsp;</span><strong><span style="color:#1F497D;">&nbsp; &nbsp; &nbsp; &nbsp; &nbsp;&nbsp;</span></strong></p>
        <p style='margin:0cm;margin-bottom:.0001pt;font-size:19px;font-family:"Tahoma","serif";text-indent:36.0pt;'><strong><span style="color:#1F497D;">&nbsp; &nbsp; &nbsp; &nbsp; &nbsp; สำคัญ&nbsp;</span></strong></p>
        <p style='margin:0cm;margin-bottom:.0001pt;font-size:19px;font-family:"Tahoma","serif";'><span style='font-size:15px;font-family:"Verdana","sans-serif";color:#212121;'>&nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; ***&nbsp;</span><span style='font-size:15px;font-family:"Cordia New","sans-serif";color:#212121;'>1.</span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:#212121;'>&nbsp;</span><span style='font-size:15px;font-family:"Cordia New","sans-serif";color:white;background:navy;'>กรุณาอย่าติด</span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:white;background:navy;'>&nbsp;</span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:white;background:navy;'>Label&nbsp;</span><span style="font-size:15px;color:white;background:navy;">สลับ หรือ ผิด เลขที่บิล นะคะ</span></p>
        <p style='margin:0cm;margin-bottom:.0001pt;font-size:19px;font-family:"Tahoma","serif";'><span style='font-size:15px;font-family:"Verdana","sans-serif";color:#212121;'>&nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp;&nbsp;</span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:#212121;'>***</span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:#212121;'>&nbsp;</span><span style='font-size:15px;font-family:"Cordia New","sans-serif";color:#212121;'>2.</span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:#212121;'>&nbsp;</span><span style='font-size:15px;font-family:"Cordia New","sans-serif";color:white;background:red;'>กรุณาอย่ามี</span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:white;background:red;'>&nbsp;</span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:white;background:red;'>Label&nbsp;</span><span style='font-size:15px;font-family:"Verdana","sans-serif";color:white;background:red;'>&nbsp;</span><span style='font-size:15px;font-family:"Cordia New","sans-serif";color:white;background:red;'>ที่ใช้แล้ว ตรวจสภาพกล่องที่นำมาใช้ ด้วยนะ</span></p>
        <p id="isPasted" style='margin:0cm;margin-bottom:.0001pt;font-size:19px;font-family:"Tahoma","serif";'><span style='font-size:13px;font-family:"Segoe Print";color:#F89D52;'>B<strong>est Regards,</strong></span></p>
        <p style='margin:0cm;margin-bottom:.0001pt;font-size:19px;font-family:"Tahoma","serif";'><strong><span style='font-size:13px;font-family:"Segoe Print";color:#F89D52;'>Kamolwan &nbsp;Aromsawa (AOR)</span></strong></p>
        <p style='margin:0cm;margin-bottom:.0001pt;font-size:19px;font-family:"Tahoma","serif";'><strong><span style='font-size:13px;font-family:"Segoe Print";color:#1D1B11;'>Customer Service</span></strong></p>
        <p style='margin:0cm;margin-bottom:.0001pt;font-size:19px;font-family:"Tahoma","serif";'><strong><span style='font-size:16px;font-family:"Segoe Print";color:black;'>Email:</span></strong><strong><span style='font-size:16px;font-family:"Calibri","sans-serif";color:black;'>&nbsp;<a href="mailto:Kamolwan@comseven.com" id="LPNoLP">Kamolwan@comseven.com</a> &nbsp;</span></strong></p>
        <p style='margin:0cm;margin-bottom:.0001pt;font-size:19px;font-family:"Tahoma","serif";'><strong><span style='font-size:16px;font-family:"Segoe UI Symbol","sans-serif";color:black;'>🔸🔸🔸🔸🔸🔸🔸🔸🔸🔸🔸🔸🔸</span></strong></p>
        <p><span style='font-size:16px;font-family:"Calibri","sans-serif";color:black;'>02-017-7777 # 7312</span></p>
        <p style='margin:0cm;margin-bottom:.0001pt;font-size:15px;font-family:"Tahoma","sans-serif";'><img src="https://i.imgur.com/zGEe9kH.jpg"></p>
        <p style='margin:0cm;margin-bottom:.0001pt;font-size:19px;font-family:"Tahoma","serif";'><br></p>
        <p style='margin:0cm;margin-bottom:.0001pt;font-size:19px;font-family:"Tahoma","serif";'><br></p>
        <p style='margin:0cm;margin-bottom:.0001pt;font-size:19px;font-family:"Tahoma","serif";'><br></p>

        </body> \
        </html>      '''       
        mail.HTMLBody = html
        mail.Attachments.Add(Source=root1.filedir + '/' + str(value1) + '.pdf')
        if From != None:
        # This line basically calls the "mail.SendUsingAccount = xyz@email.com" outlook VBA command
            mail._oleobj_.Invoke(*(64209, 0, 8, 0, From))
        
        mail.Send()
        row += 1
        if value1 is None:
            break
        elif email2 is None:
            break
        else:
            continue
sentmail()