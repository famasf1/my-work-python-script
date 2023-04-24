import openpyxl as pyxl
from tkinter import Tk, filedialog
import pyautogui
from line_notify_me.line_notify_sourcecode import notifyme
from datetime import date, timedelta, datetime
from pyperclip import copy

root = Tk()
root.excel = filedialog.askopenfilename(title="เลือกไฟล์ Excel", filetypes=(
    ('Excel Files', '*.xlsx'), ('All Files', '*.*')))
root.withdraw()
wb = pyxl.load_workbook(root.excel, data_only=True)
ws = wb.sheetnames
main_sheet = wb[ws[0]]


class function_ITEC:
    def search_button(self):
        pyautogui.press('f12')

    def docref_button(self):
        pyautogui.press('f5')


def edit_docref():
    '''
    start the operation.
    '''
    def start():  # default behavior243
        
        '''
        start the operation. Opening stockout page, set the date back 1 year and click at ID field ready for searching operation.
        '''
        pyautogui.moveTo(45, 255)
        pyautogui.sleep(.7)
        pyautogui.leftClick()
        pyautogui.sleep(10)
        pyautogui.moveTo(1753, 192)
        pyautogui.leftClick()
        pyautogui.moveTo(1752, 215)
        pyautogui.doubleClick()
        pyautogui.leftClick()
        pyautogui.press('down')
        pyautogui.press('enter')
        pyautogui.sleep(1)
        pyautogui.moveTo(294, 89)
        pyautogui.doubleClick()
        pyautogui.leftClick()

    start()
    # CHANGE DATE HERE
    #day = date.today() - timedelta(days=1)
    

    for row in range(2, main_sheet.max_row+1):
        day = main_sheet.cell(row=row, column=2).value
        time = main_sheet.cell(row=row, column=3).value
        id = main_sheet.cell(row=row, column=4).value
        branch = main_sheet.cell(row=row, column=5).value
        isinsure = main_sheet.cell(row=row, column=6).value
        count = main_sheet.cell(row=row, column=7).value
        day = datetime.strptime(day, "%b %d, %Y ")
        dayformatted = day.strftime("%d/%m/%y")
        if isinsure == "Stockout_Insure":
            pyautogui.sleep(.7)
            pyautogui.moveTo(288, 88)
            pyautogui.sleep(.7)
            pyautogui.doubleClick()
            pyautogui.typewrite(str(id))
            pyautogui.press('enter')
            pyautogui.typewrite(str(branch))
            function_ITEC().search_button()
            pyautogui.press('enter')
            function_ITEC().docref_button()
            copy("DHLเข้ารับ")
            pyautogui.hotkey('ctrl','v')
            pyautogui.typewrite(f"{count}")
            copy("ก.")
            pyautogui.hotkey('ctrl','v')
            pyautogui.typewrite(f"{dayformatted} | {time}")
            pyautogui.hotkey('alt', 'o')
            pyautogui.press('enter')

        else:
            continue
    notifyme('Docref for DHL Shipment is finished!')


if __name__ in '__main__':
    edit_docref()
