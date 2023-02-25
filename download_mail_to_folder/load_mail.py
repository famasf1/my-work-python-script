import win32com.client
import os
from datetime import date, datetime, timedelta
from openpyxl import load_workbook
import random

def main():
    '''
    Get all attachments and stored them
    '''
    outlook = win32com.client.dynamic.Dispatch("Outlook.Application").GetNamespace("MAPI")
    root_folder = outlook.GetDefaultFolder(6)
    Pickup_DHL_subfolder = root_folder.Folders.Item("Pickup DHL")
    messages = Pickup_DHL_subfolder.Items
    ######
    ## Get path
    PDF_FOLDER_PATH = r"D:\Workstuff\my-work-python-script\download_mail_to_folder\pdf"
    EXCEL_FOLDER_PATH = r"D:\Workstuff\my-work-python-script\download_mail_to_folder\xlsx"
    ######
    ## Get Time
    time = datetime.now().strftime("%d-%m-%y_%H-%M")
    today = date.today()
    yesterday = today - timedelta(days=1)
    ######FW: DHL eCommerce pick up- 2022-12-26
    ## First loop will go through messeges and find how many messege match this criteria
    subject = f"FW: DHL eCommerce pick up- {yesterday}"
    count = 0
    for m in messages:
        if m.Subject == subject:
            attachments = m.Attachments
            num_attach = len([a for a in attachments])
        ## this loop goes through each mail and read how many attachment inside
            for attachment_excel in range(1, num_attach+1):
                attachment = attachments.Item(attachment_excel)  
                if (attachment.FileName).endswith('xlsx'):
                    ###As soon as you downloaded, get value in there and instantly rename yourself
                    download_path = os.path.join(f"{EXCEL_FOLDER_PATH}", rf"{attachment.FileName}")
                    try:
                        attachment.SaveAsFile(download_path)
                        wb = load_workbook(download_path)
                        wb_name = wb.sheetnames
                        ws = wb[wb_name[0]]
                        shipment_last_6digit = ws['A2'].value[-7:]
                        print(shipment_last_6digit)
                        shipment_time = ws['B2'].value
                        str(shipment_time).replace(":","-")
                        new_name_path = os.path.join(EXCEL_FOLDER_PATH, f"{shipment_time}_{shipment_last_6digit}.xlsx".replace(":","-"))
                        os.rename(download_path,new_name_path)
                        count += 1
                    except FileExistsError:
                        pass

    print(f"Total : {count} items")

if __name__ in "__main__":
    #get_time()
    main()