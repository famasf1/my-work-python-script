import os
import pandas as pd
from datetime import date, datetime, timedelta
import subprocess
import pyautogui as pyg
from openpyxl import load_workbook
import openpyxl 
import random
import win32com.client
import warnings
import requests
from tkinter import Tk, filedialog
import tkinter
from pyperclip import copy
import sys

daystime = 1

EXCEL_READ_FILEPATH = r"D:\Workstuff\my-work-python-script\download_mail_to_folder\xlsx"
EXCEL_WRITE_FILEPATH = r"D:\Workstuff\my-work-python-script\download_mail_to_folder\Result_archiev"
time = datetime.now().strftime("%d-%m-%y_%H-%M")
today = date.today()
yesterday = today - timedelta(days=1)


class function_ITEC:
    def search_button(self):
        pyg.press('f12')

    def docref_button(self):
        pyg.press('f5')



def load_mail():
    '''
    Get all attachments and stored them
    '''
    outlook = win32com.client.dynamic.Dispatch("Outlook.Application").GetNamespace("MAPI")
    root_folder = outlook.GetDefaultFolder(6)
    Pickup_DHL_subfolder = root_folder.Folders.Item("Pickup DHL")
    messages = Pickup_DHL_subfolder.Items
    ######
    ## Get path
    PDF_FOLDER_PATH = r"D:\Workstuff\my-work-python-script\download_mail_to_folder\pdf"
    EXCEL_FOLDER_PATH = r"D:\Workstuff\my-work-python-script\download_mail_to_folder\xlsx"

    ######FW: DHL eCommerce pick up- 2022-12-26
    ## First loop will go through messeges and find how many messege match this criteria
    subject = f"FW: DHL eCommerce pick up- {yesterday}"
    count = 0
    for m in messages:
        if m.Subject == subject:
            attachments = m.Attachments
            num_attach = len([a for a in attachments])
        ## this loop goes through each mail and read how many attachment inside
            for attachment_excel in range(1, num_attach+1):
                attachment = attachments.Item(attachment_excel)  
                if (attachment.FileName).endswith('xlsx'):
                    ###As soon as you downloaded, get value in there and instantly rename yourself
                    download_path = os.path.join(f"{EXCEL_FOLDER_PATH}", rf"{attachment.FileName}")
                    try:
                        attachment.SaveAsFile(download_path)
                        wb = load_workbook(download_path)
                        wb_name = wb.sheetnames
                        ws = wb[wb_name[0]]
                        shipment_last_6digit = ws['A2'].value[-7:]
                        print(shipment_last_6digit)
                        shipment_time = ws['B2'].value
                        str(shipment_time).replace(":","-")
                        new_name_path = os.path.join(EXCEL_FOLDER_PATH, f"{shipment_time}_{shipment_last_6digit}.xlsx".replace(":","-"))
                        os.rename(download_path,new_name_path)
                        count += 1
                    except FileExistsError:
                        pass

    print(f"Total : {count} items")

def create_excel():
    ############### DAYS #############
    # if monday, set 3 = friday, 2 = saturday, 1 = sunday
    daystime = 1

    yesterday = date.today() - timedelta(days=daystime)

    #Create excel
    os.chdir(EXCEL_WRITE_FILEPATH)
    new_wb = openpyxl.Workbook()
    new_wb.save(f"Parcel Pickup {yesterday}.xlsx")
    load_new_wb = openpyxl.load_workbook(os.path.join(EXCEL_WRITE_FILEPATH, f"Parcel Pickup {yesterday}.xlsx"), data_only=True)
    sheet = load_new_wb['Sheet']
    sheet.cell(row=1,column=1).value = "Tracking"
    sheet.cell(row=1,column=2).value = "Time"
    sheet.title = f"{yesterday}"
    load_new_wb.save(filename=f"Parcel Pickup {yesterday}.xlsx")

    #Combine all into one
    os.chdir(EXCEL_READ_FILEPATH)
    with warnings.catch_warnings(record=True):
        warnings.simplefilter("always")
        for file in os.listdir(EXCEL_READ_FILEPATH):
            data = pd.read_excel(file, dtype=object) ##turn excel into dataframe
            with pd.ExcelWriter(f"..//Result_archiev//Parcel Pickup {yesterday}.xlsx", mode="a", if_sheet_exists="overlay", engine="openpyxl") as excel_file:
                excel_file.number_format = openpyxl.styles.numbers.FORMAT_TEXT
                data.to_excel(excel_file, sheet_name=f"{yesterday}", header=False, index=False, startrow=excel_file.sheets[f"{yesterday}"].max_row)

    #Load file and convert time to 24h format
    os.chdir(EXCEL_WRITE_FILEPATH)
    load_time = pd.read_excel(f"Parcel Pickup {yesterday}.xlsx")
    shipment_number = pd.DataFrame(load_time['Tracking'])

    time_column_shipment = pd.DataFrame(load_time['Time'].str.split("([\d][\d][:][\d][\d] [A-Z][A-Z])", expand=True, regex=True))
    frames = [shipment_number, time_column_shipment]
    new_data = pd.concat(frames, axis=1)
    with pd.ExcelWriter(f"Parcel Pickup {yesterday}.xlsx", mode="a", if_sheet_exists="replace", engine="openpyxl") as excel_file:
        new_data.to_excel(excel_file,sheet_name=f"{yesterday}", header=False, index=False, startrow=1)

def getITECdata():
    #stockout
    def stockout():
        ##ID49
        subprocess.Popen("C:\Program Files (x86)\Softbox\ITEC2007_49\ITECStock2007.exe")
        ##ITEC Login Script
        def login(user, pwd):
            pyg.sleep(1)
            def do_your_thing(what_field):
                pyg.click(what_field)
                pyg.write(user)
                pyg.press("enter")
                pyg.write(pwd)
                pyg.press("enter")
            while True:
                pyg.sleep(1)
                user_field = pyg.locateOnScreen(r"D:\Workstuff\my-work-python-script\asset\user_field.png")
                user_field2 = pyg.locateOnScreen(r"D:\Workstuff\my-work-python-script\asset\user_field_is1920.png")
                user_field3 = pyg.locateOnScreen(r"D:\Workstuff\my-work-python-script\asset\your_id_pls.png")
                if user_field:
                    do_your_thing(user_field)
                    break
                elif user_field2:
                    do_your_thing(user_field2)
                    break
                elif user_field3:
                    do_your_thing(user_field3)
                    break
                else:
                    pyg.sleep(1)
        login("22608", "22608")
        pyg.hotkey("alt", "k")
        pyg.press("u", 2)
        pyg.press("enter")
        pyg.sleep(10)
        pyg.click(1737,191)
        pyg.press("down")
        pyg.press("enter")
        pyg.sleep(1)
        pyg.doubleClick(401,86)
        pyg.typewrite("0")
        pyg.press("f12")
        pyg.sleep(1200) #20minute
        pyg.click(596,337)
        pyg.hotkey("ctrl","a")
        pyg.hotkey("ctrl", "c")
        pyg.sleep(300)
        stockout = pd.read_clipboard(sep='\t', dtype={'0' : 'string','16' : 'string', '11' : 'string', '15' : 'string', '18' : 'string', '21' : 'string'})
        stockoutid = stockout["Stock Out (ID)"]
        branchid = stockout["Branch (ID)"]
        booking_id = stockout["Booking ID"].str.rstrip().str.replace("Booking-DHL ID : ","").str.split(" , ", expand=True)

        frames = [stockoutid, branchid, booking_id]
        data_itec = pd.concat(frames, axis=1)
        with pd.ExcelWriter(f"Parcel Pickup {yesterday}.xlsx", mode="a", if_sheet_exists="overlay", engine="openpyxl") as excel_file:
            data_itec.to_excel(excel_file, sheet_name="stockout")
    stockout()

    #stockout insure
    def stockoutinsure():
        #49INSURE
        subprocess.Popen("C:\Program Files (x86)\Softbox\ITECInsurance_49\ITECStock2007.exe")
        def login(user, pwd):
            pyg.sleep(1)
            def do_your_thing(what_field):
                pyg.click(what_field)
                pyg.write(user)
                pyg.press("enter")
                pyg.write(pwd)
                pyg.press("enter")
            while True:
                pyg.sleep(1)
                user_field = pyg.locateOnScreen(r"D:\Workstuff\my-work-python-script\asset\user_field.png")
                user_field2 = pyg.locateOnScreen(r"D:\Workstuff\my-work-python-script\asset\user_field_is1920.png")
                user_field3 = pyg.locateOnScreen(r"D:\Workstuff\my-work-python-script\asset\your_id_pls.png")
                if user_field:
                    do_your_thing(user_field)
                    break
                elif user_field2:
                    do_your_thing(user_field2)
                    break
                elif user_field3:
                    do_your_thing(user_field3)
                    break
                else:
                    pyg.sleep(1)
        login("22608", "22608")
        pyg.hotkey("alt", "k")
        pyg.press("u", 2)
        pyg.press("enter")
        pyg.sleep(10)
        pyg.click(1737,191)
        pyg.press("down")
        pyg.press("enter")
        pyg.sleep(1)
        pyg.doubleClick(401,86)
        pyg.typewrite("0")
        pyg.press("f12")
        pyg.sleep(600) #10minute
        pyg.click(596,337)
        pyg.hotkey("ctrl","a")
        pyg.hotkey("ctrl", "c")
        pyg.sleep(180)
        stockoutinsure()
        stockout_insure = pd.read_clipboard(sep='\t')
        stockoutid_insure = stockout_insure["Stock Out (ID)"]
        branchid_insure = stockout_insure["Branch (ID)"] #
        booking_id_insure = stockout_insure["Booking ID"].str.rstrip().str.replace("Booking-DHL ID : ","").str.split(" , ", expand=True)

        frames_insure = [stockoutid_insure, branchid_insure, booking_id_insure]
        data_insure = pd.concat(frames_insure, axis=1)
        with pd.ExcelWriter(f"Parcel Pickup {yesterday}.xlsx", mode="a", if_sheet_exists="overlay", engine="openpyxl") as excel_file:
            data_insure.to_excel(excel_file, sheet_name="stockout_insure")
    stockoutinsure()

def open_excel():
    # create excel object
    excel = win32com.client.gencache.EnsureDispatch("Excel.Application")

    # excel can be visible or not
    excel.Visible = True  # False
    
    # try except for file / path
    try:
        wb = excel.Workbooks.Open(os.path.join(EXCEL_WRITE_FILEPATH, f"Parcel Pickup {yesterday}.xlsx",))
    except Exception as e:
        if e.excepinfo[5] == -2146827284:
            print(f'Failed to open spreadsheet.  Invalid filename or location: {wb}')
        else:
            raise e
        sys.exit(1)
    pyg.sleep(2)
    wb.Close()
    excel.Quit()



def setdataexcel():
    def convert24(str1):
    #check if am and equal 12
        if str1[-2:] == "AM" and str1[:2] == "12":
            return "00" + str1[2:-2]

        #otherwise just remove AM
        elif str1[-2:] == "AM":
            return str1[:-2]
        
        #but if it's pm and equal 12
        elif str1[-2:] == "PM" and str1[:2] == "12":
            return str1[:-2]

        else:
            return str(int(str1[:2]) + 12) + str1[2:5]
    load_wb = openpyxl.load_workbook(os.path.join(EXCEL_WRITE_FILEPATH, f"Parcel Pickup {yesterday}.xlsx"), data_only=True)
    load_sheet = load_wb.sheetnames
    shipment_data = load_wb[load_sheet[0]]
    stockout_normal = load_wb[load_sheet[1]]
    stockout_insurance = load_wb[load_sheet[2]]

    for row in range(2,shipment_data.max_row+1):
        shipment_data.cell(row=row, column=3).value = convert24(shipment_data.cell(row=row, column=3).value)
        shipment_data.cell(row=row, column=4).value = f'''=IF(ISNUMBER(MATCH({openpyxl.utils.quote_sheetname(shipment_data.title)}!A{row},stockout!D:D,0)),INDEX(stockout!B:B,MATCH({openpyxl.utils.quote_sheetname(shipment_data.title)}!A{row},stockout!D:D,0)),IF(ISNUMBER(MATCH({openpyxl.utils.quote_sheetname(shipment_data.title)}!A{row},stockout_insure!D:D,0)),INDEX(stockout_insure!B:B,MATCH({openpyxl.utils.quote_sheetname(shipment_data.title)}!A{row},stockout_insure!D:D,0)),""))'''
        shipment_data.cell(row=row, column=5).value = f'''=IF(ISNUMBER(MATCH({openpyxl.utils.quote_sheetname(shipment_data.title)}!A{row},stockout!D:D,0)),INDEX(stockout!C:C,MATCH({openpyxl.utils.quote_sheetname(shipment_data.title)}!A{row},stockout!D:D,0)),IF(ISNUMBER(MATCH({openpyxl.utils.quote_sheetname(shipment_data.title)}!A{row},stockout_insure!D:D,0)),INDEX(stockout_insure!C:C,MATCH({openpyxl.utils.quote_sheetname(shipment_data.title)}!A{row},stockout_insure!D:D,0)),""))'''
        shipment_data.cell(row=row, column=6).value = f'''=IF(ISNUMBER(MATCH({openpyxl.utils.quote_sheetname(shipment_data.title)}!A{row},stockout!D:D,0)),"Stockout",IF(ISNUMBER(MATCH({openpyxl.utils.quote_sheetname(shipment_data.title)}!A{row},stockout_insure!D:D,0)),"Stockout_Insure",""))'''
    load_wb.save(f"Parcel Pickup {yesterday}.xlsx")




####TODO
def load_stockout():
    def stockout():
        #root = Tk()
        #root.excel = filedialog.askopenfilename(title="เลือกไฟล์ Excel", filetypes=(('Excel Files','*.xlsx'),('All Files', '*.*')))
        #root.withdraw()
        wb = openpyxl.load_workbook(os.path.join(EXCEL_WRITE_FILEPATH, f"Parcel Pickup {yesterday}.xlsx"), data_only=True)
        ws = wb.sheetnames
        main_sheet = wb[ws[0]]


        '''
        start the operation.
        '''
        def start(): #default behavior
            '''
            start the operation. Opening stockout page, set the date back 1 year and click at ID field ready for searching operation.
            '''
            pyg.moveTo(45,255)
            pyg.sleep(.7)
            pyg.leftClick()
            pyg.sleep(10)
            pyg.moveTo(1753,192)
            pyg.leftClick()
            pyg.moveTo(1752,215)
            pyg.doubleClick()
            pyg.leftClick()
            pyg.press('down')
            pyg.press('enter')
            pyg.sleep(1)
            pyg.moveTo(294,89)
            pyg.doubleClick()
            pyg.leftClick()

        start()
        # CHANGE DATE HERE

        for row in range(2, main_sheet.max_row+1):
            day = main_sheet.cell(row=row, column=2).value
            time = main_sheet.cell(row=row, column=3).value
            id = main_sheet.cell(row=row, column=4).value
            branch = main_sheet.cell(row=row, column=5).value
            isinsure = main_sheet.cell(row=row, column=6).value
            day = datetime.strptime(day, "%b %d, %Y ")
            dayformatted = day.strftime("%d/%m/%y")
            if isinsure == "Stockout":
                pyg.sleep(.7)
                pyg.moveTo(288,88)
                pyg.sleep(.7)
                pyg.doubleClick()
                pyg.typewrite(str(id))
                pyg.press('enter')
                pyg.typewrite(str(branch))
                function_ITEC().search_button()
                pyg.press('enter')
                function_ITEC().docref_button()
                copy("DHL เข้ารับ: ")
                pyg.hotkey('ctrl','v')
                pyg.typewrite(f"{dayformatted} | {time}")
                pyg.press('tab')
                pyg.press('enter')
                pyg.press('enter')
            else: continue

        class function_ITEC:
            def search_button(self):
                pyg.press('f12')
            def docref_button(self):
                pyg.press('f5')

    def stockout_insure():

        '''
        start the operation.
        '''
        def start():  # default behavior
            '''
            start the operation. Opening stockout page, set the date back 1 year and click at ID field ready for searching operation.
            '''
            pyg.moveTo(45, 255)
            pyg.sleep(.7)
            pyg.leftClick()
            pyg.sleep(10)
            pyg.moveTo(1753, 192)
            pyg.leftClick()
            pyg.moveTo(1752, 215)
            pyg.doubleClick()
            pyg.leftClick()
            pyg.press('down')
            pyg.press('enter')
            pyg.sleep(1)
            pyg.moveTo(294, 89)
            pyg.doubleClick()
            pyg.leftClick()

        start()
        # CHANGE DATE HERE
        #day = date.today() - timedelta(days=1)
        

        for row in range(2, main_sheet.max_row+1):
            day = main_sheet.cell(row=row, column=2).value
            time = main_sheet.cell(row=row, column=3).value
            id = main_sheet.cell(row=row, column=4).value
            branch = main_sheet.cell(row=row, column=5).value
            isinsure = main_sheet.cell(row=row, column=6).value
            day = datetime.strptime(day, "%b %d, %Y ")
            dayformatted = day.strftime("%d/%m/%y")
            if isinsure == "Stockout_Insure":
                pyg.sleep(.7)
                pyg.moveTo(288, 88)
                pyg.sleep(.7)
                pyg.doubleClick()
                pyg.typewrite(str(id))
                pyg.press('enter')
                pyg.typewrite(str(branch))
                function_ITEC().search_button()
                pyg.press('enter')
                function_ITEC().docref_button()
                copy("DHL เข้ารับ: ")
                pyg.hotkey('ctrl', 'v')
                pyg.typewrite(f"{dayformatted} | {time}")
                pyg.press('tab')
                pyg.press('enter')
                pyg.press('enter')
            else:
                continue
    
    stockout_insure()
    ###SWAP TO NORMAL
    stockout()
   

### Notify me when the script is completed to LINE.
def notifyme(confirmtext):
    """
    LINE Notify - Send text to my own line.
    parameter :
    confirmtext: str (required)
    """
    
    mytoken = 'kOcQyjPGgIAgTQ4qWjTlEJZFUj7GegzGefdDEiSsYJr'
    url = 'https://notify-api.line.me/api/notify'
    data = {
        'message' : confirmtext
    }
    options = {
        'Method' : 'POST',
        'Content-Type' : 'application/x-www-form-urlencoded',
        'Authorization' : f'Bearer {mytoken}',
    }
    response = requests.post(url=url, headers=options, data=data)
    print(response.status_code)





if __name__ in "__main__":
    #get_time()
    #step 1 : load email
    load_mail()
    #step 2 : creating excel file
    create_excel()
    #step 3 : get data from ITEC
    getITECdata()
    #step 4 : write data to excel
    setdataexcel()
    #step 5 : read excel to cache it
    open_excel()
    #step 6 : load into stock out
    #TODO
    #final step : line to me when it's done
    notifyme("รวมข้อมูลเสร็จสมบูรณ์")