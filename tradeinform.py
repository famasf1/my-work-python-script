import openpyxl
from tkinter import filedialog, messagebox
from tkinter import *
import os
import ctypes
import sys
import datetime as dt
import win32com.client

def MBox(title,text,styles): # create popup windows messages
    return ctypes.windll.user32.MessageBoxW(0,text,title,styles)

root1 = Tk()
root1.withdraw()
root1.fileask = filedialog.askopenfilename(initialdir="/",title="เลือกไฟล์ Excel ข้อมูล",filetypes=(("Excel","*.xlsx"),("All Files","*.*")))
#root1.dir = filedialog.askdirectory(initialdir="/",title="เลือกที่จัดเก็บใบส่ง Trade In")

#os.chdir(root1.dir)#(sys._MEIPASS)

shsource = openpyxl.load_workbook(root1.fileask, data_only=True)


sourcewb = shsource.sheetnames
form = shsource[sourcewb[4]]
person1 = shsource[sourcewb[0]]
#try:
person2 = shsource[sourcewb[1]]
#except IndexError:
#    try:
person3 = shsource[sourcewb[2]]
#    except IndexError:
#        try:
person4 = shsource[sourcewb[3]]
#        except IndexError:
#            person2 = ''
#            person3 = ''
#            person4 = ''
    
#try:
person = (person1,person2,person3,person4)
# except NameError:
#     try:
#         person = (person1,person2,person3)
#     except NameError:
#         try:
#             person = (person1, person2)
#         except NameError:
#             person = (person1)



def readsheet():
    p = 0
    n = 1
    o = 1
    def checkme():
        if p == 4:
            MBox("ทำรายการสำเร็จ!","ทำรายการสำเร็จแล้ว!",0)
    while True:
        try:
            for o in range(1,person[p].max_row+1):
                phyid = person[p].cell(row=o,column=1).value
                etc = person[p].cell(row=o,column=3).value
                customerinfo = person[p].cell(row=o,column=4).value
                prodcode = person[p].cell(row=o,column=5).value
                prodname = person[p].cell(row=o,column=6).value
                serial = person[p].cell(row=o,column=2).value
                supplier = person[p].cell(row=o,column=8).value
                price = person[p].cell(row=o,column=9).value
                vouchercode = person[p].cell(row=o,column=10).value
                date2 = person[p].cell(row=o,column=11).value
                if phyid:
                    def write():
                        form['C4'].value = '=INDEX(Sheet1!A1:A5,MATCH(C5,Sheet1!B1:B5,0))'
                        form['C5'].value = str(supplier)
                        form['C6'].value = str(phyid)
                        t = dt.date.today()
                        form['G4'].value = str(date2) #t.strftime("%d/%m/%Y")
                        form['G5'].value = str(customerinfo)
                        form['G6'].value = str(vouchercode)
                        form['B8'].value = str(prodcode)
                        form['F8'].value = str(prodname)
                        form['B11'].value = '*' + str(serial) + '*'
                        form['B12'].value = str(serial)
                        form['H11'].value = str(price)
                        form['I11'].value = str(etc)
                        form['B1'].value = 'ผู้จัดทำ : ' + str(shsource.sheetnames[p])
                    write()
                    o += 1
                    shsource.save(root1.fileask)
                    a = win32com.client.Dispatch('Excel.Application')
                    a.visible = False
                    wb = a.Workbooks.Open(str(root1.fileask))
                    getsheet = wb.Worksheets([5])
                    getsheet.PrintOut()
                    wb.Close(True)
                    n += 1
                    if o == person[p].max_row+1: # if all row is reached maximum number
                        o = 1 # go back to row 1
                        n = 1 # restart sheet name to 1
                        # but change source to the second sheet
                        p += 1
                else: 
                    p += 1
                    checkme()
                    break
        except IndexError:
            p +=1
            checkme()
            break
        #except (IndexError, AttributeError):
            
            #MBox("ทำรายการสำเร็จ!","ทำรายการสำเร็จแล้ว!",0)
            #break



try:
    readsheet()
except Exception as e:
    messagebox.showerror('Python Error', f'{e}')