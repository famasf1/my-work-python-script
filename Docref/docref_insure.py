import openpyxl as pyxl
from tkinter import Tk, filedialog
import pyautogui
from line_notify_me.line_notify_sourcecode import notifyme

root = Tk()
root.excel = filedialog.askopenfilename(title="เลือกไฟล์ Excel", filetypes=(('Excel Files','*.xlsx'),('All Files', '*.*')))
root.withdraw()
wb = pyxl.load_workbook(root.excel, data_only=True)
ws = wb.sheetnames
main_sheet = wb[ws[0]]



class function_ITEC:
    def search_button(self):
        pyautogui.press('f12')
    def docref_button(self):
        pyautogui.press('f5')


def edit_insure():
    '''
    start the operation.
    '''
    def start(): #default behavior
        '''
        start the operation. Opening stockout page, set the date back 1 year and click at ID field ready for searching operation.
        '''
        pyautogui.moveTo(45,255)
        pyautogui.sleep(.7)
        pyautogui.leftClick()
        pyautogui.sleep(3)
        pyautogui.moveTo(1753,192)
        pyautogui.leftClick()
        pyautogui.moveTo(1752,215)
        pyautogui.doubleClick()
        pyautogui.leftClick()
        pyautogui.press('down')
        pyautogui.press('enter')
        pyautogui.sleep(1)
        pyautogui.moveTo(294,89)
        pyautogui.doubleClick()
        pyautogui.leftClick()

    start()
    for row in range(2, main_sheet.max_row+1):
        bitly_link = main_sheet.cell(row=row, column=15).value
        id = main_sheet.cell(row=row, column=12).value
        branch = main_sheet.cell(row=row, column=13).value
        if bitly_link:
            pyautogui.sleep(.7)
            pyautogui.moveTo(288,88)
            pyautogui.sleep(.7)
            pyautogui.doubleClick()
            pyautogui.typewrite(str(id))
            pyautogui.press('enter')
            pyautogui.typewrite(str(branch))
            function_ITEC().search_button()
            pyautogui.press('enter')
            function_ITEC().docref_button()
            pyautogui.typewrite(f"{bitly_link}")
            pyautogui.press('tab')
            pyautogui.press('enter')
            pyautogui.press('enter')
        else: break
    notifyme('Docref for Insure is finished!')

if __name__ in '__main__':
    edit_insure()