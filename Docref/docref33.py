import pyautogui
import openpyxl
from tkinter import filedialog, messagebox
from tkinter import *
import datetime
import pyperclip
import requests
###################### Part 1 : Start the variable process

def startup():
    try:
        global workbook
        root = Tk()
        root.excel = filedialog.askopenfilename(initialdir='/Desktop',title='เลือกไฟล์ Excel สำหรับ Stock-Out33', filetypes=(('Excel','*.xlsx'),('All Files','*.*')))
        workbook = openpyxl.load_workbook(root.excel, data_only=True)
        root.withdraw()
        worksheetname = workbook.sheetnames
        global ID33_Data, ID33BKK_Data, ID49_Data, ID49BKK_Data, ID747_Data, delivery_Failed_Data, receive
        ID33_Data = workbook[worksheetname[1]]
        ID33BKK_Data = workbook[worksheetname[6]]
        ID49_Data = workbook[worksheetname[3]]
        ID49BKK_Data = workbook[worksheetname[8]]
        ID747_Data = workbook[worksheetname[4]]
        delivery_Failed_Data = workbook[worksheetname[5]]
        receive = 'รับ'
    except Exception as e:
        messagebox.showerror('Python Error', f'{e}')
        exit()

def notifyme(confirmtext):
    mytoken = 'kOcQyjPGgIAgTQ4qWjTlEJZFUj7GegzGefdDEiSsYJr'
    url = 'https://notify-api.line.me/api/notify'
    data = {
        'message' : confirmtext
    }
    options = {
        'Method' : 'POST',
        'Content-Type' : 'application/x-www-form-urlencoded',
        'Authorization' : f'Bearer {mytoken}',
    }
    response = requests.post(url=url, headers=options, data=data)

def getdate_Obj(dateData):
    try:
        date_obj = datetime.datetime.strptime(dateData, "%Y-%m-%d %H:%M:%S.%f")
    except:
        try:
            date_obj = datetime.datetime.strptime(dateData, "%Y-%m-%d %H:%M:%S")
        except:
            try:
                date_obj = datetime.datetime.strptime(dateData, "%d-%m-%Y %H:%M:%S")
            except:
                try:
                    date_obj = datetime.datetime.strptime(dateData, "%d/%m/%Y")
                except Exception as e:
                    print(e)
                
    return date_obj.strftime('%d/%m/%y')

################################ TEST ROOM ###############################

class function_ITEC:
    def search_button(self):
        pyautogui.press('f12')
    def docref_button(self):
        pyautogui.press('f5')

#########################################################################
################################ Part 2: Reading excel and Start writing
def docref():
    def defaultref(): #Do this everytime i want to start.
        pyautogui.moveTo(45,255)
        
        pyautogui.leftClick()
        pyautogui.sleep(10)
        pyautogui.moveTo(1753,192)
        pyautogui.leftClick()
        pyautogui.moveTo(1752,215)
        pyautogui.doubleClick()
        pyautogui.leftClick()
        pyautogui.press('down')
        pyautogui.press('enter')
        
        pyautogui.moveTo(294,89)
        pyautogui.doubleClick()
        pyautogui.leftClick()

    def docref747(): #49 Return DHL
        for i in range(2,ID747_Data.max_row+1):
            out_sect = ID747_Data.cell(row=i,column=15).value
            id = ID747_Data.cell(row=i,column=12).value
            branch = ID747_Data.cell(row=i,column=13).value
            date = ID747_Data.cell(row=i, column=7).value
            formulae = f"=ifna(VLOOKUP(M{i},Data!C:G,5,0),"")"
            ID747_Data.cell(row=i,column=15).value = formulae
            workbook.save('bitly+ready.xlsx')
            if out_sect:
                
                pyautogui.moveTo(288,88)
                
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                function_ITEC().search_button()
                pyautogui.press('enter')
                function_ITEC().docref_button()
                pyperclip.copy(receive)
                pyautogui.hotkey('ctrl','v')
                pyautogui.typewrite(f"{getdate_Obj(str(date))} | {out_sect}")
                pyautogui.press('tab')
                pyautogui.press('enter')
                pyautogui.press('enter')
            else: break
        notifyme('docref33 finished!')

    def failed_toDeliver():
        for row in range(2, delivery_Failed_Data.max_row+1):
            reason = delivery_Failed_Data.cell(row=row, column=9).value
            phyid = delivery_Failed_Data.cell(row=row, column=17).value
            branch = delivery_Failed_Data.cell(row=row, column=18).value
            date = delivery_Failed_Data.cell(row=row, column=12).value
            if reason:
                
                pyautogui.moveTo(288,88)
                
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                function_ITEC().search_button()
                pyautogui.press('enter')
                function_ITEC().docref_button()
                pyperclip.copy(str(reason))
                (0.8)
                pyautogui.write(f"{pyperclip.paste()} | {getdate_Obj(str(date))}")
                pyautogui.press('tab')
                pyautogui.press('enter')
                pyautogui.press('enter')
            else: break
        docref747()


    def docref49returnbkk(): #49 Return BKK
        for i in range(2,ID49BKK_Data.max_row+1):
            date = ID49BKK_Data.cell(row=i, column=1).value
            id = ID49BKK_Data.cell(row=i,column=2).value
            branch = ID49BKK_Data.cell(row=i,column=3).value
            zone = ID49BKK_Data.cell(row=i,column=4).value          
            if id:
                
                pyautogui.moveTo(288,88)
                
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                function_ITEC().search_button()
                pyautogui.press('enter')
                function_ITEC().docref_button()
                pyperclip.copy(receive)
                pyautogui.hotkey('ctrl','v')
                pyautogui.typewrite(f'{str(getdate_Obj(str(date)))} | ')
                pyperclip.copy(str(zone))
                pyautogui.hotkey('ctrl','v')
                pyautogui.press('tab')
                pyautogui.press('enter')
                pyautogui.press('enter')
            else: break
        failed_toDeliver()
        
    def docref49return(): #49 Return DHL
        for i in range(2,ID49_Data.max_row+1):
            out_sect = ID49_Data.cell(row=i,column=15).value
            id = ID49_Data.cell(row=i,column=12).value
            branch = ID49_Data.cell(row=i,column=13).value
            date = ID49_Data.cell(row=i, column=7).value
            formulae = f"=ifna(VLOOKUP(M{i},Data!C:G,5,0),"")"
            ID49_Data.cell(row=i,column=15).value = formulae
            workbook.save('bitly+ready.xlsx')
            if out_sect:
                
                pyautogui.moveTo(288,88)
                
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                function_ITEC().search_button()
                pyautogui.press('enter')
                function_ITEC().docref_button()
                pyperclip.copy(receive)
                pyautogui.hotkey('ctrl','v')
                pyautogui.typewrite(f"{getdate_Obj(str(date))} | {out_sect}")
                pyautogui.press('tab')
                pyautogui.press('enter')
                pyautogui.press('enter')
            else: break
        docref49returnbkk()
        
    def docref33bkk(): # 33 Service Headoffice BKK
        for i in range(2,ID33BKK_Data.max_row+1):
            date = ID33BKK_Data.cell(row=i,column=1).value
            id = ID33BKK_Data.cell(row=i,column=2).value
            branch = ID33BKK_Data.cell(row=i,column=3).value
            zone = ID33BKK_Data.cell(row=i,column=4).value
            if date:
                
                pyautogui.moveTo(288,88)
                
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                function_ITEC().search_button()
                pyautogui.press('enter')
                function_ITEC().docref_button()
                pyperclip.copy(receive)
                pyautogui.hotkey('ctrl','v')
                pyautogui.typewrite(f'{getdate_Obj(str(date))} | ')
                pyperclip.copy(str(zone))
                pyautogui.hotkey('ctrl','v')
                pyautogui.press('tab')
                pyautogui.press('enter')
                pyautogui.press('enter')
            else: break
        docref49return()
    
    # THIS IS WHERE EVERYTHING STARTED!
    def docref33(): # 33 Service Headoffice DHL
        defaultref()
        for i in range(2,ID33_Data.max_row+1):
            out_sect = ID33_Data.cell(row=i,column=15).value
            id = ID33_Data.cell(row=i,column=12).value
            branch = ID33_Data.cell(row=i,column=13).value
            date = ID33_Data.cell(row=i, column=7).value
            formulae = f"=ifna(VLOOKUP(M{i},Data!C:G,5,0),"")"
            ID33_Data.cell(row=i,column=15).value = formulae
            workbook.save('bitly+ready.xlsx')
            if out_sect:
                
                pyautogui.moveTo(288,88)
                
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                function_ITEC().search_button()
                pyautogui.press('enter')
                function_ITEC().docref_button()
                pyperclip.copy(receive)
                pyautogui.hotkey('ctrl','v')
                pyautogui.typewrite(f"{getdate_Obj(str(date))} | {out_sect}")
                pyautogui.press('tab')
                pyautogui.press('enter')
                pyautogui.press('enter')
            else: break 
        docref33bkk()

    docref33()


############################ Finally  - Try running it. If there's error, print it out.
try:
    startup()
    docref()
except Exception as e:
    messagebox.showerror('Python Error',f'{e}')

print(f'Finished! Zero Error')

