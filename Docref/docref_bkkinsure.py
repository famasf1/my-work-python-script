import openpyxl as pyxl
from tkinter import Tk, filedialog
import pyautogui
from line_notify_me.line_notify_sourcecode import notifyme
import pyperclip

root = Tk()
root.excel = filedialog.askopenfilename(title="เลือกไฟล์ Excel", filetypes=(('Excel Files','*.xlsx'),('All Files', '*.*')))
root.withdraw()
wb = pyxl.load_workbook(root.excel, data_only=True)
ws = wb.sheetnames
receive = "รับแล้ว"
bkk33 = wb[ws[6]]
bkk49 = wb[ws[8]]



class function_ITEC:
    def search_button(self):
        pyautogui.press('f12')
    def docref_button(self):
        pyautogui.press('f5')


def edit_insure():
    '''
    start the operation.
    '''
    def start(): #default behavior
        '''
        start the operation. Opening stockout page, set the date back 1 year and click at ID field ready for searching operation.
        '''
        pyautogui.moveTo(45,255)
        pyautogui.sleep(.7)
        pyautogui.leftClick()
        pyautogui.sleep(10)
        pyautogui.moveTo(1753,192)
        pyautogui.leftClick()
        pyautogui.moveTo(1752,215)
        pyautogui.doubleClick()
        pyautogui.leftClick()
        pyautogui.press('down')
        pyautogui.press('enter')
        pyautogui.sleep(1)
        pyautogui.moveTo(294,89)
        pyautogui.doubleClick()
        pyautogui.leftClick()

    start()
    def docref33bkk():
        for row in range(1, bkk33.max_row+1):
            date = bkk33.cell(row=row, column=1).value
            id = bkk33.cell(row=row, column=2).value
            branch = bkk33.cell(row=row, column=3).value
            zone = bkk33.cell(row=row, column=4).value
            if date:
                pyautogui.sleep(.7)
                pyautogui.moveTo(288,88)
                pyautogui.sleep(.7)
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                function_ITEC().search_button()
                pyautogui.press('enter')
                function_ITEC().docref_button()
                pyperclip.copy(receive)
                pyautogui.typewrite(f"{zone} ")
                pyautogui.hotkey('ctrl', 'v')
                pyautogui.typewrite(f" | {date}")
                pyautogui.press('tab')
                pyautogui.press('enter')
                pyautogui.press('enter')
            else: break

    def docref49bkk():
        for row in range(1, bkk49.max_row+1):
            date = bkk49.cell(row=row, column=1).value
            id = bkk49.cell(row=row, column=2).value
            branch = bkk49.cell(row=row, column=3).value
            zone = bkk49.cell(row=row, column=4).value
            if date:
                pyautogui.sleep(.7)
                pyautogui.moveTo(288,88)
                pyautogui.sleep(.7)
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                function_ITEC().search_button()
                pyautogui.press('enter')
                function_ITEC().docref_button()
                pyperclip.copy(receive)
                pyautogui.typewrite(f"{zone} ")
                pyautogui.hotkey('ctrl', 'v')
                pyautogui.typewrite(f" | {date}")
                pyautogui.press('tab')
                pyautogui.press('enter')
                pyautogui.press('enter')
            else: break
    docref49bkk()
    docref33bkk()
    notifyme('Docref for BKK is finished!')

if __name__ in '__main__':
    edit_insure()