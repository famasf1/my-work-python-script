import openpyxl as pyxl
from tkinter import Tk, filedialog
import pyautogui
from line_notify_me.line_notify_sourcecode import notifyme
import pyperclip
import datetime
import os

class function_ITEC:
    def search_button(self):
        pyautogui.press('f12')
    def docref_button(self):
        pyautogui.press('f5')


def edit_insure():

    root = Tk()
    root.excel = filedialog.askopenfilename(title="เลือกไฟล์ Excel", filetypes=(('Excel Files','*.xlsx'),('All Files', '*.*')))
    root.withdraw()
    wb = pyxl.load_workbook(root.excel, data_only=True)
    ws = wb.sheetnames
    main_sheet = wb[ws[0]]
    receive = 'พัสดุถึง '
    receive_retdup = 'พัสดุ RET ถึง'

    '''
    start the operation.
    '''
    def start(): #default behavior
        '''
        start the operation. Opening stockout page, set the date back 1 year and click at ID field ready for searching operation.
        '''
        pyautogui.moveTo(45,255)
        pyautogui.sleep(.7)
        pyautogui.leftClick()
        pyautogui.sleep(10)
        pyautogui.moveTo(1753,192)
        pyautogui.leftClick()
        pyautogui.moveTo(1752,215)
        pyautogui.doubleClick()
        pyautogui.leftClick()
        pyautogui.press('down')
        pyautogui.press('enter')
        pyautogui.sleep(1)
        pyautogui.moveTo(294,89)
        pyautogui.doubleClick()
        pyautogui.leftClick()

    start()
    for row in range(2, main_sheet.max_row+1):
        date = main_sheet.cell(row=row, column=7).value
        id = main_sheet.cell(row=row, column=12).value
        branch = main_sheet.cell(row=row, column=13).value
        checkretdup = main_sheet.cell(row=row, column=14).value

        
        if id:
            pyautogui.sleep(.7)
            pyautogui.moveTo(288,88)
            pyautogui.sleep(.7)
            pyautogui.doubleClick()
            pyautogui.typewrite(str(id))
            pyautogui.press('enter')
            pyautogui.typewrite(str(branch))
            function_ITEC().search_button()
            pyautogui.press('enter')
            function_ITEC().docref_button()
            pyautogui.hotkey('ctrl','c')
            readData = pyperclip.paste()
            if "SVCOM7" in str(readData):
                pyperclip.copy('')
                pass
            elif "svcom7" in str(readData):
                pyperclip.copy('')
                pass
            else:
                if checkretdup == "1RETDUP":
                    pyperclip.copy(receive_retdup)
                else:
                    pyperclip.copy(receive)
                pyautogui.hotkey('ctrl','v')
                pyautogui.typewrite(f"{getdate_Obj(str(date))} ")
            pyautogui.hotkey('alt','o')
            pyautogui.press('enter')
        else: break
    notifyme('Docref is finished!')


### Converting Date
def getdate_Obj(dateData):
    try:
        date_obj = datetime.datetime.strptime(dateData, "%Y-%m-%d %H:%M:%S.%f")
    except:
        try:
            date_obj = datetime.datetime.strptime(dateData, "%Y-%m-%d %H:%M:%S")
        except:
            try:
                date_obj = datetime.datetime.strptime(dateData, "%d-%m-%Y %H:%M:%S")
            except:
                try:
                    date_obj = datetime.datetime.strptime(dateData, "%d/%m/%Y")
                except:
                    try:
                        date_obj = datetime.datetime.strptime(dateData, "%d/%m/%Y, %H:%m:%S")
                    except Exception as e:
                        print(e)
                
    return date_obj.strftime('%d/%m/%y')






if __name__ in '__main__':
    try:
        edit_insure()
    except Exception as e:
        notifyme("Docref Failed")
        print(e)
    #os.system("shutdown /s /t 1")