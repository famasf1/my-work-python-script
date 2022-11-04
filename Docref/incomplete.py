import sys
sys.path.insert(1,"D:\Workstuff\my work script\line_notify_me")

from lib2to3.pgen2 import token
from tracemalloc import start
import pyautogui
import openpyxl
from tkinter import filedialog, messagebox
from tkinter import *
import pyperclip
import datetime
import requests

### Notify me when the script is completed to LINE.
def notifyme(confirmtext):
    """
    LINE Notify - Send text to my own line.
    parameter :
    confirmtext: str (required)
    """
    mytoken = 'kOcQyjPGgIAgTQ4qWjTlEJZFUj7GegzGefdDEiSsYJr'
    url = 'https://notify-api.line.me/api/notify'
    data = {
        'message' : confirmtext
    }
    options = {
        'Method' : 'POST',
        'Content-Type' : 'application/x-www-form-urlencoded',
        'Authorization' : f'Bearer {mytoken}',
    }
    response = requests.post(url=url, headers=options, data=data)
    print(response.status_code)
############################# Part 0 : Context Menu

############################# Part 1 : variable
def startup():
    try:
        global workbook, receive, Incomplete_DataSheet
        root = Tk()
        root.excel = filedialog.askopenfilename(initialdir='/',title='เลือกไฟล์ Excel สำหรับ Stock Out 33-insure', filetypes=(('Excel','*.xlsx'),('All Files','*.*')))
        workbook = openpyxl.load_workbook(root.excel, data_only=True)
        root.withdraw()
        worksheetname = workbook.sheetnames
        Incomplete_DataSheet = workbook[worksheetname[0]]
        todate = datetime.date.today()
        yesterdate = todate - datetime.timedelta(days=1)
        yesterdate_string = str(yesterdate.strftime("%d/%m/%y"))
        receive = "สินค้าไม่ครบ | "
    except Exception as e:
        messagebox.showerror('Python Error', f'{e}')
        exit()

### Converting Date
def getdate_Obj(dateData):
    try:
        date_obj = datetime.datetime.strptime(dateData, "%Y-%m-%d %H:%M:%S.%f")
    except:
        try:
            date_obj = datetime.datetime.strptime(dateData, "%Y-%m-%d %H:%M:%S")
        except:
            try:
                date_obj = datetime.datetime.strptime(dateData, "%d-%m-%Y %H:%M:%S")
            except:
                try:
                    date_obj = datetime.datetime.strptime(dateData, "%d/%m/%Y")
                except:
                    try:
                        date_obj = datetime.datetime.strptime(dateData, "%d/%m/%Y, %H:%m:%S")
                    except Exception as e:
                        print(e)
                
    return date_obj.strftime('%d/%m/%y')


################################ TEST ROOM ###############################

class function_ITEC:
    
    def search_button(self):
        pyautogui.press('f12')
    def docref_button(self):
        pyautogui.press('f5')

#########################################################################
################################ Part 2: Reading excel and Start writing


################################# Part 2 : Function 

def docref(): 
    def defaultref(): #default behavior
        pyautogui.moveTo(45,255)
        pyautogui.sleep(.7)
        pyautogui.leftClick()
        pyautogui.sleep(3)
        pyautogui.moveTo(1753,192)
        pyautogui.leftClick()
        pyautogui.moveTo(1752,215)
        pyautogui.doubleClick()
        pyautogui.leftClick()
        pyautogui.press('down')
        pyautogui.press('enter')
        pyautogui.sleep(1)
        pyautogui.moveTo(294,89)
        pyautogui.doubleClick()
        pyautogui.leftClick()
    
    def docref49(): # 49 Trade in \ Bangkok
        for i in range(2,Incomplete_DataSheet.max_row+1):
            id = Incomplete_DataSheet.cell(row=i, column=1).value
            branch = Incomplete_DataSheet.cell(row=i, column=2).value
            reason = Incomplete_DataSheet.cell(row=i, column=3).value
            if id:
                pyautogui.sleep(.7)
                pyautogui.moveTo(278,87)
                pyautogui.sleep(.7)
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                function_ITEC().search_button()
                pyautogui.press('enter')
                function_ITEC().docref_button()
                pyperclip.copy(reason)
                pyautogui.hotkey('ctrl','v')
                pyautogui.typewrite(f" {getdate_Obj(str(datetime.datetime.today()))})")
                pyautogui.press('tab')
                pyautogui.press('enter')
                pyautogui.press('enter')
            else: break
        print("docref finished")
    docref49()

try:
    startup()
    docref()
except Exception as e:
    messagebox.showerror('Python Error',f'{e}')

print(f'Finished! Zero Error')
