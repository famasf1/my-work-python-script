#import

###############################
# WHAT DAY IS TODAY?
timedelta_date = 0
###############################

import pandas as pd
from tkinter import *
from tkinter import filedialog
import datetime
import openpyxl as pyxl
import dagdshort
import win32com.client
import os

today_DHL_mail = datetime.date.today() - datetime.timedelta(days=timedelta_date)
today_DHL_mail_month = today_DHL_mail.month
today_DHL_mail_year = int(today_DHL_mail.strftime("%y")) + 43
today_DHL_mail_strftime = today_DHL_mail.strftime("%d-%m-%y")
MAIN_FOLDER_PATH = fr"C:\Users\{os.getlogin()}\Documents\DHL"

def load_excel_from_mail():

    #Try changing dir to directory i want.
    #Except any error, then create that directory first.
    try:
        os.chdir(fr"C:\Users\{os.getlogin()}\Documents\DHL")
    except:
        os.chdir(fr"C:\Users\{os.getlogin()}\Documents")
        os.mkdir("DHL")
        os.chdir(fr"C:\Users\{os.getlogin()}\Documents\DHL")
    try:
        os.chdir(fr"C:\Users\{os.getlogin()}\Documents\DHL\dhl {today_DHL_mail_month}-{today_DHL_mail_year}")
    except:
        os.mkdir(fr"dhl {today_DHL_mail_month}-{today_DHL_mail_year}")
        os.chdir(fr"C:\Users\{os.getlogin()}\Documents\DHL\dhl {today_DHL_mail_month}-{today_DHL_mail_year}")

    outlook = win32com.client.dynamic.Dispatch("Outlook.Application").GetNamespace("MAPI")
    root_folder = outlook.GetDefaultFolder(6)
    messages = root_folder.Items
    subject = f"DHL - Com7_ShoptoShop_Signature_Report - {today_DHL_mail}"
    for m in messages:
        if m.Subject == subject:
            attachments = m.Attachments
            numattach = len([a for a in attachments])
            for attachment_excel in range(1, numattach+1):
                attachment = attachments.Item(attachment_excel)
                if (attachment.FileName).endswith("xlsx"):
                    try:
                        download_path = os.path.join(f"{MAIN_FOLDER_PATH}",f"dhl {today_DHL_mail_month}-{today_DHL_mail_year}", f"dhl {today_DHL_mail}.xlsx" )
                        attachment.SaveAsFile(download_path)
                    except FileExistsError:
                        print("File Exist")
                        pass

def read():

    '''
    Load Excel with Pandas, find and replace PHYIDINSURE | PHYID with empty whitespace. Then split out the data into array with '-'
    once that's done, combine frame and send all the data to 'whatday' function to process into Excel file. 
    '''

    #read
    global sheet
    
    #root = Tk()
    #root.excel = filedialog.askopenfilename(title='Open DHL file',filetypes=[('Excel Files', '*.xlsx'), ('All Files' , '*.*')]) #Load Excel
    
    sheet = pd.read_excel(os.path.join(f"{MAIN_FOLDER_PATH}",f"dhl {today_DHL_mail_month}-{today_DHL_mail_year}", f"dhl {today_DHL_mail}.xlsx"))
    remove_word = sheet['CCN'].replace(['PHYIDINSURE','PHYID'],'', regex=True).str.split('-', expand=True)
    rw_df = pd.concat([sheet, remove_word.iloc[:, :3]], axis=1)


    #rw_df2 = rw_df[['ID','Branch','Box Num','etc']] = pd.DataFrame(rw_df.CCN.to_list(), index=rw_df.index)
    #except ValueError:
    #    rw_df2 = rw_df[['ID','Branch','Box Num', 'etc1', 'etc2']] = pd.DataFrame(rw_df.CCN.to_list(), index=rw_df.index)
    #except:
    #    rw_df2 = rw_df[['ID','Branch','Box Num']] = pd.DataFrame(rw_df.CCN.to_list(), index=rw_df.index)

    #combine frame
    rw_df = rw_df.rename(columns={0 : 'ID', 1 : 'Branch', 2 : 'Box Num'})
    sheet = rw_df
    #rw_df.to_clipboard(sep='\t')
    insure_only()
    FCB_only()
    FCB_only_out()
    whatday()



def insure_only():
    '''
    After read() function finished. Take data from loaded Excel sheet and flitered out only value that match 'PHYIDINSURE', '33-1' using regex.
    then replace all the 'PHYIDINSURE' into blank, split '-' and return as Pandas Dataframe before sent it to whatday() function to process.
    '''

    #filter only insure from 33 to customer 
    global sheet_insure_only
    CCN_Insure = sheet[sheet['CCN'].str.contains(r"(^PHYIDINSURE.*)-(33-1)$",regex=True)]
    remove_word = CCN_Insure["CCN"].replace(['PHYIDINSURE'],'', regex=True).str.split('-')

    df_remove = pd.DataFrame(remove_word)
    #try:
    #    df_remove_as_table = df_remove[['ID','Branch','Box Num']] = pd.DataFrame(df_remove.CCN.to_list(), index=df_remove.index)
    #except:
    #    pass
    frames = [CCN_Insure, df_remove]
    sheet_insure_only = pd.concat(frames, axis=1)
    

def FCB_only():

    '''
    Separate only FCB Branch
    '''
    #filter only FCB branch
    global sheet_fcb_only

    only_fcb = sheet[sheet["POD"].str.contains("^FCB", regex=True)]
    only_fcb_phyid = only_fcb["CCN"].replace(['PHYID'],'',regex=True).str.split('-')

    #only_fcb_out = sheet[sheet["Branch"].str.contains("(^2)...", regex=True, na=False)]
    #only_fcb_out_phyid = only_fcb_out["CCN"].replace(['PHYID'],'',regex=True).str.split('-')

    df_only_fcb = pd.DataFrame(only_fcb_phyid)
    #df_only_fcb_out = pd.DataFrame(only_fcb_out_phyid)

    #try:
    #    df_only_fcb_as_table = df_only_fcb[['ID','Branch','Box Num']] = pd.DataFrame(df_only_fcb.CCN.to_list(), index=df_only_fcb.index)
    #except:
    #    pass

    frames = [only_fcb, df_only_fcb] #df_only_fcb_out]
    sheet_fcb_only = pd.concat(frames, axis=1)

def FCB_only_out():

    '''
    Separate only FCB Branch
    '''
    #filter only FCB branch
    global sheet_fcb_out_only


    only_fcb_out = sheet[sheet["Branch"].str.contains("(^2)...", regex=True, na=False)]
    only_fcb_out_phyid = only_fcb_out["CCN"].replace(['PHYID'],'',regex=True).str.split('-')

    df_only_fcb_out = pd.DataFrame(only_fcb_out_phyid)

    #try:
    #    df_only_fcb_as_table = df_only_fcb[['ID','Branch','Box Num']] = pd.DataFrame(df_only_fcb.CCN.to_list(), index=df_only_fcb.index)
    #except:
    #    pass

    frames = [only_fcb_out, df_only_fcb_out]
    sheet_fcb_out_only = pd.concat(frames, axis=1)


def whatday():

    '''
    Take data from 'read' function. Check whatday variable setting before export Pandas dataframe into Excel.
    '''

    global name
    name = f'DHL {today_DHL_mail_strftime}'
    sheet.to_excel(fr'C:\Users\Comseven\Documents\DHL\Completed\{name}.xlsx',index=False)
    sheet_insure_only.to_excel(fr'C:\Users\Comseven\Documents\DHL\Completed\{name}_INSURE_ONLY.xlsx',index=False)
    sheet_fcb_only.to_excel(fr'C:\Users\Comseven\Documents\DHL\Completed\{name}_FCB_ONLY.xlsx',index=False)
    with pd.ExcelWriter(fr'C:\Users\Comseven\Documents\DHL\Completed\{name}_FCB_ONLY.xlsx', engine="openpyxl", mode="a", if_sheet_exists="overlay") as file:
        sheet_fcb_out_only.to_excel(file, index=False, header=None, startrow=file.sheets['Sheet1'].max_row)
    

    bitly_api_activate()

def bitly_api_activate():

    '''
    Get data from selected row where URL link is located. Then convert them into shortlinks.
    '''

    dagd_connect = dagdshort.Shortener(user_agent_suffix='famasf1/dagd_shorten', max_cache_size=256)
    #network_path_test = '\\10.100.101.200\Scan_Service\jirayuth_stuff\DHL\dhl 1-66'
    load_insure_wb = pyxl.load_workbook(fr'C:\Users\Comseven\Documents\DHL\Completed\{name}_INSURE_ONLY.xlsx', data_only=True)
    #load_insure_wb = pyxl.load_workbook(fr'\\10.100.101.200\Scan_Service\jirayuth_stuff\DHL\Completed\{name}_INSURE_ONLY.xlsx', data_only=True)
    load_fcb_only = pyxl.load_workbook(fr'C:\Users\Comseven\Documents\DHL\Completed\{name}_FCB_ONLY.xlsx', data_only=True)
    #load_fcb_only = pyxl.load_workbook(fr'\\10.100.101.200\Scan_Service\jirayuth_stuff\DHL\Completed\{name}_FCB_ONLY.xlsx', data_only=True)
    load_insure_wb_sh = load_insure_wb.sheetnames
    load_fcb_wb_sheet = load_fcb_only.sheetnames
    active_Sheet = load_insure_wb[load_insure_wb_sh[0]]
    active_fcb_sheet = load_fcb_only[load_fcb_wb_sheet[0]]

    #for customer insure sheet
    def customer_link():
        for row in range(2, active_Sheet.max_row+1):

            long_link_list = []
            long_link = active_Sheet.cell(row=row, column=11).value
            long_link_list.append(long_link)
        
            
            try:
                short_link = dagd_connect.shorten_urls(long_link_list)
                short_link_result = list(short_link.values())
                active_Sheet.cell(row=row, column=15).value = str(short_link_result[0])
            except:
                active_Sheet.cell(row=row, column=15).value = ""
        load_insure_wb.save(fr'C:\Users\Comseven\Documents\DHL\Completed\{name}_INSURE_ONLY_1.xlsx')

    #for fcb sheet
    def fcb_link():
        for row in range(2,active_fcb_sheet.max_row+1):
            
            long_link_list = []
            long_link = active_fcb_sheet.cell(row=row, column=11).value
            long_link_list.append(long_link)
        
            try:
                short_link = dagd_connect.shorten_urls(long_link_list)
                short_link_result = list(short_link.values())
                active_fcb_sheet.cell(row=row, column=15).value = str(short_link_result[0])
            except:
                active_fcb_sheet.cell(row=row, column=15).value = ""
        load_fcb_only.save(fr'C:\Users\Comseven\Documents\DHL\Completed\{name}_FCB_ONLY_1.xlsx')

    customer_link()
    fcb_link()        
            
def sendmailpodfile():
    
    '''
    Find only shipment that manually book by hand, POD to headoffice.
    Then create excel file contains only that specific data.
    then send them to everyone who [might] needs it.
    '''
    global sheet_com7pod_only
    #regex
    com7list_new = ['COM7', 'Com7', 'com7', 'คลัง 49', 'สินค้าคืน', 'คุณแจง', 'พี่บล']
    com7pattern = '|'.join(com7list_new)
    sheet_com7pod_only = sheet[sheet['POD'].str.contains(com7pattern, regex=True)]
    if sheet_com7pod_only.empty:
        print("It's empty!")
        pass
    else:
        sheet_com7pod_only.to_excel(fr'C:\Users\Comseven\Documents\DHL\Completed\{name}_POD_TO_HEADOFFICE.xlsx', index=False)
        '''
        After getting that podfile, send mail to anyone responsible.
        '''
        outlook = win32com.client.dynamic.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        From = None
        for myEmailAddress in outlook.Session.Accounts:
            if "jirayuth.p@comseven.com" in str(myEmailAddress):
                From = myEmailAddress
                break
        mail.Subject = f"รายงานสินค้าตีกลับ - Book มือประจำวันที่ {today_DHL_mail_strftime}"
        mail.To = "jambo5167@gmail.com; Pairin@COMSEVEN.COM; pitchsukran.p@comseven.com; pawarisa.k@comseven.com; jambo5167@gmail.com"
        mail.Attachments.Add(fr'C:\Users\{os.getlogin()}\Documents\DHL\Completed\DHL {today_DHL_mail_strftime}_POD_TO_HEADOFFICE.xlsx')
        html = f'''
        <html> \
        <body> \
        <p>ขออนุญาตินำส่งไฟล์รายงานพร้อมลายเซ็นสินค้าประเภทตีกลับ - บุ๊คมือ ประจำวันที่ {today_DHL_mail_strftime}</p>
        <p>***อีเมลนี้เป็นอีเมลอัตโนมัติ</p>
        <p>ขอบคุณครับ</p>
        </body> \
        </html>
        '''
        mail.HTMLBody = html
        if From != None:
        # This line basically calls the "mail.SendUsingAccount = xyz@email.com" outlook VBA command
            mail._oleobj_.Invoke(*(64209, 0, 8, 0, From))
        mail.Send()

if __name__ in '__main__':
    load_excel_from_mail()
    read()
    sendmailpodfile()
    #com7podonly()
    #bitly_api_activate_once()


##############################################################################################################################



##### LEGACY ####

def bitly_api_activate_once():

    '''
    Get data from selected row where URL link is located. Then convert them into shortlinks.
    ***THIS IS MEANT TO BE USE ONCE***
    '''

    dagd_connect = dagdshort.Shortener(user_agent_suffix='famasf1/dagd_shorten', max_cache_size=256)

    load_insure_wb = pyxl.load_workbook(fr"C:\Users\Comseven\Documents\DHL\Completed\11-65\All_Nov_FCB.xlsx", data_only=True)
    load_insure_wb_sh = load_insure_wb.sheetnames
    active_Sheet = load_insure_wb[load_insure_wb_sh[0]]
    for row in range(2, active_Sheet.max_row+1):

        long_link_list = []
        long_link = active_Sheet.cell(row=row, column=11).value
        long_link_list.append(long_link)
    
        short_link = dagd_connect.shorten_urls(long_urls=long_link_list)
        short_link_result = list(short_link.values())
        active_Sheet.cell(row=row, column=15).value = str(short_link_result[0])
        load_insure_wb.save(fr"C:\Users\Comseven\Documents\DHL\Completed\11-65\All_Nov_FCB_1.xlsx")

def network_path_test(path):
    test_wb = pyxl.load_workbook(path, data_only=True)
    test_ws = test_wb.sheetnames
    sheet1 = test_wb[test_ws[0]]
    print(sheet1.cell(row=2, column=1).value)
