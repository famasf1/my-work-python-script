import sys
sys.path.insert(1,"D:\Workstuff\my work script\line_notify_me")

from lib2to3.pgen2 import token
from tracemalloc import start
import pyautogui
import openpyxl
from tkinter import filedialog, messagebox
from tkinter import *
import pyperclip
import datetime
import requests

### Notify me when the script is completed to LINE.
def notifyme(confirmtext):
    """
    LINE Notify - Send text to my own line.
    parameter :
    confirmtext: str (required)
    """
    mytoken = 'kOcQyjPGgIAgTQ4qWjTlEJZFUj7GegzGefdDEiSsYJr'
    url = 'https://notify-api.line.me/api/notify'
    data = {
        'message' : confirmtext
    }
    options = {
        'Method' : 'POST',
        'Content-Type' : 'application/x-www-form-urlencoded',
        'Authorization' : f'Bearer {mytoken}',
    }
    response = requests.post(url=url, headers=options, data=data)
    print(response.status_code)
############################# Part 0 : Context Menu

############################# Part 1 : variable
def startup():
    try:
        global workbook
        root = Tk()
        root.excel = filedialog.askopenfilename(initialdir='/',title='เลือกไฟล์ Excel สำหรับ Stock Out 33-insure', filetypes=(('Excel','*.xlsx'),('All Files','*.*')))
        workbook = openpyxl.load_workbook(root.excel, data_only=True)
        root.withdraw()
        worksheetname = workbook.sheetnames
        global Insure_33_Data, Info_33_Data, ID49Tradein_BKK, ID49Tradein, yesterdate_string, yesterdate, receive
        Insure_33_Data = workbook[worksheetname[0]]
        Info_33_Data = workbook[worksheetname[7]]
        ID49Tradein_BKK = workbook[worksheetname[9]]
        ID49Tradein = workbook[worksheetname[2]]
        todate = datetime.date.today()
        yesterdate = todate - datetime.timedelta(days=1)
        yesterdate_string = str(yesterdate.strftime("%d/%m/%y"))
        receive = 'รับ'
    except Exception as e:
        messagebox.showerror('Python Error', f'{e}')
        exit()

### Converting Date
def getdate_Obj(dateData):
    try:
        date_obj = datetime.datetime.strptime(dateData, "%Y-%m-%d %H:%M:%S.%f")
    except:
        try:
            date_obj = datetime.datetime.strptime(dateData, "%Y-%m-%d %H:%M:%S")
        except:
            try:
                date_obj = datetime.datetime.strptime(dateData, "%d-%m-%Y %H:%M:%S")
            except:
                try:
                    date_obj = datetime.datetime.strptime(dateData, "%d/%m/%Y")
                except:
                    try:
                        date_obj = datetime.datetime.strptime(dateData, "%d/%m/%Y, %H:%m:%S")
                    except Exception as e:
                        print(e)
                
    return date_obj.strftime('%d/%m/%y')


################################ TEST ROOM ###############################

class function_ITEC:
    
    def search_button(self):
        pyautogui.press('f12')
    def docref_button(self):
        pyautogui.press('f5')

#########################################################################
################################ Part 2: Reading excel and Start writing


################################# Part 2 : Function 

def docref(): 
    def defaultref(): #default behavior
        pyautogui.moveTo(45,255)
        pyautogui.sleep(1)
        pyautogui.leftClick()
        pyautogui.sleep(3)
        pyautogui.moveTo(1753,192)
        pyautogui.leftClick()
        pyautogui.moveTo(1752,215)
        pyautogui.doubleClick()
        pyautogui.leftClick()
        pyautogui.press('down')
        pyautogui.press('enter')
        pyautogui.sleep(3)
        pyautogui.moveTo(294,89)
        pyautogui.doubleClick()
        pyautogui.leftClick()

    def docref49():
        for i in range(1,ID49Tradein.max_row+1):
            out_sect = ID49Tradein.cell(row=i,column=15).value
            id = ID49Tradein.cell(row=i,column=12).value
            branch = ID49Tradein.cell(row=i,column=13).value
            date = ID49Tradein.cell(row=i, column=7).value
            formulae = f"=ifna(VLOOKUP(M{i},Data!C:G,5,0),"")"
            Insure_33_Data.cell(row=i,column=15).value = formulae
            workbook.save('bitly+ready.xlsx')
            if out_sect:
                pyautogui.sleep(3)
                pyautogui.moveTo(288,88)
                pyautogui.sleep(1)
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                function_ITEC().search_button()
                pyautogui.press('enter')
                function_ITEC().docref_button()
                pyautogui.hotkey('ctrl','c')
                readData = pyperclip.paste()
                if "SVCOM7" in str(readData):
                    pyperclip.copy('')
                    pass
                elif "svcom7" in str(readData):
                    pyperclip.copy('')
                    pass
                else:
                    pyperclip.copy(receive)
                    pyautogui.hotkey('ctrl','v')
                    pyautogui.typewrite(f"{getdate_Obj(str(date))} | {out_sect}")
                pyautogui.press('tab')
                pyautogui.press('enter')
                pyautogui.press('enter')
            else: break 
        notifyme('Docref166 Finished!')

    def docref166bkk(): #33 ITEC Insure \ Bangkok
        for i in range(1,Info_33_Data.max_row+1):
            date = Info_33_Data.cell(row=i,column=1).value
            id = Info_33_Data.cell(row=i,column=2).value
            branch = Info_33_Data.cell(row=i,column=3).value
            zone = Info_33_Data.cell(row=i,column=4).value
            if date:
                pyautogui.sleep(1)
                pyautogui.moveTo(278,87)
                pyautogui.sleep(1)
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                function_ITEC().search_button()
                pyautogui.press('enter')
                function_ITEC().docref_button()
                pyperclip.copy(receive)
                pyautogui.hotkey('ctrl','v')
                pyautogui.typewrite(f"{getdate_Obj(str(date))} | ")
                pyperclip.copy(str(zone))
                pyautogui.hotkey('ctrl','v')
                pyautogui.press('tab')
                pyautogui.press('enter')
                pyautogui.press('enter')
            else: break
        print("docref166 finishied! next : docref49tradeinbkk")
        docref49tradeinbkk()

    def docref166(): # 33 ITEC Insure
        defaultref()
        for i in range(1,Insure_33_Data.max_row+1):
            out_sect = Insure_33_Data.cell(row=i,column=15).value
            id = Insure_33_Data.cell(row=i,column=12).value
            branch = Insure_33_Data.cell(row=i,column=13).value
            date = Insure_33_Data.cell(row=i, column=7).value
            formulae = f"=ifna(VLOOKUP(M{i},Data!C:G,5,0),"")"
            Insure_33_Data.cell(row=i,column=15).value = formulae
            workbook.save('bitly+ready.xlsx')
            if out_sect:
                pyautogui.sleep(3)
                pyautogui.moveTo(288,88)
                pyautogui.sleep(1)
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                function_ITEC().search_button()
                pyautogui.press('enter')
                function_ITEC().docref_button()
                pyperclip.copy(receive)
                pyautogui.hotkey('ctrl','v')
                pyautogui.typewrite(f"{getdate_Obj(str(date))} | {out_sect}")
                pyautogui.press('tab')
                pyautogui.press('enter')
                pyautogui.press('enter')
            else: break
        print("docref166 finished. up next, docref166bkk")
        docref166bkk()
    
    def docref49tradeinbkk(): # 49 Trade in \ Bangkok
        for i in range(1,ID49Tradein_BKK.max_row+1):
            date = ID49Tradein_BKK.cell(row=i,column=1).value
            id = ID49Tradein_BKK.cell(row=i,column=2).value
            branch = ID49Tradein_BKK.cell(row=i,column=3).value
            zone = ID49Tradein_BKK.cell(row=i,column=4).value
            if date:
                pyautogui.sleep(1)
                pyautogui.moveTo(278,87)
                pyautogui.sleep(1)
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                function_ITEC().search_button()
                pyautogui.press('enter')
                function_ITEC().docref_button()
                pyperclip.copy(receive)
                pyautogui.hotkey('ctrl','v')
                pyautogui.typewrite(f"{getdate_Obj(str(date))} | ")
                pyperclip.copy(str(zone))
                pyautogui.hotkey('ctrl','v')
                pyautogui.press('tab')
                pyautogui.press('enter')
                pyautogui.press('enter')
            else: break
        print("docref49tradeinbkk finished. next up, docref49!")
        docref49()
    docref166()


try:
    startup()
    docref()
except Exception as e:
    messagebox.showerror('Python Error',f'{e}')

print(f'Finished! Zero Error')
