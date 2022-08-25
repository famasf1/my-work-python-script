import pyautogui
import openpyxl
from tkinter import filedialog, messagebox
from tkinter import *
import datetime
import pyperclip
import requests

try:
    root = Tk()
    root.excel = filedialog.askopenfilename(initialdir='/Desktop',title='เลือกไฟล์ Excel สำหรับ Stock-Out49-Insure', filetypes=(('Excel','*.xlsx'),('All Files','*.*')))
    workbook = openpyxl.load_workbook(root.excel, data_only=True)
    root.withdraw()
    worksheetname = workbook.sheetnames
    greenBox_Data = workbook[worksheetname[0]]
    receive = 'รับ'
except Exception as e:
    messagebox.showerror('Python Error', f'{e}')
    exit()

def notifyme(confirmtext):
    mytoken = 'kOcQyjPGgIAgTQ4qWjTlEJZFUj7GegzGefdDEiSsYJr'
    url = 'https://notify-api.line.me/api/notify'
    data = {
        'message' : confirmtext
    }
    options = {
        'Method' : 'POST',
        'Content-Type' : 'application/x-www-form-urlencoded',
        'Authorization' : f'Bearer {mytoken}',
    }
    response = requests.post(url=url, headers=options, data=data)

def getdate_Obj(dateData):
    try:
        date_obj = datetime.datetime.strptime(dateData, "%Y-%m-%d %H:%M:%S.%f")
    except:
        try:
            date_obj = datetime.datetime.strptime(dateData, "%Y-%m-%d %H:%M:%S")
        except:
            try:
                date_obj = datetime.datetime.strptime(dateData, "%d-%m-%Y %H:%M:%S")
            except:
                try:
                    date_obj = datetime.datetime.strptime(dateData, "%d/%m/%Y")
                except Exception as e:
                    print(e)
                
    return date_obj.strftime('%d/%m/%y')

def docref():
    def defaultref(): #Do this everytime i want to start.
        pyautogui.moveTo(45,255)
        pyautogui.sleep(1)
        pyautogui.leftClick()
        pyautogui.sleep(3)
        pyautogui.moveTo(1432,192)
        pyautogui.leftClick()
        pyautogui.moveTo(1427,215)
        pyautogui.doubleClick()
        pyautogui.leftClick()
        pyautogui.press('down')
        pyautogui.press('enter')
        pyautogui.sleep(3)
        pyautogui.moveTo(278,87)
        pyautogui.doubleClick()
        pyautogui.leftClick()
    def docrefgreenbox(): # 49 Trade in \ Bangkok
        for i in range(1,greenBox_Data.max_row+1):
            date = greenBox_Data.cell(row=i,column=1).value
            id = greenBox_Data.cell(row=i,column=2).value
            branch = greenBox_Data.cell(row=i,column=3).value
            boxid = greenBox_Data.cell(row=i,column=4).value
            if date:
                pyautogui.sleep(1)
                pyautogui.moveTo(278,87)
                pyautogui.sleep(1)
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                pyautogui.moveTo(1554,54)
                pyautogui.leftClick()
                pyautogui.moveTo(345,56)
                pyautogui.leftClick()
                pyperclip.copy(str(boxid))
                pyautogui.hotkey('ctrl','v')
                pyautogui.typewrite(f"/{getdate_Obj(str(date))}")
                pyautogui.moveTo(701,483)
                pyautogui.leftClick()
                pyautogui.press('enter')
            else: break
    defaultref()
    docrefgreenbox()

try:
    docref()
    notifyme('ใส่ข้อมูลกล่องเขียวสำเร็จ')
except Exception as e:
    messagebox.showerror('Python Error',f'{e}')