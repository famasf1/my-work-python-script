from line_notify_me.line_notify_sourcecode import notifyme
import pyautogui as pyg
import openpyxl
from tkinter import filedialog
from tkinter import *
import pyperclip

def main(): 
    root = Tk()
    root.excel = filedialog.askopenfilename(title='Open Excel', filetypes=([( 'Excel Files','*.xlsx',),( 'All Files','*.*',)]))
    root.withdraw()
    workbook = openpyxl.load_workbook(root.excel)
    sheet = workbook.sheetnames
    default_sheet = workbook[sheet[0]]
    
    def only_first():
        '''
        Run only for first row of data
        '''
        
        pyg.click(306,139) ##only for first time
        pyg.sleep(3)
        pyg.click(602,297)

    def aging_Cat():
        
        '''
        first check if next row in col A is empty.

        if empty = not next product yet. Do the loop.

        in the loop, copy a row from centralized page. then search if
        it match first value in col B.

        if it's not match, try next row in col B.
        if match, insert number.
        
        '''

        pyg.write(branch)
        pyg.press('tab',3)
        pyg.write(productcode)
        pyg.press('f12')
        pyg.sleep(1.5)
        pyg.press('enter')
        pyg.sleep(1.5)
        if pyg.locateOnScreen(rf"D:\Workstuff\my-work-python-script\rotate\asset\ret_error.png",grayscale=True):
            pyg.press('enter')
            default_sheet.cell(row=product, column=4).value = 'Failed'
            pass
        else:
            pyg.press('enter')
        pyg.sleep(1.5)
        pyg.write('49')
        pyg.hotkey('alt','f')
        pyg.sleep(1)
        pyg.press('tab',11)
        pyg.hotkey('ctrl','c')
        val = pyperclip.paste()
        if val != number:
            pyg.write(str(number))
        else:
            pass
        pyg.hotkey('alt','s')
        pyg.sleep(2)
        if pyg.locateCenterOnScreen(rf"D:\Workstuff\my-work-python-script\rotate\asset\ret_error.png",grayscale=True):
            pyg.press('enter')
            pyg.hotkey('alt','x')
            default_sheet.cell(row=product, column=4).value = 'Failed'
        else:
            pyg.press('y')
            pyg.press('enter')
            default_sheet.cell(row=product, column=4).value = 'Success'
        workbook.save(root.excel)

    only_first()
    #first loop for main page
    for product in range(2,default_sheet.max_row+1):
        productcode = default_sheet.cell(row=product, column=1).value
        branch = default_sheet.cell(row=product, column=2).value
        number = default_sheet.cell(row=product, column=3).value
        success = default_sheet.cell(row=product,column=4).value
        pyg.sleep(1)
        pyg.click(107,106,2)
            #first check if next row in col A is empty
            #if empty = not next product yet. Do the loop.
            #in the loop, copy a row from centralized page. then search if
            # it match first value in col B
            # if it's not match, try next row in col B
            # if match, insert number
        if productcode:
            aging_Cat()
        ###########################################################
    notifyme('Rotate finished!')
        #operation success. save it and break to start next iter

def test_room():
    root = Tk()
    root.excel = filedialog.askopenfilename(title='Open Excel', filetypes=([( 'Excel Files','*.xlsx',),( 'All Files','*.*',)]))
    root.withdraw()
    workbook = openpyxl.load_workbook(root.excel)
    sheet = workbook.sheetnames
    default_sheet = workbook[sheet[1]]

    for i in range(3,default_sheet.max_row+1):
        productcode = default_sheet.cell(row=i, column=1).value
        branch = default_sheet.cell(row=i, column=2).value
        number = default_sheet.cell(row=i, column=4).value
        val = pyperclip.paste() 
        while True:
            if productcode == None:
                i += 1
                break
            else:
                print(productcode)
                print(i)
                break
        
def test_room2():
    val = 7
    number = 2
    if pyperclip.paste() != number:
        pyg.write(str(number))
    else:
        pass



if __name__ in '__main__':
    #test_room2()
    main()
