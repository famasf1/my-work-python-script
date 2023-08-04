from operator import index
from random import *
from line_notify_me.line_notify_sourcecode import notifyme
from tkinter import filedialog, messagebox
from tkinter import *
import pyautogui as pyg
import openpyxl
import pyperclip
import numpy as np
############### CLASS ################

###MAKE SURE YOU CHECK CUSTOM EMPLOYEE VARIABLE FIRST


### List all employee ###
class Employeelist:
    #เขียนครบ ตามด้วยไอดีพนักงาน
    def __init__(self, index, name, staffid=str):
        self.name = name
        self.staffid = staffid
        self.index = index
    #ขานชื่อ - เรียกชื่อจริง
    #ขานเลขที่ - เรียกเลขที่
    def ขานชื่อ(self):
        return self.name

    def ขานเลขพนักงาน(self):
        return self.staffid

    def get_index(self):
        return self.index

########################################

##### SETTING

def start_script():
    try:
        global workbook, worksheet,readpicerrorfound, ok, employeelist_index, customemployee_index, employeelist_name, employeelist_id, checking
        global จิรายุทธ, วรัญญู, วรวุฒิ, กิตติคุณ, สราวุธ, กิติพงษ์, ณัฐพงษ์
        ###################################### LIST EMPLOYEE #####################################
        จิรายุทธ = Employeelist(0,r'ครบ / โบ้ | ', '22608') 
        วรัญญู = Employeelist(1,r'ครบ / ตั้ม | ', '26425')
        วรวุฒิ = Employeelist(2, r'ครบ / ดิว | ','22073')
        กิตติคุณ = Employeelist(3,r'ครบ / ก็อต | ', '24021')
        สราวุธ = Employeelist(4,r'ครบ / เอก | ','23267')
        กิติพงษ์ = Employeelist(5, r'ครบ / กิต | ', '26308')
        ณัฐพงษ์ = Employeelist(6, r'ครบ / เพรช | ', '26181' )
        #ไพรินทร์ = Employeelist(7, 'ครบ / ไพริน', '1815')
        employeelist_index = [กิตติคุณ.index, สราวุธ.index, วรวุฒิ.index, ณัฐพงษ์.index, วรัญญู.index, กิติพงษ์.index]
        customemployee_index = [จิรายุทธ.index]
        ##########################################################################################

        root = Tk()
        root.excel = filedialog.askopenfilename(initialdir='/Desktop',title='เลือกไฟล์ Excel สำหรับ Stock-In', filetypes=(('Excel','*.xlsx'),('All Files','*.*')))

        workbook = openpyxl.load_workbook(root.excel, data_only=True)
        root.withdraw()
        sheet = workbook.sheetnames
        worksheet = workbook[sheet[0]] 
        readpicerrorfound = pyg.locateCenterOnScreen('asset/checkerror.png')
        ok = pyg.locateOnScreen('asset/ok.png')

    except Exception as e:
        messagebox.showerror('Python Error', f'{e}')
        exit()

############################
############################
############################
###### MAIN FUNCTION #######

def main(): #short for default behavior
    ### First time starting
    start_script()

## WTF?
## if i remove it, will it break?

    def firstStart(start):
        if start == 1:
            pyg.hotkey('alt','k')
            pyg.press('i')
        else:
            pass
    ### Next

    def nextStart(next):
        if next == 1:
            pass
        else:
            pyg.hotkey('alt','k')
            pyg.press('i')

    def start_in():
        for i in range(2, worksheet.max_row+1):
            
            stockoutid = worksheet.cell(row=i, column=1).value
            date = worksheet.cell(row=i, column=2).value
            custom = worksheet.cell(row=i, column=3).value
            etc = worksheet.cell(row=i, column=4).value
            if stockoutid:

                nextStart(i)
                
                firstStart(i)
                print('Start Stock In')

                #pick staff id
                r = np.random.default_rng()
                ###### employeelist_index
                #employeelist_index = [กิตติคุณ.index, สราวุธ.index, วรวุฒิ.index, ณัฐพงษ์.index, วรัญญู.index, กิติพงษ์.index]
                ###### customemployee_index

                #### 2/04

                #[สราวุธ.index, วรวุฒิ.index, วรัญญู.index, กิตติคุณ.index]
                if custom == "custom":
                    ###If someone take a day off, customize this.
                    #custom_prob = [.1,.25,.1,.15,.25,.15]
                    #custom_prob = np.array(custom_prob)
                    #custom_prob_sum = sum(custom_prob)
                    #a = 1/custom_prob_sum
                    #customprob_Scaled = [e*a for e in custom_prob]
                    rng = r.choice(customemployee_index,)# p=customprob_Scaled)
                else:
                    prob = [.15,.3,.15,.2,.15,.15]
                    prob = np.array(prob)
                    prob_sum = sum(prob)
                    b = 1/prob_sum
                    prob_Scaled = [e*b for e in prob]
                    rng = r.choice(employeelist_index, p=prob_Scaled)
                match rng:
                    case 0:
                        pyg.typewrite(จิรายุทธ.staffid)
                    case 1:
                        pyg.typewrite(วรัญญู.staffid)
                    case 2:
                        pyg.typewrite(วรวุฒิ.staffid)
                    case 3:
                        pyg.typewrite(กิตติคุณ.staffid)
                    case 4:
                        pyg.typewrite(สราวุธ.staffid)
                    case 5:
                        pyg.typewrite(กิติพงษ์.staffid)
                    case 6:
                        pyg.typewrite(ณัฐพงษ์.staffid)
                pressenter(2)
                pyg.typewrite(str(stockoutid)) #stockout
                pressenter(2)

                #name
                match rng:
                    case 0:
                        pyperclip.copy(จิรายุทธ.name)
                        pyg.hotkey('ctrl','v')
                    case 1:
                        pyperclip.copy(วรัญญู.name)
                        pyg.hotkey('ctrl','v')
                    case 2:
                        pyperclip.copy(วรวุฒิ.name)
                        pyg.hotkey('ctrl', 'v')
                    case 3:
                        pyperclip.copy(กิตติคุณ.name)
                        pyg.hotkey('ctrl', 'v')
                    case 4:
                        pyperclip.copy(สราวุธ.name)
                        pyg.hotkey('ctrl', 'v')
                    case 5:
                        pyperclip.copy(กิติพงษ์.name)
                        pyg.hotkey('ctrl', 'v')
                    case 6:
                        pyperclip.copy(ณัฐพงษ์.name)
                        pyg.hotkey('ctrl', 'v')
                        
                ##ready
                ##now check if value exist
                pyg.sleep(1.85)
                pyg.typewrite(f"Date : {date}")
                pyg.sleep(1
                )
                if etc:
                    pyperclip.copy(etc)
                    pyg.typewrite(" | ")
                    pyg.hotkey('ctrl', 'v')
                pyg.hotkey('alt','f')
                pyg.press('o')
                pyg.sleep(5)
                try:
                    if pyg.locateCenterOnScreen(r"D:\Workstuff\my-work-python-script\asset\ret_error.png", grayscale=True):
                        pressenter(1)
                        worksheet.cell(row=i, column=4).value = 'Failed'
                        pyg.press('esc')
                        pass
                    else:
                        pressenter(1)
                        pyg.press('left')
                        pressenter(1)
                        foundstockbill = None
                        while (foundstockbill == None):
                            try:
                                foundstockbill = pyg.locateCenterOnScreen(r"D:\Workstuff\my-work-python-script\asset\foundstockbill.png", grayscale=True, confidence=.77)
                            except Exception as e:
                                print(e)
                                continue
                        if foundstockbill:
                            worksheet.cell(row=i, column=4).value = 'Success'
                            pressenter(4)
                            continue
                    workbook.save(r"D:\Workstuff\my-work-python-script\Autostockin\stockin49_1.xlsx")
                except Exception as e:
                    messagebox.showerror('Python Error', f'{e}')
                    
            else: break

    start_in()
    notifyme('Stock In Complete!')

####################################################################
######### Extra Function that doesn't involve with any of the above

def custom_comment(comment):
    pyperclip.copy(comment)
    pyg.hotkey('ctrl', 'v')

def pressenter(numberoftimes):
    for i in range(0, numberoftimes):
        pyg.press('enter')
        if i == numberoftimes:
            break

def clickleft(numberoftimes):
    for i in range(0, numberoftimes):
        pyg.leftClick()
        pyg.sleep(1.5)
        if i == numberoftimes:
            break

def test():

    จิรายุทธ = Employeelist(0,'ครบ / โบ้', '22608') 
    วรัญญู = Employeelist(1,'ครบ / ตั้ม', '25175')
    วรวุฒิ = Employeelist(3, 'ครบ / ดิว','22073')
    กิตติคุณ = Employeelist(4,'ครบ / ก็อต', '24021')
    สราวุธ = Employeelist(5,'ครบ / เอก','23267')
    กิติคุณ = Employeelist(6, 'ครบ / กิต', '25120')
    employeelist_index = [วรัญญู.index, กิติคุณ.index, กิตติคุณ.index, สราวุธ.index, วรวุฒิ.index, จิรายุทธ.index]
    r = np.random.default_rng()
    #rng = r.choice(employeelist_index, p=[.200,.200,.200,.067,.067,.067,.199])

try:
    #test()
    main()

except Exception as e:
    messagebox.showerror('Python Error', f'{e}')
    exit()








































