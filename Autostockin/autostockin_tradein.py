from operator import index
from line_notify_me.line_notify_sourcecode import notifyme
from tkinter import filedialog, messagebox
from tkinter import *
import pyautogui as pyg
import openpyxl
import pyperclip
import random
############### CLASS ################

### List all employee ###
class Employeelist:
    #เขียนครบ ตามด้วยไอดีพนักงาน
    def __init__(self,index, name, staffid=str):
        self.name = name
        self.staffid = staffid
        self.index = index
    #ขานชื่อ - เรียกชื่อจริง
    #ขานเลขที่ - เรียกเลขที่
    def ขานชื่อ(self):
        return self.name

    def ขานเลขพนักงาน(self):
        return self.staffid

    def get_index(self):
        return self.index

########################################

##### SETTING

def start_script():
    try:
        global workbook, worksheet,readpicerrorfound, ok, employeelist_index, employeelist_name, employeelist_id, checking
        global จิรายุทธ, วรัญญู, มรกต, วุฒิภัทร, วรวุฒิ, กิตติคุณ, สราวุธ
        ###################################### LIST EMPLOYEE #####################################
        จิรายุทธ = Employeelist(0,'ครบ / โบ้', '22608') 
        วรัญญู = Employeelist(1,'ครบ / ตั้ม', '24179')
        มรกต = Employeelist(2,'ครบ / ปาน', '23947')
        วุฒิภัทร = Employeelist(3,'ครบ / มาร์ค','23800')
        วรวุฒิ = Employeelist(4, 'ครบ / ดิว','22073')
        กิตติคุณ = Employeelist(5,'ครบ / ก็อต', '24021')
        สราวุธ = Employeelist(6,'ครบ / เอก','23267')
        employeelist_index = [วรวุฒิ.index, กิตติคุณ.index, สราวุธ,index]
        ##########################################################################################

        root = Tk()
        root.excel = filedialog.askopenfilename(initialdir='/Desktop',title='เลือกไฟล์ Excel สำหรับ Stock-Out33', filetypes=(('Excel','*.xlsx'),('All Files','*.*')))

        workbook = openpyxl.load_workbook(root.excel, data_only=True)
        root.withdraw()
        sheet = workbook.sheetnames
        worksheet = workbook[sheet[0]] 
        readpicerrorfound = pyg.locateCenterOnScreen('asset/checkerror.png')
        ok = pyg.locateOnScreen('asset/ok.png')

    except Exception as e:
        messagebox.showerror('Python Error', f'{e}')
        exit()

############################
############################
############################
###### MAIN FUNCTION #######

def main(): #short for default behavior
    ### First time starting
    start_script()

## WTF?
## if i remove it, will it break?

    def firstStart(start):
        if start == 1:
            pyg.hotkey('alt','k')
            pyg.press('i')
        else:
            pass
    ### Next

    def nextStart(next):
        if next == 1:
            pass
        else:
            pyg.hotkey('alt','k')
            pyg.press('i')

    def start_in():
        for i in range(1, worksheet.max_row+1):
            stockoutid = worksheet.cell(row=i, column=1).value
            if stockoutid:
                nextStart(i)
                firstStart(i)
                print('Start Stock In')

                #pick staff id
                r = random.choice(employeelist_index)
                match r:
                    case 0:
                        pyg.typewrite(วรวุฒิ.staffid)
                    case 1:
                        pyg.typewrite(กิตติคุณ.staffid)
                    case 2:
                        pyg.typewrite(สราวุธ.staffid)
                pressenter(2)
                pyg.typewrite(str(stockoutid)) #stockout
                pressenter(2)

                #name
                match r:
                    case 0:
                        pyperclip.copy(วรวุฒิ.name)
                        pyg.hotkey('ctrl','v')
                    case 1:
                        pyperclip.copy(กิตติคุณ.name)
                        pyg.hotkey('ctrl','v')
                    case 2:
                        pyperclip.copy(สราวุธ.name)
                        pyg.hotkey('ctrl','v')
                ##ready
                ##now check if value exist
                pyg.hotkey('alt','f')
                pyg.press('o')
                pyg.sleep(5)
                try:
                    if pyg.locateCenterOnScreen(r"D:\Workstuff\my-work-python-script\asset\ret_error.png", grayscale=True):
                        pressenter(1)
                        worksheet.cell(row=i, column=2).value = 'Failed'
                        pyg.press('esc')
                        pass
                    else:
                        pressenter(1)
                        pyg.press('left')
                        pressenter(1)
                        foundstockbill = None
                        while (foundstockbill == None):
                            try:
                                foundstockbill = pyg.locateCenterOnScreen(r"D:\Workstuff\my-work-python-script\asset\foundstockbill.png", grayscale=True, confidence=.77)
                            except Exception as e:
                                print(e)
                                continue
                        if foundstockbill:
                            worksheet.cell(row=i, column=2).value = 'Success'
                            pressenter(4)
                            continue
                except Exception as e:
                    messagebox.showerror('Python Error', f'{e}')
            else: break

    start_in()
    workbook.save(r"D:\Workstuff\my-work-python-script\Autostockin\stockin49_1.xlsx")
    notifyme('Stock In Complete!')

####################################################################
######### Extra Function that doesn't involve with any of the above

def custom_comment(comment):
    pyperclip.copy(comment)
    pyg.hotkey('ctrl', 'v')

def pressenter(numberoftimes):
    for i in range(0, numberoftimes):
        pyg.press('enter')
        if i == numberoftimes:
            break

def clickleft(numberoftimes):
    for i in range(0, numberoftimes):
        pyg.leftClick()
        pyg.sleep(1.5)
        if i == numberoftimes:
            break

def test():
    จิรายุทธ = Employeelist(0,'ครบ / โบ้', 22608) 
    วรัญญู = Employeelist(1,'ครบ / ตั้ม', 24179)
    มรกต = Employeelist(2,'ครบ / ปาน', 23947)
    วุฒิภัทร = Employeelist(3,'ครบ / มาร์ค',23800)
    employeelist_index = [วรวุฒิ.index, กิตติคุณ.index, สราวุธ,index]
    r = random.choice(employeelist_index)
    print(r)

try:
    #test()
    main()
except Exception as e:
    messagebox.showerror('Python Error', f'{e}')
    exit()








































