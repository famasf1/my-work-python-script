import pyautogui
import openpyxl
from tkinter import filedialog, messagebox
from tkinter import *
import pyperclip
import datetime
############################# Part 0 : Context Menu

############################# Part 1 : variable
try:
    root = Tk()
    root.excel = filedialog.askopenfilename(initialdir='/',title='เลือกไฟล์ Excel สำหรับ Stock Out 33-insure', filetypes=(('Excel','*.xlsx'),('All Files','*.*')))
    workbook = openpyxl.load_workbook(root.excel, data_only=True)
    root.withdraw()
    worksheetname = workbook.sheetnames
    Insure_33_Data = workbook[worksheetname[0]]
    Info_33_Data = workbook[worksheetname[4]]
    ID49Tradein_BKK = workbook[worksheetname[7]]
    ID49Tradein = workbook[worksheetname[8]]
    todate = datetime.date.today()
    yesterdate = todate - datetime.timedelta(days=1)
    yesterdate_string = str(yesterdate.strftime("%d/%m/%y"))
    receive = 'รับ'
except Exception as e:
    messagebox.showerror('Python Error', f'{e}')
    exit()
################################# Part 2 : Function 
def docref(): 
    def defaultref(): #default behavior
        pyautogui.moveTo(45,255)
        pyautogui.sleep(1)
        pyautogui.leftClick()
        pyautogui.sleep(3)
        pyautogui.moveTo(1432,192)
        pyautogui.leftClick()
        pyautogui.moveTo(1434,215)
        pyautogui.doubleClick()
        pyautogui.leftClick()
        pyautogui.press('down')
        pyautogui.press('enter')
        pyautogui.sleep(3)
        pyautogui.moveTo(278,87)
        pyautogui.doubleClick()
        pyautogui.leftClick()

    def docref49():
        for i in range(1,ID49Tradein.max_row+1):
            out_sect = ID49Tradein.cell(row=i,column=15).value
            id = ID49Tradein.cell(row=i,column=12).value
            branch = ID49Tradein.cell(row=i,column=13).value
            date = ID49Tradein.cell(row=i, column=7).value
            try:
                date_obj = datetime.datetime.strptime(date, "%Y-%m-%d %H:%M:%S")
            except TypeError:
                pass
            formulae = f"=ifna(VLOOKUP(M{i},Data!C:G,5,0),"")"
            Insure_33_Data.cell(row=i,column=15).value = formulae
            workbook.save('bitly+ready.xlsx')
            if out_sect:
                pyautogui.sleep(3)
                pyautogui.moveTo(288,88)
                pyautogui.sleep(1)
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                pyautogui.moveTo(1554,54)
                pyautogui.leftClick()
                pyautogui.moveTo(345,56)
                pyautogui.leftClick()
                pyautogui.hotkey('ctrl','c')
                readData = pyperclip.paste()
                if "SVCOM7" in str(readData):
                    pyperclip.copy('')
                    pass
                else:
                    pyperclip.copy(receive)
                    pyautogui.hotkey('ctrl','v')
                    pyautogui.typewrite(f"{date_obj.strftime('%d/%m/%y')} | {out_sect}")
                pyautogui.moveTo(701,483)
                pyautogui.leftClick()
                pyautogui.press('enter')
            else: break 

    def docref166bkk(): #33 ITEC Insure \ Bangkok
        for i in range(1,Insure_33_Data.max_row+1):
            date = Info_33_Data.cell(row=i,column=2).value
            id = Info_33_Data.cell(row=i,column=5).value
            branch = Info_33_Data.cell(row=i,column=6).value
            zone = Info_33_Data.cell(row=i,column=9).value
            if date:
                pyautogui.sleep(1)
                pyautogui.moveTo(278,87)
                pyautogui.sleep(1)
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                pyautogui.moveTo(1554,54)
                pyautogui.leftClick()
                pyautogui.moveTo(345,56)
                pyautogui.leftClick()
                pyperclip.copy(receive)
                pyautogui.hotkey('ctrl','v')
                pyautogui.typewrite(f"{date} | ")
                pyperclip.copy(str(zone))
                pyautogui.hotkey('ctrl','v')
                pyautogui.moveTo(701,483)
                pyautogui.leftClick()
                pyautogui.press('enter')
            else: break
        docref49tradeinbkk()

    def docref166(): # 33 ITEC Insure
        defaultref()
        for i in range(1,Insure_33_Data.max_row+1):
            out_sect = Insure_33_Data.cell(row=i,column=15).value
            id = Insure_33_Data.cell(row=i,column=12).value
            branch = Insure_33_Data.cell(row=i,column=13).value
            date = Insure_33_Data.cell(row=i, column=7).value
            try:
                date_obj = datetime.datetime.strptime(date, "%Y-%m-%d %H:%M:%S")
            except TypeError:
                pass
            formulae = f"=ifna(VLOOKUP(M{i},Data!C:G,5,0),"")"
            Insure_33_Data.cell(row=i,column=15).value = formulae
            workbook.save('bitly+ready.xlsx')
            if out_sect:
                pyautogui.sleep(3)
                pyautogui.moveTo(288,88)
                pyautogui.sleep(1)
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                pyautogui.moveTo(1554,54)
                pyautogui.leftClick()
                pyautogui.moveTo(345,56)
                pyautogui.leftClick()
                pyperclip.copy(receive)
                pyautogui.hotkey('ctrl','v')
                pyautogui.typewrite(f"{date_obj.strftime('%d/%m/%y')} | {out_sect}")
                pyautogui.moveTo(701,483)
                pyautogui.leftClick()
                pyautogui.press('enter')
            else: break
        docref166bkk()
    
    def docref49tradeinbkk(): # 49 Trade in \ Bangkok
        for i in range(1,ID49Tradein_BKK.max_row+1):
            date = ID49Tradein_BKK.cell(row=i,column=1).value
            id = ID49Tradein_BKK.cell(row=i,column=2).value
            branch = ID49Tradein_BKK.cell(row=i,column=3).value
            zone = ID49Tradein_BKK.cell(row=i,column=4).value
            if date:
                pyautogui.sleep(1)
                pyautogui.moveTo(278,87)
                pyautogui.sleep(1)
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                pyautogui.moveTo(1554,54)
                pyautogui.leftClick()
                pyautogui.moveTo(345,56)
                pyautogui.leftClick()
                pyperclip.copy(receive)
                pyautogui.hotkey('ctrl','v')
                pyautogui.typewrite(f"{date} | ")
                pyperclip.copy(str(zone))
                pyautogui.hotkey('ctrl','v')
                pyautogui.moveTo(701,483)
                pyautogui.leftClick()
                pyautogui.press('enter')
            else: break
        docref49()
    docref49()


try:
    docref()
except Exception as e:
    messagebox.showerror('Python Error',f'{e}')

print(f'Finished! Zero Error')
