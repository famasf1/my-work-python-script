
import tkinter
import openpyxl
import pyautogui as pyg
from tkinter import Tk, filedialog, messagebox
import tkinter as tk

def createlabel(text1,placex,placey,font=10): ##str text, x, y
    label = tk.Label(text=text1,font=font)
    label.place(x=placex,y=placey)

def create_button_tkinter(text1,command,placex,placey): #str text, function, x, y
    tk.Button(text=text1, command=command).place(x=placex,y=placey)

def press_enter(number):
    for n in range(0,number):
        pyg.press('enter')

root = Tk()
root.title("Recheck Phone List")
root.geometry("540x250")
root.rowconfigure(0,minsize=800, weight=1)
root.columnconfigure(1,minsize=800, weight=1)

############
def readData():
    root.excel = filedialog.askopenfilename(title="เลือกไฟล์ Excel",filetypes=(("Excel Files","*.xlsx"),("All Files","*.*")))
    sheet = openpyxl.load_workbook(root.excel, data_only=True)
    sheet_name = sheet.sheetnames
    readData.xcite_sheet = sheet[sheet_name[0]]
    readData.apple_sheet = sheet[sheet_name[1]]
    readData.remobie_sheet = sheet[sheet_name[2]]
    readData.compasia_sheet = sheet[sheet_name[3]]
    readData.yellowstar_sheet = sheet[sheet_name[4]]
    tk.Label(text=f"ไฟล์โหลดเรียบร้อยแล้วที่ {root.excel}'").place(x=25,y=75)

def xcite_start():
    root.state("iconic")
    pyg.moveTo(242,231)
    pyg.leftClick()
    pyg.leftClick()
    pyg.sleep(2)
    pyg.write(officer_entry.get())
    press_enter(2)
    pyg.write("AS0006")
    pyg.moveTo(157,787) #untick send to dealer
    pyg.leftClick()
    pyg.moveTo(228,264)
    pyg.leftClick()
    for index in range(2, readData.xcite_sheet.max_row+1):
        serial_from_itec = readData.xcite_sheet.cell(row=index, column=3).value
        serial_from_techtrade = readData.xcite_sheet.cell(row=index, column=4).value
        customer_name = readData.xcite_sheet.cell(row=index, column=5).value
        phy_id = readData.xcite_sheet.cell(row=index, column=2).value
        voucher = readData.xcite_sheet.cell(row=index, column=10).value
        docno = readData.xcite_sheet.cell(row=index, column=11).value
        voucher_code = readData.xcite_sheet.cell(row=index, column=12).value
        invoice_number = readData.xcite_sheet.cell(row=index, column=13).value
        if serial_from_itec:
            pyg.write(str(serial_from_itec))
            pyg.press("right",presses=3)
            pyg.sleep(1)
            if pyg.locateOnScreen('ret_error.png', grayscale=True):
                break
            if serial_from_techtrade != None:
                pyg.write(f"{serial_from_techtrade} |  ")
            pyg.write(f"{phy_id}  |  {customer_name}  |  {voucher}  |  {docno}  |  {voucher_code}  |  {invoice_number}")
            press_enter(1)

    messagebox.showinfo("Complete!","ลงรายการครบเรียบร้อย กรุณาตรวจสอบ")
        
def remobie_start():
    root.state("iconic")
    pyg.moveTo(242,231)
    pyg.leftClick()
    pyg.leftClick()
    pyg.sleep(2)
    pyg.write(officer_entry.get())
    press_enter(2)
    pyg.write("RM0004")
    pyg.moveTo(157,787) #untick send to dealer
    pyg.leftClick()
    pyg.moveTo(228,264)
    pyg.leftClick()
    for index in range(1, readData.remobie_sheet.max_row+1):
        serial_from_itec = readData.remobie_sheet.cell(row=index, column=3).value
        serial_from_techtrade = readData.remobie_sheet.cell(row=index, column=4).value
        customer_name = readData.remobie_sheet.cell(row=index, column=5).value
        phy_id = readData.remobie_sheet.cell(row=index, column=2).value
        voucher = readData.remobie_sheet.cell(row=index, column=10).value
        docno = readData.remobie_sheet.cell(row=index, column=11).value
        voucher_code = readData.remobie_sheet.cell(row=index, column=12).value
        invoice_number = readData.remobie_sheet.cell(row=index, column=13).value
        if serial_from_itec:
            pyg.write(str(serial_from_itec))
            pyg.press("right",presses=3)
            pyg.sleep(1)
            if pyg.locateOnScreen('ret_error.png', grayscale=True):
                break
            if serial_from_techtrade != None:
                pyg.write(f"{serial_from_techtrade} |  ")
            pyg.write(f"{phy_id}  |  {customer_name}  |  {voucher}  |  {docno}  |  {voucher_code}  |  {invoice_number}")
            press_enter(1)

    messagebox.showinfo("Complete!","ลงรายการครบเรียบร้อย กรุณาตรวจสอบ")

def apple_bkk_start():
    root.state("iconic")
    pyg.moveTo(242,231)
    pyg.leftClick()
    pyg.leftClick()
    pyg.sleep(2)
    pyg.write(officer_entry.get())
    press_enter(2)
    pyg.write("AG0011")
    pyg.moveTo(157,787) #untick send to dealer
    pyg.leftClick()
    pyg.moveTo(228,264)
    pyg.leftClick()
    for index in range(1, readData.apple_sheet.max_row+1):
        serial_from_itec = readData.apple_sheet.cell(row=index, column=3).value
        serial_from_techtrade = readData.apple_sheet.cell(row=index, column=4).value
        customer_name = readData.apple_sheet.cell(row=index, column=5).value
        phy_id = readData.apple_sheet.cell(row=index, column=2).value
        voucher = readData.apple_sheet.cell(row=index, column=10).value
        docno = readData.apple_sheet.cell(row=index, column=11).value
        voucher_code = readData.apple_sheet.cell(row=index, column=12).value
        invoice_number = readData.apple_sheet.cell(row=index, column=13).value
        if serial_from_itec:
            pyg.write(str(serial_from_itec))
            pyg.press("right",presses=3)
            pyg.sleep(1)
            if pyg.locateOnScreen('ret_error.png', grayscale=True):
                break
            if serial_from_techtrade != None:
                pyg.write(f"{serial_from_techtrade} |  ")
            pyg.write(f"{phy_id}  |  {customer_name}  |  {voucher}  |  {docno}  |  {voucher_code}  |  {invoice_number}")
            press_enter(1)
    messagebox.showinfo("Complete!","ลงรายการครบเรียบร้อย กรุณาตรวจสอบ")

def compasia_start():
    root.state("iconic")
    pyg.moveTo(242,231)
    pyg.leftClick()
    pyg.leftClick()
    pyg.sleep(2)
    pyg.write(officer_entry.get())
    press_enter(2)
    pyg.write("220076")
    pyg.moveTo(157,787) #untick send to dealer
    pyg.leftClick()
    pyg.moveTo(228,264)
    pyg.leftClick()
    for index in range(1, readData.compasia_sheet.max_row+1):
        serial_from_itec = readData.compasia_sheet.cell(row=index, column=3).value
        serial_from_techtrade = readData.compasia_sheet.cell(row=index, column=4).value
        customer_name = readData.compasia_sheet.cell(row=index, column=5).value
        phy_id = readData.compasia_sheet.cell(row=index, column=2).value
        voucher = readData.compasia_sheet.cell(row=index, column=10).value
        docno = readData.compasia_sheet.cell(row=index, column=11).value
        voucher_code = readData.compasia_sheet.cell(row=index, column=12).value
        invoice_number = readData.compasia_sheet.cell(row=index, column=13).value
        if serial_from_itec:
            pyg.write(str(serial_from_itec))
            pyg.press("right",presses=3)
            pyg.sleep(1)
            if pyg.locateOnScreen('ret_error.png', grayscale=True):
                break
            if serial_from_techtrade != None:
                pyg.write(f"{serial_from_techtrade} |  ")
            pyg.write(f"{phy_id}  |  {customer_name}  |  {voucher}  |  {docno}  |  {voucher_code}  |  {invoice_number}")
            press_enter(1)
    messagebox.showinfo("Complete!","ลงรายการครบเรียบร้อย กรุณาตรวจสอบ")

def yellowstar_start():
    root.state("iconic")
    pyg.moveTo(242,231)
    pyg.leftClick()
    pyg.leftClick()
    pyg.sleep(2)
    pyg.write(officer_entry.get())
    press_enter(2)
    pyg.write("YS0004")
    pyg.moveTo(157,787) #untick send to dealer
    pyg.leftClick()
    pyg.moveTo(228,264)
    pyg.leftClick()
    for index in range(1, readData.yellowstar_sheet.max_row+1):
        serial_from_itec = readData.yellowstar_sheet.cell(row=index, column=3).value
        serial_from_techtrade = readData.yellowstar_sheet.cell(row=index, column=4).value
        customer_name = readData.yellowstar_sheet.cell(row=index, column=5).value
        phy_id = readData.yellowstar_sheet.cell(row=index, column=2).value
        voucher = readData.yellowstar_sheet.cell(row=index, column=10).value
        docno = readData.yellowstar_sheet.cell(row=index, column=11).value
        voucher_code = readData.yellowstar_sheet.cell(row=index, column=12).value
        invoice_number = readData.yellowstar_sheet.cell(row=index, column=13).value 
        if serial_from_itec:
            pyg.write(str(serial_from_itec))
            pyg.press("right",presses=3)
            pyg.sleep(1)
            if pyg.locateOnScreen('ret_error.png', grayscale=True):
                break
            if serial_from_techtrade != None:
                pyg.write(f"{serial_from_techtrade} |  ")
            pyg.write(f"{phy_id}  |  {customer_name}  |  {voucher}  |  {docno}  |  {voucher_code}  |  {invoice_number}")
            press_enter(1)
    messagebox.showinfo("Complete!","ลงรายการครบเรียบร้อย กรุณาตรวจสอบ")    

hello = tk.Label(text="Hello!").pack()
officer = createlabel("เลขพนักงาน", 200, 20,20)
officer_entry = tk.Entry(master=root)
officer_entry.place(x=200,y=40)
greeting = create_button_tkinter("เลือกไฟล์ Excel",readData,50,20)
xcite = create_button_tkinter("Xcite", xcite_start, 70,180)
apple_bkk = create_button_tkinter("Apple Bangkok",apple_bkk_start,120,180)
remobie = create_button_tkinter("Remobie",remobie_start,230,180)
compasia = create_button_tkinter("Comp Asia",compasia_start, 310,180)
yellowstar = create_button_tkinter("Yellow Star",yellowstar_start,400,180)
selectsup = tk.Label(text="เลือก Supplier ที่ต้องการใส่ข้อมูล",font=("TH Sarabun",30)).place(relx=.5,rely=.5,anchor="center")

if __name__ == "__main__":
    root.mainloop()




#tick out send to dealer
#load from excel serial row, then phyid row


## if serial is wrong. try another one.
## if phyid is wrong, stop completely.

## else nothing gone wrong. stop.