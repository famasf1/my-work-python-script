import sys

from tkinter import filedialog, messagebox
import pyautogui as pyg
import openpyxl
import pyperclip
import tkinter as tk
from line_notify_me.line_notify_sourcecode import notifyme

## image

isthisnull = pyg.locateOnScreen(r"D:\Workstuff\my-work-python-script\Return\asset\NOT_NULL.png")
nothingleft = pyg.locateOnScreen(r"D:\Workstuff\my-work-python-script\Return\asset\nothing_error.png")


### create label above entry
def createlabel(text1,placex,placey):
    label = tk.Label(text=text1)
    label.place(x=placex,y=placey)

### Create Windows Interface for automated program without
### Changing Code constantly
root = tk.Tk()
root.title("Return Automation")
root.geometry("580x270")

root.rowconfigure(0,minsize=800, weight=1)
root.columnconfigure(1,minsize=800, weight=1)

hello = tk.Label(text="Hello!").pack()
get_supcode = createlabel("Supplier Code",25,70)
sup_code = tk.Entry(master=root)
sup_code.place(x=25,y=90)
get_supname = createlabel("Supplier Name",160,70)
sup_name = tk.Entry(master=root)
sup_name.place(x=160,y=90)
get_doc_Date = createlabel("Doc Date",290,70)
doc_Date = tk.Entry(master=root)
doc_Date.place(x=290,y=90)
get_rts = createlabel("เลข Com 7 RTS",160,120)
getRTS = tk.Entry(master=root)
getRTS.place(x=160,y=140)
get_numberofrow = createlabel("เลขแถว Excel ที่ต้องการให้เริ่ม",420,70)
getnumRow = tk.Entry(master=root)
getnumRow.place(x=420,y=90)

def clear_all_entry():
    sup_code.delete(0,'end')
    sup_name.delete(0,'end')
    doc_Date.delete(0,'end')
    getRTS.delete(0,'end')

### shorten autopress enter function
def press_enter(number):
    for n in range(0,number):
        pyg.press('enter')
### this is where the code launch

def readexcelAgain():
    openpyxl.load_workbook(readData.directoryhere, data_only=True)

def readData():
    try:
        root.state('iconic')
        directoryhere = filedialog.askopenfilename(title="เลือกไฟล์ Excel ที่มีข้อมูล", filetypes=(("Excel","*.xlsx"),('All Files','*.*')))
        root.state('normal')
    except Exception as e:
        messagebox.showerror(title="Error!",message=f"{e}")
        root.state('normal')
    tk.Label(text=f"ไฟล์โหลดเรียบร้อยแล้วที่ {directoryhere}'").place(x=25,y=45)
    data = openpyxl.load_workbook(directoryhere, data_only=True)
    datasheet = data.sheetnames
    readData.datasheet1 = data[datasheet[0]] #VAT default
    readData.datasheet2 = data[datasheet[1]] #Return
    readData.settingsheet = data[datasheet[2]] #Setting
    readData.datasheet3 = data[datasheet[3]] #NOVAT Check
    readData.data66 = data[datasheet[4]] #Not Stock From 18
    ################## Read Excel ##########
    readData.supcode = sup_code.get()
    readData.supname = sup_name.get()
    readData.docdatedata = doc_Date.get()
    readData.com7rts = getRTS.get()
    readData.getnumRow = int(getnumRow.get())

date_sent = tk.Button(master=root, command=readexcelAgain, text="Reinitialized Excel")
date_sent.place(x=25,y=140)
root.state('normal')


## VatBOT start grinding
def Vat_start_here():
    root.state('iconic')
    pyg.sleep(3)
### Vat bot (มี Vat)
    def VATbot_Start():
        pyg.sleep(2)
        for i in range(readData.getnumRow,readData.datasheet1.max_row+1): #skip row 1
            product_Code = readData.datasheet1.cell(row=i,column=2).value
            number = readData.datasheet1.cell(row=i,column=5).value
            billtype = readData.datasheet1.cell(row=i,column=23).value
            serial = readData.datasheet1.cell(row=i, column=4).value
            if billtype == "18":
                pyg.sleep(0.50)
                pyg.write(str(product_Code))
                pyg.press('Down')
                pyg.press('Down')
                pyg.sleep(1)
                if pyg.locateOnScreen(r'D:\Workstuff\my-work-python-script\Return\asset\nothing_error.png', confidence=.9): 
                    pyg.press('Esc')
                    press_enter(1)
                    pyg.press('Down')
                    continue
                pyg.moveTo(360,455)
                pyg.leftClick()
                press_enter(1)
                pyg.sleep(1)
                if pyg.locateOnScreen(r'D:\Workstuff\my-work-python-script\Return\asset\ret_error.png', grayscale=True):
                    print('error')
                    press_enter(1)
                    pyg.press('Up')
                    pyg.press('Down')
                    pyg.press('Down')
                elif pyg.locateCenterOnScreen(r'D:\Workstuff\my-work-python-script\Return\asset\ret_error2.png', grayscale=True):
                    pyg.press('Esc')
                    pyg.press('Esc')
                    pyg.press('Esc')
                    pyg.press('Down')
                    pyg.press('Down')
        root.state('normal')   
    pyg.write('22608')
    press_enter(2)
    pyg.write(readData.supcode)
    press_enter(2)
    pyg.write(readData.com7rts)
    press_enter(1)
    pyg.press('Down')
    pyg.moveTo(166,139)
    pyg.leftClick()
    press_enter(1)
    pyg.sleep(0.5)
    pyperclip.copy(readData.supname)
    pyg.sleep(0.5)
    pyg.write(f"{pyg.hotkey('ctrl','v')} | Doc Date : {readData.docdatedata}".replace("None",""))
    pyg.moveTo(124,233)
    pyg.leftClick()
    VATbot_Start()
    pyg.press('Up')
    pyg.hotkey('ctrl','a')
    pyg.hotkey('ctrl','c')
    notifyme('ตัดยอดสำเร็จ พร้อมใส่จำนวน')
    root.state('normal')


### No VAT bot (ไม่มีภาษี VAT)
def NOVAT_start_here():
    root.state('iconic')
    pyg.sleep(3)
### Vat bot (มี Vat)
    def bot_Start():
        pyg.sleep(2)
        for i in range(readData.getnumRow,readData.datasheet1.max_row+1): #skip row 1
            product_Code = readData.datasheet1.cell(row=i,column=2).value
            number = readData.datasheet1.cell(row=i,column=5).value
            billtype = readData.datasheet1.cell(row=i,column=23).value
            serial = readData.datasheet1.cell(row=i, column=4).value
            if billtype == "18":
                pyg.sleep(0.50)
                pyg.write(str(product_Code))
                pyg.press('Down')
                pyg.press('Down')
                pyg.sleep(1)
                if pyg.locateOnScreen(r'D:\Workstuff\my-work-python-script\Return\asset\nothing_error.png', confidence=.9): 
                    pyg.press('Esc')
                    press_enter(1)
                    pyg.press('Down')
                    continue
                pyg.moveTo(360,455)
                pyg.leftClick()
                press_enter(1)
                pyg.sleep(1)
                if pyg.locateOnScreen(r'D:\Workstuff\my-work-python-script\Return\asset\ret_error.png', grayscale=True):
                    print('error')
                    press_enter(1)
                    pyg.press('Up')
                    pyg.press('Down')
                    pyg.press('Down')
                elif pyg.locateCenterOnScreen(r'D:\Workstuff\my-work-python-script\Return\asset\ret_error2.png', grayscale=True):
                    pyg.press('Esc')
                    pyg.press('Esc')
                    pyg.press('Esc')
                    pyg.press('Down')
                    pyg.press('Down')
        root.state('normal')   
    pyg.write('22608')
    press_enter(2)
    pyg.write(readData.supcode)
    press_enter(2)
    pyg.write(readData.com7rts)
    press_enter(2)
    pyg.press('Down')
    pyg.moveTo(166,139)
    pyg.leftClick()
    press_enter(1)
    pyg.sleep(0.5)
    pyperclip.copy(readData.supname)
    pyg.sleep(0.5)
    pyg.write(f"{pyg.hotkey('ctrl','v')} | Doc Date : {readData.docdatedata}")
    pyg.moveTo(124,233)
    pyg.leftClick()
    bot_Start()
    pyg.press('Up')
    pyg.hotkey('ctrl','a')
    pyg.hotkey('ctrl','c')
    notifyme('ตัดยอดสำเร็จ พร้อมใส่จำนวน')
    root.state('normal')


## Put in number of items
def number_Input():
    root.state('iconic')
    pyg.sleep(3)
    pyg.press('Up')
    pyg.press('Right')
    pyg.hotkey('ctrl', 'Up')
    pyg.sleep(2)
    for i in range(readData.getnumRow, readData.datasheet2.max_row+1):
        numberitem = readData.datasheet2.cell(row=i, column=31).value #column 31
        if numberitem:
            pyg.typewrite(str(numberitem))
            press_enter(1)
            pyg.sleep(0.56)
        else:
            break
    notifyme('ใส่จำนวนสำเร็จแล้ว')
    root.state('normal')

# Stock out to 73
def stock_to73():
    ### Function to press down until you can't
    def press_down_again(times):
        pyg.press('Down',presses=times)

    root.state('iconic')
    pyg.sleep(3)
    pyg.write('22608')
    press_enter(2)
    pyg.write('73')
    press_enter(2)
    pyg.write(readData.com7rts)
    press_enter(2)
    pyperclip.copy(readData.supname)
    pyg.write(f"{pyg.hotkey('ctrl','v')} | {readData.supcode} | Doc Date : {readData.docdatedata} ")
    pyg.moveTo(231,216)
    pyg.leftClick()
    press_Again = 1
    number_Item_sofar = 1

    def itemalreadytakenException(presses):
        press_enter(1)
        pyg.press('Down')
        pyg.sleep(1.3)
        press_down_again(presses)
        press_enter(1)
        pyg.sleep(1)

    for i in range(readData.getnumRow, readData.data66.max_row+1):
        product_Code = readData.data66.cell(row=i,column=1).value
        product_Name = readData.data66.cell(row=i,column=2).value
        column3toint = readData.data66.cell(row=i, column=3).value
        number_Item = int(column3toint)
        #serial_Item = readData.data66.cell(row=i, column=4).value
        ### if productcode is found
        if product_Code:
            if number_Item == 1:
                pyg.write(str(product_Code))
                pyg.press('Right')
                press_enter(1)
                pyg.sleep(0.5)
                print(number_Item)
                continue
            else:
                print(f'Start {product_Name} {number_Item_sofar}/{number_Item}' )
                pyg.sleep(0.5)
                while number_Item_sofar <= number_Item: ## while number of total item and number of item so far is not 0, press time start at 1
                    try: #write product code, press right and then enter
                        pyg.write(str(product_Code))
                        pyg.press('Right')
                        press_enter(1)
                        pyg.sleep(1.2)
                        if pyg.locateOnScreen(r"D:\Workstuff\my-work-python-script\Return\asset\NOT_NULL.png", confidence=.7, grayscale=True): #If image input value found and this is not null, add number of items by 1 then continues
                            print('Image Found!')
                            press_enter(1)
                            pyg.sleep(1.3)
                            if pyg.locateCenterOnScreen(r"D:\Workstuff\my-work-python-script\Return\asset\ret_error.png", grayscale=True, confidence=.9): #mean item already taken
                                print('There is nothing left!')
                                itemalreadytakenException(press_Again)
                                press_Again += 1
                                number_Item_sofar += 1
                                print(f"Select another list completed. Currently i have to press down {press_Again} times")
                                print(f'Continues {number_Item_sofar}/{number_Item}' )
                                if number_Item_sofar > number_Item: #if number of items so far is more than total number, reset.
                                    print('Resetting back to 1')
                                    number_Item_sofar = 1
                                    press_Again = 1
                                    break
                            else:
                                number_Item_sofar += 1
                                print('Enter | Pass')
                                print(f'Continues {number_Item_sofar}/{number_Item}' )
                        else:
                            number_Item_sofar += 1
                            print(f'Continues {number_Item_sofar}/{number_Item}' )
                            print('Operation Completed! Continues...')
                            if number_Item_sofar > number_Item: #if number of items so far is more than total number, reset.
                                print('Resetting back to 1')
                                number_Item_sofar = 1
                                press_Again = 1
                                break
                            else:
                                continue
                        
        
                    except Exception:
                        pass
                        #pyg.write(str(number_Item))
                        #pyg.press('Left')
                        #pyg.press('Left')
                        #pyg.press('Left')  
                else:
                    continue
    notifyme('ตัดยอด 73 เสร็จสิ้น')
    root.state('normal')

def stock_to73_Noserial():
    root.state('iconic')
    pyg.sleep(3)
    pyg.write('22608')
    press_enter(2)
    pyg.write('73')
    press_enter(2)
    pyg.write(readData.com7rts)
    press_enter(2)
    pyperclip.copy(readData.supname)
    pyg.write(f"{pyg.hotkey('ctrl','v')} | {readData.supcode} | Doc Date : {readData.docdatedata} ")
    pyg.moveTo(231,216)
    pyg.leftClick()
    for i in range(readData.getnumRow, readData.data66.max_row+1):
        product_Code = readData.data66.cell(row=i,column=1).value
        number_Item = readData.data66.cell(row=i, column=3).value
        if product_Code:
            pyg.write(str(product_Code))
            pyg.press('Right')
            pyg.press('Right')
            pyg.write(str(number_Item))
            pyg.press('Left')
            pyg.press('Left')
            pyg.press('Left')            
            press_enter(1)
    root.state('normal')

def create_button_tkinter(text1,command,placex,placey):
    tk.Button(text=text1, command=command).place(x=placex,y=placey)

def restart_Bot():
    pyg.sleep(2)
    for i in range(readData.getnumRow,readData.datasheet1.max_row+1): #skip row 1
        product_Code = readData.datasheet1.cell(row=i,column=2).value
        number = readData.datasheet1.cell(row=i,column=5).value
        billtype = readData.datasheet1.cell(row=i,column=23).value
        serial = readData.datasheet1.cell(row=i, column=4).value
        if billtype == "18":
            pyg.sleep(0.50)
            pyg.write(str(product_Code))
            pyg.press('Down')
            pyg.press('Down')
            pyg.sleep(1)
            if pyg.locateOnScreen(r'D:\Workstuff\my-work-python-script\Return\asset\nothing_error.png', confidence=.9): 
                pyg.press('Esc')
                press_enter(1)
                pyg.press('Down')
                continue
            pyg.moveTo(360,455)
            pyg.leftClick()
            press_enter(1)
            pyg.sleep(1)
            if pyg.locateOnScreen(r'D:\Workstuff\my-work-python-script\Return\asset\ret_error.png', grayscale=True):
                print('error')
                press_enter(1)
                pyg.press('Up')
                pyg.press('Down')
                pyg.press('Down')
            elif pyg.locateCenterOnScreen(r'D:\Workstuff\my-work-python-script\Return\asset\ret_error2.png', grayscale=True):
                pyg.press('Esc')
                pyg.press('Esc')
                pyg.press('Esc')
                pyg.press('Down')
                pyg.press('Down')
    root.state('normal')   


greeting = create_button_tkinter("Browse",readData,250,20)
vatstart = create_button_tkinter("เริ่ม (มี VAT)",Vat_start_here,25,190)
novat_start = create_button_tkinter("เริ่ม (ไม่มี VAT)",NOVAT_start_here,110,190)
number_times = create_button_tkinter("ใส่จำนวนที่คืน",number_Input,205,190)
stock_out_73 = create_button_tkinter("โอนบิล 66 ไป 73",stock_to73,300,190)
stock_out_73_noserialize = create_button_tkinter("โอนบิล 66 ไป 73 (ไม่มี Serial)",stock_to73_Noserial,300,150)
clear_all = create_button_tkinter("เคลียร์ช่อง",clear_all_entry,405,190)
renew = create_button_tkinter("Custom Row",restart_Bot,475,190)

if __name__ == "__main__":
    root.mainloop()
    
        