from tkinter import filedialog, messagebox
import pyautogui as pyg
import openpyxl
import pyperclip
import tkinter as tk
import os

not_null_asset_location = r"asset\NOT_NULL.png"
nothing_error = r"asset\nothing_error.png"

isthisnull = pyg.locateOnScreen(os.path.join(os.getcwd(),not_null_asset_location))
nothingleft = pyg.locateOnScreen(os.path.join(os.getcwd(), nothing_error))


import requests
### Notify me when the script is completed to LINE.
def notifyme(confirmtext):
    """
    LINE Notify - Send text to my own line.
    parameter :
    confirmtext: str (required)
    """
    
    mytoken = 'kOcQyjPGgIAgTQ4qWjTlEJZFUj7GegzGefdDEiSsYJr'
    url = 'https://notify-api.line.me/api/notify'
    data = {
        'message' : confirmtext
    }
    options = {
        'Method' : 'POST',
        'Content-Type' : 'application/x-www-form-urlencoded',
        'Authorization' : f'Bearer {mytoken}',
    }
    response = requests.post(url=url, headers=options, data=data)
    print(response.status_code)


### create label above entry
def createlabel(text1,placex,placey):
    label = tk.Label(text=text1)
    label.place(x=placex,y=placey)

def create_button_tkinter(text1,command,placex,placey):
    tk.Button(text=text1, command=command).place(x=placex,y=placey)

### Create Windows Interface for automated program without
### Changing Code constantly
root = tk.Tk()
root.title("Return Automation")
root.geometry("580x270")

hello = tk.Label(text="Hello!").pack()
get_branch_to = createlabel("Branch To",25,70)
branch_to = tk.Entry(master=root)
branch_to.place(x=25,y=90)
get_comment = createlabel("Comment",160,70)
comment = tk.Entry(master=root)
comment.place(x=160,y=90)
get_staffid = createlabel("เลข ID พนักงาน",295,70)
staffid = tk.Entry(master=root)
staffid.place(x=295,y=90)
get_numberofrow = createlabel("เลขแถว Excel ที่ต้องการให้เริ่ม",430,70)
getnumRow = tk.Entry(master=root)
getnumRow.place(x=430,y=90)

def clear_all_entry():
    branch_to.delete(0,'end')
    comment.delete(0,'end')

### shorten autopress enter function
def press_enter(number):
    for n in range(0,number):
        pyg.press('enter')

def readData():
    try:
        root.state('iconic')
        directoryhere = filedialog.askopenfilename(title="เลือกไฟล์ Excel ที่มีข้อมูล", filetypes=(("Excel","*.xlsx"),('All Files','*.*')))
        root.state('normal')
    except Exception as e:
        messagebox.showerror(title="Error!",message=f"{e}")
        root.state('normal')
    tk.Label(text=f"ไฟล์โหลดเรียบร้อยแล้วที่ {directoryhere}'").place(x=25,y=45)
    data = openpyxl.load_workbook(directoryhere, data_only=True)
    datasheet = data.sheetnames
    readData.datasheet1 = data[datasheet[0]] #VAT default
    ################## Read Excel ##########

    readData.branch_to = branch_to.get()
    readData.comment = comment.get()
    readData.staffid = staffid.get()
    readData.getnumRow = int(getnumRow.get())


# Stock out
def stock_out():
    ### Function to press down until you can't
    def press_down_again(times):
        pyg.press('Down',presses=times)

    root.state('iconic')
    pyg.sleep(3)
    pyg.write(readData.staffid)
    press_enter(2)
    pyg.write(readData.branch_to)
    press_enter(4)
    comment = pyperclip.copy(readData.comment)
    pyperclip.paste()
    pyg.moveTo(231,216)
    pyg.leftClick()
    press_Again = 1
    number_Item_sofar = 1

    def itemalreadytakenException(presses):
        press_enter(1)
        pyg.press('Down')
        pyg.sleep(1.3)
        press_down_again(presses)
        press_enter(1)
        pyg.sleep(1)

    for i in range(readData.getnumRow, readData.datasheet1.max_row+1):
        product_Code = readData.datasheet1.cell(row=i,column=1).value
        product_Name = readData.datasheet1.cell(row=i,column=2).value
        column3toint = readData.datasheet1.cell(row=i, column=3).value
        number_Item = int(column3toint)
        #serial_Item = readData.datasheet1.cell(row=i, column=4).value
        ### if productcode is found
        if product_Code:
            if number_Item == 1:
                pyg.write(str(product_Code))
                pyg.press('Right')
                press_enter(1)
                pyg.sleep(0.5)
                print(number_Item)
                continue
            else:
                print(f'Start {product_Name} {number_Item_sofar}/{number_Item}' )
                pyg.sleep(0.5)
                while number_Item_sofar <= number_Item: ## while number of total item and number of item so far is not 0, press time start at 1
                    try: #write product code, press right and then enter
                        pyg.write(str(product_Code))
                        pyg.press('Right')
                        press_enter(1)
                        pyg.sleep(1.2)
                        if pyg.locateOnScreen(fr"{os.path.join(os.getcwd(),not_null_asset_location)}", confidence=.7, grayscale=True): #If image input value found and this is not null, add number of items by 1 then continues
                            print('Image Found!')
                            press_enter(1)
                            pyg.sleep(1.3)
                            if pyg.locateCenterOnScreen(fr"{os.path.join(os.getcwd(), nothing_error)}", grayscale=True, confidence=.9): #mean item already taken
                                print('There is nothing left!')
                                itemalreadytakenException(press_Again)
                                press_Again += 1
                                number_Item_sofar += 1
                                print(f"Select another list completed. Currently i have to press down {press_Again} times")
                                print(f'Continues {number_Item_sofar}/{number_Item}' )
                                if number_Item_sofar > number_Item: #if number of items so far is more than total number, reset.
                                    print('Resetting back to 1')
                                    number_Item_sofar = 1
                                    press_Again = 1
                                    break
                            else:
                                number_Item_sofar += 1
                                print('Enter | Pass')
                                print(f'Continues {number_Item_sofar}/{number_Item}' )
                        else:
                            number_Item_sofar += 1
                            print(f'Continues {number_Item_sofar}/{number_Item}' )
                            print('Operation Completed! Continues...')
                            if number_Item_sofar > number_Item: #if number of items so far is more than total number, reset.
                                print('Resetting back to 1')
                                number_Item_sofar = 1
                                press_Again = 1
                                break
                            else:
                                continue
                        
        
                    except Exception:
                        pass
                        #pyg.write(str(number_Item))
                        #pyg.press('Left')
                        #pyg.press('Left')
                        #pyg.press('Left')  
                else:
                    continue
    notifyme(f'ตัดยอด {readData.branch_to} เสร็จสิ้น')
    root.state('normal')

def restart_out():
    press_Again = 1
    number_Item_sofar = 1
    pyg.sleep(10)

    def press_down_again(times):
        pyg.press('Down',presses=times)

    def itemalreadytakenException(presses):
        press_enter(1)
        pyg.press('Down')
        pyg.sleep(1.3)
        press_down_again(presses)
        press_enter(1)
        pyg.sleep(1)

    for i in range(readData.getnumRow, readData.datasheet1.max_row+1):
        product_Code = readData.datasheet1.cell(row=i,column=1).value
        product_Name = readData.datasheet1.cell(row=i,column=2).value
        column3toint = readData.datasheet1.cell(row=i, column=3).value
        number_Item = int(column3toint)
        #serial_Item = readData.datasheet1.cell(row=i, column=4).value
        ### if productcode is found
        if product_Code:
            if number_Item == 1:
                pyg.write(str(product_Code))
                pyg.press('Right')
                press_enter(1)
                pyg.sleep(0.5)
                print(number_Item)
                continue
            else:
                print(f'Start {product_Name} {number_Item_sofar}/{number_Item}' )
                pyg.sleep(0.5)
                while number_Item_sofar <= number_Item: ## while number of total item and number of item so far is not 0, press time start at 1
                    try: #write product code, press right and then enter
                        pyg.write(str(product_Code))
                        pyg.press('Right')
                        press_enter(1)
                        pyg.sleep(1.2)
                        if pyg.locateOnScreen(fr"{os.path.join(os.getcwd(),not_null_asset_location)}", confidence=.7, grayscale=True): #If image input value found and this is not null, add number of items by 1 then continues
                            print('Image Found!')
                            press_enter(1)
                            pyg.sleep(1.3)
                            if pyg.locateCenterOnScreen(fr"{os.path.join(os.getcwd(), nothing_error)}", grayscale=True, confidence=.9): #mean item already taken
                                print('There is nothing left!')
                                itemalreadytakenException(press_Again)
                                press_Again += 1
                                number_Item_sofar += 1
                                print(f"Select another list completed. Currently i have to press down {press_Again} times")
                                print(f'Continues {number_Item_sofar}/{number_Item}' )
                                if number_Item_sofar > number_Item: #if number of items so far is more than total number, reset.
                                    print('Resetting back to 1')
                                    number_Item_sofar = 1
                                    press_Again = 1
                                    break
                            else:
                                number_Item_sofar += 1
                                print('Enter | Pass')
                                print(f'Continues {number_Item_sofar}/{number_Item}' )
                        else:
                            number_Item_sofar += 1
                            print(f'Continues {number_Item_sofar}/{number_Item}' )
                            print('Operation Completed! Continues...')
                            if number_Item_sofar > number_Item: #if number of items so far is more than total number, reset.
                                print('Resetting back to 1')
                                number_Item_sofar = 1
                                press_Again = 1
                                break
                            else:
                                continue
                        
        
                    except Exception:
                        pass
                        #pyg.write(str(number_Item))
                        #pyg.press('Left')
                        #pyg.press('Left')
                        #pyg.press('Left')  
                else:
                    continue
    notifyme('ตัดยอด 390 เสร็จสิ้น')
    root.state('normal')



greeting = create_button_tkinter("Browse",readData,250,20)
stock_out = create_button_tkinter("โอนบิล",stock_out,250,190)
restart = create_button_tkinter("เริ่มจากแถวเดิม", restart_out, 300,190)

if __name__ == "__main__":
    root.mainloop()