from tkinter import filedialog, messagebox
from tkinter import *
import pyautogui as pyg
import openpyxl
import pyperclip
import random

try:
    root = Tk()
    root.excel = filedialog.askopenfilename(initialdir='/Desktop',title='เลือกไฟล์ Excel สำหรับ Stock-Out33', filetypes=(('Excel','*.xlsx'),('All Files','*.*')))
    workbook = openpyxl.load_workbook(root.excel, data_only=True)
    root.withdraw()
    sheet = workbook.sheetnames
    worksheet = workbook[sheet[0]]
    numb = [0,1,2]
    receive = ['ครบ / ดิว','ครบ / ก็อต','ครบ / เอก']
    idnumber = ['22073','23017','23267']
    readpicerrorfound = pyg.locateCenterOnScreen('checkerror.png')
except Exception as e:
    messagebox.showerror('Python Error', f'{e}')
    exit()

def pressenter(numberoftimes):
    for i in range(0, numberoftimes):
        pyg.press('enter')
        if i == numberoftimes:
            break

def clickleft(numberoftimes):
    for i in range(0, numberoftimes):
        pyg.leftClick()
        pyg.sleep(0.5)
        if i == numberoftimes:
            break


def defaultbeh():

    def firstStart(start):
        if start == 1:
            pyg.moveTo(238,145)
            pyg.leftClick()
            pyg.sleep(3.5)
        else:
            pass
    
    def nextStart(next):
        if next == 1:
            pass
        else:
            pyg.moveTo(238,145)
            pyg.leftClick()
            pyg.sleep(3.5)

    for i in range(1, worksheet.max_row+1):
        stockoutid = worksheet.cell(row=i, column=1).value
        etc = worksheet.cell(row=i, column=2).value
        if stockoutid:
            nextStart(i)
            firstStart(i)
            pyg.moveTo(184,65)
            clickleft(2)
            print('Start Stock In')
            if etc == 'm': 
                pyg.typewrite('7538')
            elif etc == 'โบ้':
                pyg.typewrite('22608')
            elif etc == 'pairin':
                pyg.typewrite('1815')
            elif etc == 'ปาน':
                pyg.typewrite('22929')
            else: 
                r = random.choice(numb)
                pyg.typewrite(idnumber[r])
            pressenter(2)
            pyg.typewrite(str(stockoutid)) #stockout
            pressenter(2)
            if etc == 'm': # for my queen
                pyperclip.copy('ครบ / เอ็ม')
                pyg.hotkey('ctrl', 'v')
            elif etc == 'โบ้':
                pyperclip.copy('ครบ / โบ้')
                pyg.hotkey('ctrl', 'v')
            elif etc == 'pairin':
                pyperclip.copy('ครบ / ไพรินทร์')
                pyg.hotkey('ctrl', 'v')
            elif etc == 'ปาน':
                pyperclip.copy('ครบ / ปาน')
                pyg.hotkey('ctrl', 'v')           
            else:
                pyperclip.copy(receive[r])
                pyg.hotkey('ctrl', 'v')
            pyg.moveTo(66,1023)
            clickleft(1) # this is where stock in started
            print('Press Ok')
            try:
                if readpicerrorfound:
                    pressenter(1)
                    print('Found Error')
                    continue
                else:
                    print('Error is not Found')
                    pressenter(1)
                    pyg.press('left')
                    pressenter(4)
                    pyg.sleep(2)
                    continue
            except Exception as e:
                messagebox.showerror('Python Error', f'{e}')
        else: break


try:
    defaultbeh()
except Exception as e:
    messagebox.showerror('Python Error', f'{e}')
    exit()








































