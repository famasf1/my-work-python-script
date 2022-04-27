from math import prod
import pyautogui as pyg
import openpyxl
import pyperclip
            
def press_enter(number):
    for n in range(0,number):
        pyg.press('enter')

directoryhere = r"C:\Users\Comseven\my-work-python-script\returndataready.xlsx"
data = openpyxl.load_workbook(directoryhere, data_only=True)
datasheet = data.sheetnames
datasheet1 = data[datasheet[0]] #VAT default
datasheet2 = data[datasheet[1]] #Return
settingsheet = data[datasheet[2]] #Setting
datasheet3 = data[datasheet[3]] #NOVAT Check
data66 = data[datasheet[4]]

supcode = settingsheet['B1'].value
supname = settingsheet['B2'].value
docdatedata = settingsheet['A3'].value
datesent = settingsheet['B3'].value
com7rts = settingsheet['B4'].value


def VATbot_Start():
    pyg.sleep(2)
    for i in range(524,datasheet1.max_row+1): #skip row 1
        product_Code = datasheet1.cell(row=i,column=2).value
        number = datasheet1.cell(row=i,column=5).value
        billtype = datasheet1.cell(row=i,column=23).value
        if billtype == 18:
            pyg.sleep(0.50)
            pyg.write(str(product_Code))
            pyg.press('Down')
            pyg.press('Down')
            pyg.moveTo(360,455)
            pyg.leftClick()
            press_enter(1)
            pyg.sleep(0.8)
            pyg.locateOnScreen('ret_error.png', grayscale=True)
            if pyg.locateOnScreen('ret_error.png', grayscale=True):
                print('error')
                press_enter(1)
                pyg.press('Up')
                pyg.press('Down')
                pyg.press('Down')
            pyg.locateCenterOnScreen('ret_error2.png', grayscale=True)
           # if nothing_Left:
           #     print('nothing')
           #     pyg.press('Esc')
           #     pyg.press('Enter')
           #     pyg.press('Down')
            if pyg.locateCenterOnScreen('ret_error2.png', grayscale=True):
                pyg.press('Esc')
                pyg.press('Esc')
                pyg.press('Esc')
                pyg.press('Down')
                pyg.press('Down')                


def bot_Start():
    pyg.sleep(2)
    for i in range(2,datasheet3.max_row+1): #skip row 1
        product_Code = datasheet3.cell(row=i,column=2).value
        number = datasheet3.cell(row=i,column=5).value
        billtype = datasheet3.cell(row=i,column=23).value
        if billtype == 18:
            pyg.sleep(0.50)
            pyg.write(str(product_Code))
            pyg.press('Down')
            pyg.press('Down')
            pyg.sleep(0.8)
            press_enter(1)
            pyg.sleep(1)
            pyg.locateOnScreen('ret_error.png', grayscale=True)
            pyg.locateCenterOnScreen('ret_error2.png', grayscale=True)
           # if nothing_Left:
           #     print('nothing')
           #     pyg.press('Esc')
           #     pyg.press('Enter')
           #     pyg.press('Down')
            if pyg.locateOnScreen('ret_error.png', grayscale=True):
                print('error')
                press_enter(1)
                pyg.press('Up')
                pyg.press('Down')
                pyg.press('Down')
            elif pyg.locateCenterOnScreen('ret_error2.png', grayscale=True):
                pyg.press('Esc')
                pyg.press('Esc')
                pyg.press('Esc')
                pyg.press('Down')
                pyg.press('Down')    

def Vat_start_here():
    pyg.write('22608')
    press_enter(2)
    pyg.write(supcode)
    press_enter(2)
    pyg.write(com7rts)
    press_enter(1)
    pyg.press('Down')
    pyg.moveTo(166,139)
    pyg.leftClick()
    press_enter(1)
    pyg.sleep(0.5)
    pyperclip.copy(supname)
    pyg.sleep(0.5)
    pyg.write(f"{pyg.hotkey('ctrl','v')} | {docdatedata} : {datesent}")
    pyg.moveTo(124,233)
    pyg.leftClick()
    VATbot_Start()
    pyg.press('Up')
    pyg.hotkey('ctrl','a')
    pyg.hotkey('ctrl','c')

def NOVAT_start_here():
    pyg.write('22608')
    press_enter(2)
    pyg.write(supcode)
    press_enter(2)
    pyg.write(com7rts)
    press_enter(1)
    pyg.moveTo(166,139)
    pyg.leftClick()
    press_enter(1)
    pyg.sleep(0.5)
    pyperclip.copy(supname)
    pyg.sleep(0.5)
    pyg.write(f"{pyg.hotkey('ctrl','v')} | {docdatedata} : {datesent}")
    pyg.moveTo(124,233)
    pyg.leftClick()
    bot_Start()
    pyg.press('Up')
    pyg.hotkey('ctrl','a')
    pyg.hotkey('ctrl','c')

def stilltest():
    pyg.press('Up')
    pyg.press('Right')
    pyg.hotkey('ctrl', 'Up')
    pyg.sleep(2)
    for i in range(1, datasheet2.max_row+1):
        numberitem = datasheet2.cell(row=i, column=31).value
        pyg.typewrite(str(numberitem))
        press_enter(1)
        pyg.sleep(0.56)

def stock_to73():
    #pyg.moveTo(63,147)
    #pyg.leftClick()
    pyg.write('22608')
    press_enter(2)
    pyg.write('73')
    press_enter(2)
    pyg.write(com7rts)
    press_enter(2)
    pyperclip.copy(supname)
    pyg.write(f"{supcode} | {pyg.hotkey('ctrl','v')}")
    pyg.moveTo(231,216)
    pyg.leftClick()
    for i in range(2, data66.max_row+1):
        product_Code = data66.cell(row=i,column=4).value
        pyg.write(str(product_Code))
        press_enter(1)

pyg.sleep(1)

#Vat_start_here()
#NOVAT_start_here()
#bot_Start()
#VATbot_Start()
#stilltest()
stock_to73()