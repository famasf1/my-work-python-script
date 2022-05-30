import pyautogui
import openpyxl
from tkinter import filedialog, messagebox
from tkinter import *
import datetime
import pyperclip
###################### Part 1 : Start the variable process
try:
    root = Tk()
    root.excel = filedialog.askopenfilename(initialdir='/Desktop',title='เลือกไฟล์ Excel สำหรับ Stock-Out33', filetypes=(('Excel','*.xlsx'),('All Files','*.*')))
    workbook = openpyxl.load_workbook(root.excel, data_only=True)
    root.withdraw()
    worksheetname = workbook.sheetnames
    ID33_Data = workbook[worksheetname[1]]
    ID33BKK_Data = workbook[worksheetname[6]]
    ID49_Data = workbook[worksheetname[3]]
    ID49BKK_Data = workbook[worksheetname[8]]
    ID747_Data = workbook[worksheetname[4]]
    delivery_Failed_Data = workbook[worksheetname[5]]
    receive = 'รับ'
except Exception as e:
    messagebox.showerror('Python Error', f'{e}')
    exit()

def getdate_Obj(dateData):
    try:
        date_obj = datetime.datetime.strptime(dateData, "%Y-%m-%d %H:%M:%S.%f")
    except:
        try:
            date_obj = datetime.datetime.strptime(dateData, "%Y-%m-%d %H:%M:%S")
        except:
            try:
                date_obj = datetime.datetime.strptime(dateData, "%d-%m-%Y %H:%M:%S")
            except:
                try:
                    date_obj = datetime.datetime.strptime(dateData, "%d/%m/%Y")
                except Exception as e:
                    print(e)
                
    return date_obj.strftime('%d/%m/%y')

################################ TEST ROOM ###############################



#########################################################################
################################ Part 2: Reading excel and Start writing
def docref():
    def defaultref(): #Do this everytime i want to start.
        pyautogui.moveTo(45,255)
        pyautogui.sleep(1)
        pyautogui.leftClick()
        pyautogui.sleep(3)
        pyautogui.moveTo(1432,192)
        pyautogui.leftClick()
        pyautogui.moveTo(1427,215)
        pyautogui.doubleClick()
        pyautogui.leftClick()
        pyautogui.press('down')
        pyautogui.press('enter')
        pyautogui.sleep(3)
        pyautogui.moveTo(278,87)
        pyautogui.doubleClick()
        pyautogui.leftClick()

    def docref747(): #49 Return DHL
        for i in range(1,ID747_Data.max_row+1):
            out_sect = ID747_Data.cell(row=i,column=15).value
            id = ID747_Data.cell(row=i,column=12).value
            branch = ID747_Data.cell(row=i,column=13).value
            date = ID747_Data.cell(row=i, column=7).value
            formulae = f"=ifna(VLOOKUP(M{i},Data!C:G,5,0),"")"
            ID747_Data.cell(row=i,column=15).value = formulae
            workbook.save('bitly+ready.xlsx')
            if out_sect:
                pyautogui.sleep(1)
                pyautogui.moveTo(278,87)
                pyautogui.sleep(1)
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                pyautogui.moveTo(1554,54)
                pyautogui.leftClick()
                pyautogui.moveTo(345,56)
                pyautogui.leftClick()
                pyperclip.copy(receive)
                pyautogui.hotkey('ctrl','v')
                pyautogui.typewrite(f"{getdate_Obj(str(date))} | {out_sect}")
                pyautogui.moveTo(701,483)
                pyautogui.leftClick()
                pyautogui.sleep(0.8)
                pyautogui.moveTo(855,505)
                pyautogui.leftClick()
            else: break

    def failed_toDeliver():
        for row in range(1, delivery_Failed_Data.max_row+1):
            reason = delivery_Failed_Data.cell(row=row, column=9).value
            phyid = delivery_Failed_Data.cell(row=row, column=17).value
            branch = delivery_Failed_Data.cell(row=row, column=18).value
            date = delivery_Failed_Data.cell(row=row, column=12).value
            if reason:
                pyautogui.sleep(1)
                pyautogui.moveTo(278,87)
                pyautogui.sleep(1)
                pyautogui.doubleClick()
                pyautogui.typewrite(str(phyid))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                pyautogui.moveTo(1554,54)
                pyautogui.leftClick()
                pyautogui.moveTo(345,56)
                pyautogui.leftClick()
                pyperclip.copy(str(reason))
                pyautogui.sleep(0.8)
                pyautogui.write(f"{pyperclip.paste()} | {getdate_Obj(str(date))}")
                pyautogui.moveTo(701,483)
                pyautogui.leftClick()
                pyautogui.sleep(0.8)
                pyautogui.moveTo(855,505)
                pyautogui.leftClick()
            else: break
        docref747()


    def docref49returnbkk(): #49 Return BKK
        for i in range(1,ID49BKK_Data.max_row+1):
            date = ID49BKK_Data.cell(row=i, column=1).value
            id = ID49BKK_Data.cell(row=i,column=2).value
            branch = ID49BKK_Data.cell(row=i,column=3).value
            zone = ID49BKK_Data.cell(row=i,column=4).value          
            if id:
                pyautogui.sleep(1)
                pyautogui.moveTo(278,87)
                pyautogui.sleep(1)
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                pyautogui.moveTo(1554,54)
                pyautogui.leftClick()
                pyautogui.moveTo(345,56)
                pyautogui.leftClick()
                pyperclip.copy(receive)
                pyautogui.hotkey('ctrl','v')
                pyautogui.typewrite(f'{str(getdate_Obj(str(date)))} | ')
                pyperclip.copy(str(zone))
                pyautogui.hotkey('ctrl','v')
                pyautogui.moveTo(701,483)
                pyautogui.leftClick()
                pyautogui.sleep(0.8)
                pyautogui.moveTo(855,505)
                pyautogui.leftClick()
            else: break
        failed_toDeliver()
        
    def docref49return(): #49 Return DHL
        for i in range(1,ID49_Data.max_row+1):
            out_sect = ID49_Data.cell(row=i,column=15).value
            id = ID49_Data.cell(row=i,column=12).value
            branch = ID49_Data.cell(row=i,column=13).value
            date = ID49_Data.cell(row=i, column=7).value
            formulae = f"=ifna(VLOOKUP(M{i},Data!C:G,5,0),"")"
            ID49_Data.cell(row=i,column=15).value = formulae
            workbook.save('bitly+ready.xlsx')
            if out_sect:
                pyautogui.sleep(1)
                pyautogui.moveTo(278,87)
                pyautogui.sleep(1)
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                pyautogui.moveTo(1554,54)
                pyautogui.leftClick()
                pyautogui.moveTo(345,56)
                pyautogui.leftClick()
                pyperclip.copy(receive)
                pyautogui.hotkey('ctrl','v')
                pyautogui.typewrite(f"{getdate_Obj(str(date))} | {out_sect}")
                pyautogui.moveTo(701,483)
                pyautogui.leftClick()
                pyautogui.sleep(0.8)
                pyautogui.moveTo(855,505)
                pyautogui.leftClick()
            else: break
        docref49returnbkk()
        
    def docref33bkk(): # 33 Service Headoffice BKK
        for i in range(1,ID33BKK_Data.max_row+1):
            date = ID33BKK_Data.cell(row=i,column=2).value
            id = ID33BKK_Data.cell(row=i,column=5).value
            branch = ID33BKK_Data.cell(row=i,column=6).value
            zone = ID33BKK_Data.cell(row=i,column=9).value
            if date:
                pyautogui.sleep(1)
                pyautogui.moveTo(278,87)
                pyautogui.sleep(1)
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                pyautogui.moveTo(1554,54)
                pyautogui.leftClick()
                pyautogui.moveTo(345,56)
                pyautogui.leftClick()
                pyperclip.copy(receive)
                pyautogui.hotkey('ctrl','v')
                pyautogui.typewrite(f'{getdate_Obj(str(date))} | ')
                pyperclip.copy(str(zone))
                pyautogui.hotkey('ctrl','v')
                pyautogui.moveTo(701,483)
                pyautogui.leftClick()
                pyautogui.sleep(0.8)
                pyautogui.moveTo(855,505)
                pyautogui.leftClick()
            else: break
        docref49return()
    
    # THIS IS WHERE EVERYTHING STARTED!
    def docref33(): # 33 Service Headoffice DHL
        defaultref()
        for i in range(1,ID33_Data.max_row+1):
            out_sect = ID33_Data.cell(row=i,column=15).value
            id = ID33_Data.cell(row=i,column=12).value
            branch = ID33_Data.cell(row=i,column=13).value
            date = ID33_Data.cell(row=i, column=7).value
            formulae = f"=ifna(VLOOKUP(M{i},Data!C:G,5,0),"")"
            ID33_Data.cell(row=i,column=15).value = formulae
            workbook.save('bitly+ready.xlsx')
            if out_sect:
                pyautogui.sleep(1)
                pyautogui.moveTo(278,87)
                pyautogui.sleep(1)
                pyautogui.doubleClick()
                pyautogui.typewrite(str(id))
                pyautogui.press('enter')
                pyautogui.typewrite(str(branch))
                pyautogui.moveTo(1554,54)
                pyautogui.leftClick()
                pyautogui.moveTo(345,56)
                pyautogui.leftClick()
                pyperclip.copy(receive)
                pyautogui.hotkey('ctrl','v')
                pyautogui.typewrite(f"{getdate_Obj(str(date))} | {out_sect}")
                pyautogui.moveTo(701,483)
                pyautogui.leftClick()
                pyautogui.sleep(0.8)
                pyautogui.moveTo(855,505)
                pyautogui.leftClick()
            else: break 
        docref33bkk()
    docref33()


############################ Finally  - Try running it. If there's error, print it out.
try:
    docref()
except Exception as e:
    messagebox.showerror('Python Error',f'{e}')

print(f'Finished! Zero Error')

