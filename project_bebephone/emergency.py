import os, glob, shutil
import openpyxl
from tkinter import Tk, filedialog
import itertools
import natsort as ns

asset_directory = os.path.join(r"D:\Workstuff\my-work-python-script\project_bebephone", r"Asset")

root = Tk()

#get asset
PRINT_dir = r"D:\Workstuff\my-work-python-script\project_bebephone\PRINT"
asset_dir_56 = os.path.join(asset_directory, r"block56template.xlsm")
block56_wb = openpyxl.load_workbook(asset_dir_56, data_only=True, keep_vba=True)
block56_ws = block56_wb.sheetnames
block56 = block56_wb[block56_ws[0]]
#get data
get_wb = filedialog.askopenfilename(title="เลือกไฟล์ข้อมูล Excel", filetypes=([("*Excel Files", "*.xlsx"), ("All Files", "*.*")]))
data_wb = openpyxl.load_workbook(get_wb, data_only=True)
data_sheet = data_wb.sheetnames
get_sheet = data_wb[data_sheet[0]]
def main():

        #bebe row
    bebe_start_row_number = 2

    #barcode row
    com7barcode_startrow = 3

    #barcode text
    com7barcodetext_startrow = 4

    #data
    data_startrow = 3

    ran = False
    col = 2
    folder_count = 1

    for total_print_row in range(3,get_sheet.max_row+1):
        bebecode = get_sheet.cell(row=total_print_row,column = 2).value #0166
        com7productcode = get_sheet.cell(row=total_print_row, column = 1).value #51691
        total_sticker = get_sheet.cell(row=total_print_row, column=3).value
        total_print_data = get_sheet.cell(row=total_print_row, column = 4).value #8 #int
        try:
            os.chdir(r"D:\Workstuff\my-work-python-script\project_bebephone")
            os.mkdir("Block 56")
            os.chdir(r"D:\Workstuff\my-work-python-script\project_bebephone\Block 56")
        except FileExistsError:
            pass
        os.chdir(r"D:\Workstuff\my-work-python-script\project_bebephone\Block 56")
        #1
        #b
        block56['B3'] = bebecode
        block56['B4'] = f'''=code128("{com7productcode}")'''
        block56['B5'] = com7productcode
        #d
        block56['D3'] = bebecode
        block56['D4'] = f'''=code128("{com7productcode}")'''
        block56['D5'] = com7productcode
        #f
        block56['F3'] = bebecode
        block56['F4'] = f'''=code128("{com7productcode}")'''
        block56['F5'] = com7productcode 
        #h
        block56['H3'] = bebecode
        block56['H4'] = f'''=code128("{com7productcode}")'''
        block56['H5'] = com7productcode 

        #b
        block56['B8'] = bebecode
        block56['B9'] = f'''=code128("{com7productcode}")'''
        block56['B10'] = com7productcode
        #d
        block56['D8'] = bebecode
        block56['D9'] = f'''=code128("{com7productcode}")'''
        block56['D10'] = com7productcode
        #f
        block56['F8'] = bebecode
        block56['F9'] = f'''=code128("{com7productcode}")'''
        block56['F10'] = com7productcode 
        #h
        block56['H8'] = bebecode
        block56['H9'] = f'''=code128("{com7productcode}")'''
        block56['H10'] = com7productcode 

        #b
        block56['B13'] = bebecode
        block56['B14'] = f'''=code128("{com7productcode}")'''
        block56['B15'] = com7productcode
        #d
        block56['D13'] = bebecode
        block56['D14'] = f'''=code128("{com7productcode}")'''
        block56['D15'] = com7productcode
        #f
        block56['F13'] = bebecode
        block56['F14'] = f'''=code128("{com7productcode}")'''
        block56['F15'] = com7productcode 
        #h
        block56['H13'] = bebecode
        block56['H14'] = f'''=code128("{com7productcode}")'''
        block56['H15'] = com7productcode 

        #b
        block56['B18'] = bebecode
        block56['B19'] = f'''=code128("{com7productcode}")'''
        block56['B20'] = com7productcode
        #d
        block56['D18'] = bebecode
        block56['D19'] = f'''=code128("{com7productcode}")'''
        block56['D20'] = com7productcode
        #f
        block56['F18'] = bebecode
        block56['F19'] = f'''=code128("{com7productcode}")'''
        block56['F20'] = com7productcode 
        #h
        block56['H18'] = bebecode
        block56['H19'] = f'''=code128("{com7productcode}")'''
        block56['H20'] = com7productcode 

        #b
        block56['B23'] = bebecode
        block56['B24'] = f'''=code128("{com7productcode}")'''
        block56['B25'] = com7productcode
        #d
        block56['D23'] = bebecode
        block56['D24'] = f'''=code128("{com7productcode}")'''
        block56['D25'] = com7productcode
        #f
        block56['F23'] = bebecode
        block56['F24'] = f'''=code128("{com7productcode}")'''
        block56['F25'] = com7productcode 
        #h
        block56['H23'] = bebecode
        block56['H24'] = f'''=code128("{com7productcode}")'''
        block56['H25'] = com7productcode 

        #b
        block56['B28'] = bebecode
        block56['B29'] = f'''=code128("{com7productcode}")'''
        block56['B30'] = com7productcode
        #d
        block56['D28'] = bebecode
        block56['D29'] = f'''=code128("{com7productcode}")'''
        block56['D30'] = com7productcode
        #f
        block56['F28'] = bebecode
        block56['F29'] = f'''=code128("{com7productcode}")'''
        block56['F30'] = com7productcode 
        #h
        block56['H28'] = bebecode
        block56['H29'] = f'''=code128("{com7productcode}")'''
        block56['H30'] = com7productcode 

        #b
        block56['B33'] = bebecode
        block56['B34'] = f'''=code128("{com7productcode}")'''
        block56['B35'] = com7productcode
        #d
        block56['D33'] = bebecode
        block56['D34'] = f'''=code128("{com7productcode}")'''
        block56['D35'] = com7productcode
        #f
        block56['F33'] = bebecode
        block56['F34'] = f'''=code128("{com7productcode}")'''
        block56['F35'] = com7productcode 
        #h
        block56['H33'] = bebecode
        block56['H34'] = f'''=code128("{com7productcode}")'''
        block56['H35'] = com7productcode 

        #b
        block56['B38'] = bebecode
        block56['B39'] = f'''=code128("{com7productcode}")'''
        block56['B40'] = com7productcode
        #d
        block56['D38'] = bebecode
        block56['D39'] = f'''=code128("{com7productcode}")'''
        block56['D40'] = com7productcode
        #f
        block56['F38'] = bebecode
        block56['F39'] = f'''=code128("{com7productcode}")'''
        block56['F40'] = com7productcode 
        #h
        block56['H38'] = bebecode
        block56['H39'] = f'''=code128("{com7productcode}")'''
        block56['H40'] = com7productcode 

        #b
        block56['B43'] = bebecode
        block56['B44'] = f'''=code128("{com7productcode}")'''
        block56['B45'] = com7productcode
        #d
        block56['D43'] = bebecode
        block56['D44'] = f'''=code128("{com7productcode}")'''
        block56['D45'] = com7productcode
        #f
        block56['F43'] = bebecode
        block56['F44'] = f'''=code128("{com7productcode}")'''
        block56['F45'] = com7productcode 
        #h
        block56['H43'] = bebecode
        block56['H44'] = f'''=code128("{com7productcode}")'''
        block56['H45'] = com7productcode 

        #b
        block56['B48'] = bebecode
        block56['B49'] = f'''=code128("{com7productcode}")'''
        block56['B50'] = com7productcode
        #d
        block56['D48'] = bebecode
        block56['D49'] = f'''=code128("{com7productcode}")'''
        block56['D50'] = com7productcode
        #f
        block56['F48'] = bebecode
        block56['F49'] = f'''=code128("{com7productcode}")'''
        block56['F50'] = com7productcode 
        #h
        block56['H48'] = bebecode
        block56['H49'] = f'''=code128("{com7productcode}")'''
        block56['H50'] = com7productcode 

        #b
        block56['B53'] = bebecode
        block56['B54'] = f'''=code128("{com7productcode}")'''
        block56['B55'] = com7productcode
        #d
        block56['D53'] = bebecode
        block56['D54'] = f'''=code128("{com7productcode}")'''
        block56['D55'] = com7productcode
        #f
        block56['F53'] = bebecode
        block56['F54'] = f'''=code128("{com7productcode}")'''
        block56['F55'] = com7productcode 
        #h
        block56['H53'] = bebecode
        block56['H54'] = f'''=code128("{com7productcode}")'''
        block56['H55'] = com7productcode 

        #b
        block56['B58'] = bebecode
        block56['B59'] = f'''=code128("{com7productcode}")'''
        block56['B60'] = com7productcode
        #d
        block56['D58'] = bebecode
        block56['D59'] = f'''=code128("{com7productcode}")'''
        block56['D60'] = com7productcode
        #f
        block56['F58'] = bebecode
        block56['F59'] = f'''=code128("{com7productcode}")'''
        block56['F60'] = com7productcode 
        #h
        block56['H58'] = bebecode
        block56['H59'] = f'''=code128("{com7productcode}")'''
        block56['H60'] = com7productcode 

        #b
        block56['B63'] = bebecode
        block56['B64'] = f'''=code128("{com7productcode}")'''
        block56['B65'] = com7productcode
        #d
        block56['D63'] = bebecode
        block56['D64'] = f'''=code128("{com7productcode}")'''
        block56['D65'] = com7productcode
        #f
        block56['F63'] = bebecode
        block56['F64'] = f'''=code128("{com7productcode}")'''
        block56['F65'] = com7productcode 
        #h
        block56['H63'] = bebecode
        block56['H64'] = f'''=code128("{com7productcode}")'''
        block56['H65'] = com7productcode 

        #b
        block56['B68'] = bebecode
        block56['B69'] = f'''=code128("{com7productcode}")'''
        block56['B70'] = com7productcode
        #d
        block56['D68'] = bebecode
        block56['D69'] = f'''=code128("{com7productcode}")'''
        block56['D70'] = com7productcode
        #f
        block56['F68'] = bebecode
        block56['F69'] = f'''=code128("{com7productcode}")'''
        block56['F70'] = com7productcode 
        #h
        block56['H68'] = bebecode
        block56['H69'] = f'''=code128("{com7productcode}")'''
        block56['H70'] = com7productcode 
        block56_wb.save(f"{bebecode}.xlsm")


def movefiletofolders():
    def grouper(S, n): #https://stackoverflow.com/questions/12559055/for-every-x-number-of-files-create-new-directory-and-move-files-using-python
        iterator = iter(S)
        while True:
            item = list(itertools.islice(iterator, n))
            print(item)
            if len(item) == 0:
                break
            yield item
    os.chdir(r"D:\Workstuff\my-work-python-script\project_bebephone\Block 56")
    fnames = ns.natsorted(glob.glob('*.xlsm'))
    for i, fnames in enumerate(grouper(fnames, 53)):
        dirname = f"{i + 1}"
        try:
            os.mkdir(dirname)
        except FileExistsError:
            pass
        for fname in fnames:
            shutil.move(fname, dirname)
            with open(f"filelist {i + 1}.txt", mode="a") as textfile:
                textfile.writelines(f"{fname}\n")
        shutil.move(f"filelist {i + 1}.txt", dirname)
        

if __name__ in "__main__"  :
    main()
    movefiletofolders()