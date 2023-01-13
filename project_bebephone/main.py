import os, glob, shutil
import openpyxl
from tkinter import Tk, filedialog
import itertools
import natsort as ns

asset_directory = os.path.join(r"C:\Users\jambo\OneDrive\เดสก์ท็อป\Workspaces\my-work-python-script\project_bebephone", r"Asset")

root = Tk()
#get data
get_wb = filedialog.askopenfilename(title="เลือกไฟล์ข้อมูล Excel", filetypes=([("*Excel Files", "*.xlsx"), ("All Files", "*.*")]))
data_wb = openpyxl.load_workbook(get_wb, data_only=True)
data_sheet = data_wb.sheetnames
get_sheet = data_wb[data_sheet[0]]

try:
    os.mkdir("PRINT")
except FileExistsError:
    pass

class generate_block:
    #get data
    get_wb = filedialog.askopenfilename(title="เลือกไฟล์ข้อมูล Excel", filetypes=([("*Excel Files", "*.xlsx"), ("All Files", "*.*")]))
    data_wb = openpyxl.load_workbook(get_wb, data_only=True)
    data_sheet = data_wb.sheetnames
    get_sheet = data_wb[data_sheet[1]]
    def block90():
        #get asset
        PRINT_dir = r"D:\Workstuff\my-work-python-script\project_bebephone\PRINT"
        asset_dir_90 = os.path.join(asset_directory, r"block90template.xlsm")
        block90_wb = openpyxl.load_workbook(asset_dir_90, data_only=True, keep_vba=True)
        block90_ws = block90_wb.sheetnames
        block90 = block90_wb[block90_ws[0]]

        #setting
        bebecode_startrow = 2
        com7barcode_startrow = 3
        com7barcodetext_startrow = 4
        row_gap = 4
        bebecode_endrow = 74
        com7barcode_endrow = 75
        com7barcodetext_endrow = 76
        start_col = 2
        col_gap = 3
        #end setting

        for row in range(3,get_sheet.max_row + 1):
            bebecode = get_sheet.cell(row=row, column=2).value
            com7barcode = get_sheet.cell(row=row, column=3).value
            #calculate column needed
            block90_need = get_sheet.cell(row=row, column=5).value

            #row first
            #bebecode
            for beberow in range(bebecode_startrow, bebecode_endrow, row_gap):
                for col in range(start_col, block90_need, col_gap):
                    block90.cell(row=beberow, column=col).value = bebecode

            for com7barcoderow in range(com7barcode_startrow, com7barcode_endrow, row_gap):
                for col in range(start_col, block90_need, col_gap):
                    block90.cell(row=com7barcoderow, column=col).value = f'''=code128("{com7barcode}")'''

            for com7barcodetextrow in range(com7barcodetext_startrow, com7barcodetext_endrow, row_gap):
                for col in range(start_col, block90_need, col_gap):
                    block90.cell(row=com7barcodetextrow, column=col).value = com7barcode
            
            block90_wb.save(f"{bebecode}.xlsm")
            break


    def block56():
        #get asset
        PRINT_dir = r"D:\Workstuff\my-work-python-script\project_bebephone\PRINT"
        asset_dir_56 = os.path.join(asset_directory, r"block56template.xlsm")
        block56_wb = openpyxl.load_workbook(asset_dir_56, data_only=True, keep_vba=True)
        block56_ws = block56_wb.sheetnames
        block56 = block56_wb[block56_ws[0]]
        #setting loop
        bebecode_startrow = 4
        com7barcode_startrow = 5
        com7barcodetext_startrow = 6
        for row in range(3,get_sheet.max_row + 1):
            bebecode = get_sheet.cell(row=row, column=2).value
            com7barcode = get_sheet.cell(row=row, column=3).value
            block56_need = get_sheet.cell(row=row, column=7).value

    def block117():
        #get asset
        PRINT_dir = r"D:\Workstuff\my-work-python-script\project_bebephone\PRINT"
        asset_dir_117 = os.path.join(asset_directory, r"block117template.xlsm")
        block117_wb = openpyxl.load_workbook(asset_dir_117, data_only=True, keep_vba=True)
        block117_ws = block117_wb.sheetnames
        block117 = block117_wb[block117_ws[0]]
        #setting loop
        bebecode_startrow = 3
        com7barcode_startrow = 4
        com7barcodetext_startrow = 5
        for row in range(3,get_sheet.max_row + 1):
            bebecode = get_sheet.cell(row=row, column=2).value
            com7barcode = get_sheet.cell(row=row, column=3).value
            block117_need = get_sheet.cell(row=row, column=6).value



if __name__ in "__main__":
    generate_block.block90()
