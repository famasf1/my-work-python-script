from logging import logProcesses
from operator import le
from re import L
from turtle import right
from selenium import webdriver
from time import sleep
import pyperclip
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from subprocess import CREATE_NO_WINDOW
from selenium.webdriver.chrome.options import Options
import pandastable
import pandas as pd
import auto_checktrack
import openpyxl
from selenium.common.exceptions import NoSuchElementException

#### I will separate the whole thing into 2 section. Prepatation, Interface and Logic.
## - Preparation will be code from test_idea_tracker
## - Interface part is for production
## - Logic is everything below Preparation

###############################
######## : Logic : ############
###############################

##Tried to practice how to use class in Python by create a nested function lmao
class launch_browser():
     ##################################### ##################################### #####################################
    '''
    This class is for launching selected browser. It has 2 methods.
    1. launch_browser_chrome will launch Google Chrome
    2. launch_browser_firefox will launch Firefox
    '''
     ##################################### ##################################### #####################################
    def launch_browser_chrome(self):
        chrome_service = ChromeService(ChromeDriverManager().install())
        chrome_service.creationflags = CREATE_NO_WINDOW
        chrome_option = Options()
        chrome_option.add_experimental_option("detach",True)
        chrome_option.add_experimental_option("useAutomationExtension", False)
        chrome_option.add_experimental_option("excludeSwitches",["enable-automation"])
        driver = webdriver.Chrome(options=chrome_option,service=chrome_service)
        driver.maximize_window()
        return driver
    def launch_browser_firefox(self):
        firefox_service = FirefoxService(GeckoDriverManager().install())
        driver = webdriver.Firefox(service=firefox_service)
        driver.maximize_window()
        return driver

def gather_Data():
    ##################################### ##################################### #####################################
    '''
    This function will gather data from the clipboard, Flitered everything out then pass along all the values to calculate in 'tracker' function
    '''
    ##################################### ##################################### #####################################    

    global result

    df_data = pd.DataFrame(pd.read_clipboard(sep='\t'))

    header = list(df_data.columns)
    donotdrop = ['Doc Date','Branch (Name)','Branch To (Name)','Booking ID']
    new_data = df_data[donotdrop]
    new_data['Booking ID'] = new_data['Booking ID'].str.replace('Booking-DHL ID : ','')
    new = new_data.dropna()['Booking ID']   #.to_csv('retrieve_tracking.txt', header=None, index=None)
    new = pd.DataFrame(new)
    new['Booking ID'] = new.apply(lambda x: x['Booking ID'].split(' , '), axis=1)
    result = new.explode('Booking ID').to_clipboard(index=False)
    tracker()


def tracker():
    ##################################### ##################################### #####################################
    '''
    Launch Browser, find element in the web and search it. Then call 'retrievetrackingcode' to fetch all the data into pandastable
    '''
    ##################################### ##################################### #####################################
    global driver
    driver = launch_browser().launch_browser_chrome()##change firefox to chrome to use chrome instead
    driver.get("https://ecommerceportal.dhl.com/track/")


    def retrievetrackingcode():

        ##################################### ##################################### #####################################
        '''
        This function will fetch all value in txt files, then loop it before sending the result into tracker()
        '''
        ##################################### ##################################### #####################################

        global ws
        ws = openpyxl.Workbook()
        ws1 = ws.create_sheet('Result',0)
        ws1.cell(row=1, column=1).value = "PHYID"
        ws1.cell(row=1, column=2).value = "Tracking Number"
        ws1.cell(row=1, column=3).value = "สถานะล่าสุด"
        ws1.cell(row=1, column=4).value = "รายละเอียด"
        ws1.cell(row=1, column=5).value = "น้ำหนัก"
        ws1.cell(row=1, column=6).value = "วันที่"

        with open(r'D:\Workstuff\my-work-python-script\Auto_Tracker (dev)\retrieve_tracking.txt', 'r+') as text: #fetch all value
            global file_len
            file_len = len(text.readlines())
            time_cyc = int(file_len / 50) + 2
            for loop_times in range(1,time_cyc):
                i = 0
                if loop_times > 1: #as soon as loop_times is more than first time.
                    text.seek(i + 900 * (loop_times - 1)) #go to character 900
                    print(text.seek(i + 900 * (loop_times - 1)))
                    driver.find_element(By.XPATH,"//textarea[@id='trackItNowForm:trackItNowSearchBox']").clear()
                else:
                    text.seek(i)
                for index, value in enumerate(text): #iterate through all of them first
                    index_add_one = index + 1
                    driver.find_element(By.XPATH,"//textarea[@id='trackItNowForm:trackItNowSearchBox']").send_keys(value)
                    if index % 49 == 0 and index != 0:
                        driver.find_element(By.ID,'trackItNowForm:searchSkuBtn').click()
                        break
                        # ^ = start-with
                        # * = contains
                        #Is this regex?
                    elif index_add_one % file_len == 0 and index != 0:
                        driver.find_element(By.ID,'trackItNowForm:searchSkuBtn').click()
                        break
                    else:
                        continue
                WebDriverWait(driver,10000).until(EC.visibility_of_element_located((By.XPATH, "//label[contains(@id,'trackItNowForm') and(contains(@class,'TrackingNumber'))]"))).text
                
                #### In each cycle, get data and then for every 50 rows refresh DHL Pages.

                def get_data_to_Excel():
                    for i in range(0,49): ##test 
                        try:
                            element = WebDriverWait(driver,10000).until(EC.element_to_be_clickable((By.CSS_SELECTOR,f"[id^='trackItNowForm'][id*=':{i-1}:'][class*='ui-commandlink ui-widget'][onclick*='PrimeFaces']")))
                            driver.execute_script("arguments[0].click();", element)
                            sleep(1.8)
                            refid = driver.find_element(By.XPATH, "//h3[contains(@class, 'track-number-heading')]").text
                            status_track = driver.find_element(By.CSS_SELECTOR,f"[id^='trackItNowForm'][id*=':0:'][class*='TrackStatus']").text
                            track_details = driver.find_element(By.CSS_SELECTOR,"[class*='ui-outputlabel ui-widget ShipmentDetails']").text
                            timeanddate = driver.find_element(By.CSS_SELECTOR,f"[id^='trackItNowForm'][id*=':0:'][id*='dateandtime'][class*='TrackTimeAndDate']").text
                            receiver = driver.find_element(By.CSS_SELECTOR, "[class*='ui-outputlabel ui-widget d-lg-none']").text
                            if file_len < 50:
                                end = file_len
                            else:
                                end = 50
                            if loop_times > 1:
                                start = 50 * (loop_times - 1)
                                end = 50 * loop_times
                            else:
                                start = 0
                                end = 50
                            ws1.cell(row=start+2, column=1).value = refid
                            ws1.cell(row=start+2, column=2).value = status_track
                            ws1.cell(row=start+2, column=3).value = receiver
                            ws1.cell(row=start+2, column=4).value = track_details
                            ws1.cell(row=start+2, column=5).value = timeanddate
                            start += 1
                            print([ws1.cell(row=i+2, column=1).value,ws1.cell(row=i+2, column=2).value,ws1.cell(row=i+2, column=3).value,ws1.cell(row=i+2, column=4).value,ws1.cell(row=i+2, column=5).value])
                            ### Go to excel row
                            ### for every cycle, take 50 and times with loop cycle amouth.
                            ### leaving i undisturbed                                  
                            sleep(.9)
                            driver.find_element(By.ID, "trackItNowForm:backbutton").click()
                            sleep(.4)
                            #quit = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID,"trackItNowForm:backbutton")))
                            #driver.execute("arguments[0].click();",quit)
                            if i > 49:
                                i = 0
                                driver.refresh()
                                driver.find_element(By.XPATH,"//textarea[@id='trackItNowForm:trackItNowSearchBox']").clear()
                                break
                            elif i % file_len == 0 and i != 0:
                                print('break')
                                break
                            else:
                                i += 1
                                print(i)
                                continue
                        except NoSuchElementException as e:
                            print(e)
                get_data_to_Excel()
        ws.save("get_data_dhl.xlsx")
    retrievetrackingcode()
    

    


def test_room():

    ##################################### ##################################### #####################################
    '''
    This function is purely for testing. Delete after production-ready
    '''
    ##################################### ##################################### #####################################
    driver = launch_browser().launch_browser_chrome()##change firefox to chrome to use chrome instead
    driver.get("https://ecommerceportal.dhl.com/track/")
    with open(r'D:\Workstuff\my-work-python-script\Auto_Tracker (dev)\retrieve_tracking.txt', 'r+') as text: #fetch all value
        file_len = len(text.readlines())
        time_cyc = int(file_len / 50) + 2
        for loop_times in range(1,time_cyc):
            i = 0
            if loop_times > 1: #as soon as loop_times is more than first time.
                text.seek(i + 900 * (loop_times - 1)) #go to character 900
                driver.find_element(By.XPATH,"//textarea[@id='trackItNowForm:trackItNowSearchBox']").clear()
            else:
                text.seek(i)
            for index, value in enumerate(text): #iterate through all of them first
                index_add_one = index + 1
                driver.find_element(By.XPATH,"//textarea[@id='trackItNowForm:trackItNowSearchBox']").send_keys(value)
                if index_add_one % 50 == 0 and index != 0:
                    print('hit 50')
                    break
                    # ^ = start-with
                    # * = contains
                    #Is this regex?
                elif index_add_one % file_len == 0 and index != 0:
                    print('EOL')
                    break
                else:
                    continue    
            sleep(2)  
            end = 50
            if loop_times > 1:
                start = 50 * (loop_times - 1)
                end = 50 * loop_times
                print(f"{start} || {end}")
            else:
                start = 0
                end = 50
            for r in range(start,end): ##excel row
                print(f'{r} / {end}' )
                if r == 50:
                    driver.refresh()
                    driver.find_element(By.XPATH,"//textarea[@id='trackItNowForm:trackItNowSearchBox']").clear()
                    sleep(.7)

if __name__ in "__main__":
    ##tracker is the main function
    #gather_Data()
    tracker()
    #retrievetrackingcode()

    #test_room()